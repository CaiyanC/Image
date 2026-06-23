# -*- coding: utf-8 -*-
import openpyxl
from collections import defaultdict

path = r"C:\Users\wnt\Desktop\产品库元数据.xlsx"
wb = openpyxl.load_workbook(path, data_only=True)
ws = wb["产品库元数据"]

sku_rows = defaultdict(list)
for row_idx in range(6, ws.max_row + 1):
    sku = ws.cell(row=row_idx, column=2).value
    if sku and str(sku).strip():
        sku_val = str(sku).strip()
        name = ws.cell(row=row_idx, column=4).value
        sku_rows[sku_val].append((row_idx, name))

dupes = {k: v for k, v in sku_rows.items() if len(v) > 1}
print(f"Duplicate SKUs ({len(dupes)}):")
for sku, rows in dupes.items():
    print(f"\n  {sku}:")
    for r in rows:
        print(f"    Row {r[0]}: {r[1]}")
