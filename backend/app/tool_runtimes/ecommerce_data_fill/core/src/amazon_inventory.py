"""Independent Amazon inventory workbook fill pipeline.

The weekly update owns commercial/factory fields.  The FBA export owns
available stock and ageing buckets.  Formula columns in the target workbook
remain formulas and are never replaced with calculated values.
"""
from __future__ import annotations

import re
from collections import defaultdict
from copy import copy
from pathlib import Path

from openpyxl import load_workbook

from src.excel_utils import SheetContext, build_sheet_context, build_column_map
from src.models import Issue
from src.source_builders import parse_number


WEEKLY_FIELDS = {
    "产品名称": 1, "SKU": 2, "成本": 3, "单价": 4, "15天销量": 5,
    "30天销量": 6, "工厂库存": 8, "工厂采购在途库存数": 9, "FBA仓在途库存": 10,
}
FBA_FIELDS = {
    "FBA可用库存": 4, "30天内库龄": 5, "31-60天库龄": 6,
    "61-90天库龄": 7, "91-180天库龄": 8, "181-270天库龄": 9,
    "271-330天库龄": 10, "331-365天库龄": 11, "大于365天库龄": 12,
}
TARGET_FIELDS = {
    "SKU": 1, "品名": 2, "成本": 3, "单价": 4, "15天销量": 5,
    "30天销量": 6, "聚水潭库存": 7, "聚水潭采购在途库存数": 8,
    "FBA可用库存": 9, "FBA仓在途库存": 10,
    **{field: column + 9 for field, column in FBA_FIELDS.items() if field != "FBA可用库存"},
}
WEEKLY_HEADER_ALIASES = {
    "产品名称": ("品名",),
    "工厂采购在途库存数": ("采购在途", "工厂采购在途库存"),
    "FBA仓在途库存": ("FBA在途", "FBA在途库存"),
}


def _site(value: str) -> str:
    value = str(value or "").replace(" ", "")
    for token, site in (("北美", "north_america"), ("日本", "japan"), ("英国", "uk"), ("德国", "germany")):
        if token in value:
            return site
    return ""


def _sku(value: object) -> str:
    return str(value or "").strip().upper()


def _base_sku(value: str) -> str:
    """Allow only an unambiguous terminal version suffix bridge."""
    return re.sub(r"-(?:0?1)$", "", value)


def _site_sheets(workbook) -> dict[str, object]:
    result = {}
    for ws in workbook.worksheets:
        site = _site(ws.title)
        if site:
            result[site] = ws
    return result


def _weekly_context(ws) -> SheetContext:
    """Find the weekly header by canonical names with documented aliases."""
    for row_number in range(1, min(ws.max_row, 10) + 1):
        available = build_column_map(ws, row_number)
        normalized_available = {str(name).strip().casefold(): column for name, column in available.items()}
        columns: dict[str, int] = {}
        for field in WEEKLY_FIELDS:
            candidates = (field, *WEEKLY_HEADER_ALIASES.get(field, ()))
            column = next((normalized_available.get(candidate.casefold()) for candidate in candidates if candidate.casefold() in normalized_available), None)
            if column is None:
                break
            columns[field] = column
        else:
            return SheetContext(header_row=row_number, columns=columns, data_start_row=row_number + 1)
    raise ValueError(f"Could not find weekly header row in sheet {ws.title}")


def _weekly_index(ws) -> dict[str, dict]:
    context = _weekly_context(ws)
    rows = {}
    for row_number in range(context.data_start_row, ws.max_row + 1):
        sku = _sku(ws.cell(row_number, context.columns["SKU"]).value)
        if not sku:
            continue
        rows[sku] = {field: ws.cell(row_number, context.columns[field]).value for field in WEEKLY_FIELDS}
    return rows


def _fba_index(ws) -> dict[str, dict]:
    """Aggregate duplicate FBA records instead of choosing an arbitrary row."""
    context = build_sheet_context(ws, ["SKU", *FBA_FIELDS])
    totals: dict[str, dict] = defaultdict(lambda: {field: 0.0 for field in FBA_FIELDS})
    seen: set[str] = set()
    for row_number in range(context.data_start_row, ws.max_row + 1):
        sku = _sku(ws.cell(row_number, context.columns["SKU"]).value)
        if not sku:
            continue
        seen.add(sku)
        for field in FBA_FIELDS:
            value = parse_number(ws.cell(row_number, context.columns[field]).value)
            if value is not None:
                totals[sku][field] += value
    return {sku: values for sku, values in totals.items() if sku in seen}


def _target_context(ws) -> SheetContext:
    context = build_sheet_context(ws, [*TARGET_FIELDS, "可用的总库存", "可销售天数"])
    for name, column in list(context.columns.items()):
        if name.startswith("是否断货"):
            context.columns["是否断货"] = column
            break
    if "是否断货" not in context.columns:
        raise ValueError(f"Could not find 是否断货 column in sheet {ws.title}")
    return context


def _resolve_weekly(target_sku: str, weekly: dict[str, dict]) -> tuple[dict | None, str]:
    if target_sku in weekly:
        return weekly[target_sku], target_sku
    base = _base_sku(target_sku)
    candidates = [sku for sku in weekly if _base_sku(sku) == base]
    if len(candidates) == 1:
        return weekly[candidates[0]], candidates[0]
    return None, ""


def _snapshot_row(ws, row_number: int) -> list[dict]:
    return [
        {
            "value": cell.value,
            "style": copy(cell._style),
            "number_format": cell.number_format,
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "protection": copy(cell.protection),
        }
        for cell in ws[row_number]
    ]


def _restore_row(ws, row_number: int, snapshot: list[dict]) -> None:
    for column, source in enumerate(snapshot, start=1):
        cell = ws.cell(row_number, column)
        cell.value = source["value"]
        cell._style = copy(source["style"])
        cell.number_format = source["number_format"]
        cell.font = copy(source["font"])
        cell.fill = copy(source["fill"])
        cell.border = copy(source["border"])
        cell.alignment = copy(source["alignment"])
        cell.protection = copy(source["protection"])


def _clear_business_row(ws, row_number: int, columns: dict[str, int]) -> None:
    for column in range(1, max(columns.values()) + 1):
        ws.cell(row_number, column).value = None


def _restore_target_formulas(ws, row_number: int, columns: dict[str, int], threshold_days: int | None) -> None:
    from openpyxl.utils import get_column_letter

    inventory = get_column_letter(columns["聚水潭库存"])
    fba_available = get_column_letter(columns["FBA可用库存"])
    fba_transit = get_column_letter(columns["FBA仓在途库存"])
    total = get_column_letter(columns["可用的总库存"])
    sales_30 = get_column_letter(columns["30天销量"])
    sales_days = get_column_letter(columns["可销售天数"])
    ws.cell(row_number, columns["可用的总库存"]).value = f"={inventory}{row_number}+{fba_available}{row_number}+{fba_transit}{row_number}"
    ws.cell(row_number, columns["可销售天数"]).value = f"=({total}{row_number}/{sales_30}{row_number})*30"
    ws.cell(row_number, columns["是否断货"]).value = f"={sales_days}{row_number}-{threshold_days}" if threshold_days in {90, 120} else None


def _threshold_days(value: object) -> int | None:
    match = re.fullmatch(r"=L\d+-(90|120)", str(value or ""))
    return int(match.group(1)) if match else None


def fill_amazon_inventory(target_path: Path, weekly_path: Path, fba_path: Path, issues: list[Issue] | None = None) -> dict[str, int]:
    """Fill a copy of the newest Amazon inventory workbook in place."""
    issues = issues if issues is not None else []
    target = load_workbook(target_path)
    weekly_book = load_workbook(weekly_path, read_only=True, data_only=True)
    fba_book = load_workbook(fba_path, read_only=True, data_only=True)
    counts = {
        "written_rows": 0, "missing_weekly": 0, "missing_fba": 0,
        "source_only_weekly": 0, "added_weekly_rows": 0, "cleared_target_rows": 0,
    }
    try:
        weekly_sheets = _site_sheets(weekly_book)
        fba_sheets = _site_sheets(fba_book)
        for target_ws in target.worksheets:
            site = _site(target_ws.title)
            if not site:
                continue
            target_context = _target_context(target_ws)
            target_columns = target_context.columns
            if site not in weekly_sheets:
                issues.append(Issue("WARNING", "amazon_inventory", "amazon_inventory_weekly", "本周更新表缺少对应站点，未同步该站点 SKU 清单", weekly_path.name, target_ws.title, "", "", "站点", "亚马逊库存", "SKU", "补充对应站点 Sheet 后重跑。"))
                continue
            weekly = _weekly_index(weekly_sheets[site])
            fba = _fba_index(fba_sheets[site]) if site in fba_sheets else {}
            existing_rows = [row for row in range(target_context.data_start_row, target_ws.max_row + 1) if _sku(target_ws.cell(row, target_columns["SKU"]).value)]
            # A clean test template deliberately has no SKU values. Formula
            # rows still provide the required styles and 90/120-day threshold.
            template_rows = existing_rows or [
                row for row in range(target_context.data_start_row, target_ws.max_row + 1)
                if any(target_ws.cell(row, target_columns[field]).value not in (None, "") for field in ("可用的总库存", "可销售天数", "是否断货"))
            ]
            snapshots = {row: _snapshot_row(target_ws, row) for row in template_rows}
            existing_by_sku = {_sku(target_ws.cell(row, target_columns["SKU"]).value): row for row in existing_rows}
            resolved_existing: dict[str, int] = {}
            threshold_by_weekly_sku: dict[str, int] = {}
            for target_sku, row in existing_by_sku.items():
                _, weekly_sku = _resolve_weekly(target_sku, weekly)
                if weekly_sku and weekly_sku not in resolved_existing:
                    resolved_existing[weekly_sku] = row
                    threshold = _threshold_days(target_ws.cell(row, target_columns["是否断货"]).value)
                    if threshold is not None:
                        threshold_by_weekly_sku[weekly_sku] = threshold
            default_snapshot = snapshots[template_rows[-1]] if template_rows else None
            for offset, (weekly_sku, weekly_row) in enumerate(weekly.items()):
                row_number = target_context.data_start_row + offset
                matched_source_row = (
                    existing_by_sku.get(weekly_sku)
                    or resolved_existing.get(weekly_sku)
                )
                source_row = matched_source_row or (row_number if row_number in snapshots else None)
                if source_row is not None:
                    _restore_row(target_ws, row_number, snapshots[source_row])
                    if matched_source_row is None:
                        counts["added_weekly_rows"] += 1
                elif default_snapshot is not None:
                    _restore_row(target_ws, row_number, default_snapshot)
                    counts["added_weekly_rows"] += 1
                else:
                    target_ws.insert_rows(row_number)
                    counts["added_weekly_rows"] += 1
                target_ws.cell(row_number, target_columns["SKU"]).value = weekly_sku
                for source_field, target_field in (("产品名称", "品名"), ("成本", "成本"), ("单价", "单价"), ("15天销量", "15天销量"), ("30天销量", "30天销量"), ("工厂库存", "聚水潭库存"), ("工厂采购在途库存数", "聚水潭采购在途库存数"), ("FBA仓在途库存", "FBA仓在途库存")):
                    target_ws.cell(row_number, target_columns[target_field]).value = weekly_row[source_field]
                # FBA exports occasionally use a unique terminal -01/-1
                # variant while the weekly roster uses the base SKU. Reuse
                # the same conservative, same-site bridge as weekly matching.
                fba_row, _ = _resolve_weekly(weekly_sku, fba)
                if fba_row is None:
                    counts["missing_fba"] += 1
                    for field in FBA_FIELDS:
                        target_ws.cell(row_number, target_columns[field]).value = None
                    issues.append(Issue("WARNING", "amazon_inventory", "fba_inventory", "本周 SKU 未在 FBA 仓库明细找到，FBA 字段已留空", fba_path.name, target_ws.title, str(row_number), weekly_sku, "SKU", "亚马逊库存", "FBA字段", "确认该 SKU 是否无 FBA 库存或遗漏导出。"))
                else:
                    for field in FBA_FIELDS:
                        target_ws.cell(row_number, target_columns[field]).value = fba_row[field]
                threshold_days = threshold_by_weekly_sku.get(weekly_sku)
                # K/L are deterministic. M needs an explicit business class:
                # a new SKU may not inherit another SKU's 90/120-day threshold.
                _restore_target_formulas(target_ws, row_number, target_columns, threshold_days)
                if threshold_days is None:
                    issues.append(Issue("WARNING", "amazon_inventory", "amazon_inventory_weekly", "SKU 缺少主推款/长尾款阈值，未伪造断货公式", weekly_path.name, weekly_sheets[site].title, str(row_number), weekly_sku, "是否断货", "亚马逊库存", "M列公式", "请业务确认该 SKU 为主推款（120 天）或长尾款（90 天）后填写。"))
                counts["written_rows"] += 1
            for row_number in template_rows:
                if row_number >= target_context.data_start_row + len(weekly):
                    _clear_business_row(target_ws, row_number, target_columns)
                    counts["cleared_target_rows"] += 1
        target.calculation.fullCalcOnLoad = True
        target.calculation.forceFullCalc = True
        target.save(target_path)
    finally:
        target.close()
        weekly_book.close()
        fba_book.close()
    return counts
