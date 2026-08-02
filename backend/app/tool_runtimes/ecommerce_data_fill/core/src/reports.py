from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook

from src.models import DetectionResult, Issue, SheetWriteResult
from src.runtime import FillConfig
from src.validators import role_display_name


def write_recognition_report(output_path: Path, rows: list[dict], fill_config: FillConfig | None = None) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "文件识别报告"
    ws.append(["文件角色", "是否必需", "识别状态", "实际文件名", "识别命中的Sheet", "实际处理的Sheet", "跳过的Sheet", "命中的关键字段", "缺失字段", "影响范围", "备注"])
    for row in rows:
        ws.append(
            [
                row.get("文件角色", ""),
                row.get("是否必需", ""),
                row.get("识别状态", ""),
                row.get("实际文件名", ""),
                row.get("识别命中的Sheet", ""),
                row.get("实际处理的Sheet", ""),
                row.get("跳过的Sheet", ""),
                row.get("命中的关键字段", ""),
                row.get("缺失字段", ""),
                row.get("影响范围", ""),
                row.get("备注", ""),
            ]
        )
    if fill_config:
        ws2 = wb.create_sheet("本次配置")
        ws2.append(["配置项", "实际值", "来源"])
        for field_name, source in fill_config.source_map.items():
            ws2.append([field_name, getattr(fill_config, field_name), source])
    wb.save(output_path)


def write_issue_report(output_path: Path, issues: list[Issue]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "异常清单"
    ws.append(["异常级别", "模块", "文件角色", "文件名", "Sheet", "行号", "SKU", "字段", "异常说明", "影响目标表", "影响目标字段", "建议处理方式"])
    for issue in issues:
        ws.append([issue.level, issue.module, issue.file_role, issue.file_name, issue.sheet, issue.row_number, issue.sku, issue.field, issue.message, issue.target_table, issue.target_field, issue.suggestion])
    wb.save(output_path)


def write_review_report(output_path: Path, rows: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "建议人工复核"
    ws.append(["结果表", "结果Sheet", "结果定位", "结果字段", "SKU", "商品名称", "来源文件", "来源Sheet", "来源行号", "来源字段", "来源定位", "异常说明", "建议核对动作"])
    for row in rows:
        ws.append(
            [
                row.get("结果表", ""),
                row.get("结果Sheet", row.get("Sheet", "")),
                row.get("结果定位", row.get("定位信息", "")),
                row.get("结果字段", row.get("异常字段", "")),
                row.get("SKU", ""),
                row.get("商品名称", ""),
                row.get("来源文件", ""),
                row.get("来源Sheet", ""),
                row.get("来源行号", ""),
                row.get("来源字段", ""),
                row.get("来源定位", ""),
                row.get("异常说明", ""),
                row.get("建议核对动作", ""),
            ]
        )
    wb.save(output_path)


def write_audit_report(output_path: Path, sheet_rows: dict[str, list[dict]]) -> None:
    wb = Workbook()
    required_sheets = [
        "W27_3_SKU明细_原有与重填差异",
        "W27_7_库存补货_原有与重填差异",
        "W27_近30天销量来源核对",
        "W27_异常SKU源表对账",
        "开普乐_源销售明细_来源核对",
        "开普乐_源库存快照_来源核对",
    ]
    headers = {
        "W27_3_SKU明细_原有与重填差异": ["目标Sheet", "平台", "店铺名称", "SKU编码", "商品名称", "原目标表是否已有", "本次是否写入", "差异类型", "来源文件", "来源Sheet", "来源字段", "关键值", "说明"],
        "W27_7_库存补货_原有与重填差异": ["目标Sheet", "平台", "店铺名称", "SKU编码", "商品名称", "原目标表是否已有", "本次是否写入", "差异类型", "来源文件", "来源Sheet", "来源字段", "关键值", "说明"],
        "W27_近30天销量来源核对": ["平台", "店铺名称", "SKU编码", "商品名称", "近30天销量值", "来源文件", "来源Sheet", "来源字段", "来源原始值", "是否补充来源"],
        "W27_异常SKU源表对账": ["SKU编码", "来源文件", "来源Sheet", "来源行号", "店铺", "商品编码", "商品名称", "销售数量", "净销量", "销售金额", "净销售额", "销售成本", "净销售成本", "成本价", "最终写入销售额取自哪一列", "最终写入成交件数取自哪一列", "最终写入单件成本取自哪一列", "判断"],
        "开普乐_源销售明细_来源核对": ["日期", "渠道_标准", "店铺/客户", "SKU", "产品名称", "销售额_元", "销售成本_元", "平台利润_元", "退款金额_元", "广告费_元", "来源文件", "来源Sheet", "来源字段", "说明"],
        "开普乐_源库存快照_来源核对": ["日期", "渠道_标准", "仓库/区域", "SKU/分类", "成本价", "库存数量", "库存金额_元", "可售天数", "类型", "来源文件", "来源Sheet", "来源字段", "说明"],
    }

    first_ws = wb.active
    first_ws.title = required_sheets[0]
    for index, sheet_name in enumerate(required_sheets):
        ws = first_ws if index == 0 else wb.create_sheet(sheet_name)
        header = headers[sheet_name]
        ws.append(header)
        for row in sheet_rows.get(sheet_name, []):
            ws.append([row.get(column, "") for column in header])
    wb.save(output_path)


def build_recognition_rows(
    detections: dict[str, DetectionResult],
    reusable_targets: dict[str, Path],
    expected_roles: list[str],
) -> list[dict]:
    rows = []
    for role in expected_roles:
        detection = detections.get(role)
        if detection:
            rows.append(
                {
                    "文件角色": role_display_name(role),
                    "是否必需": "是" if role in {"w27_target", "kepule_target"} else "否",
                    "识别状态": detection.status,
                    "实际文件名": detection.path.name if detection.path else "",
                    "识别命中的Sheet": ",".join(detection.matched_sheets),
                    "实际处理的Sheet": ",".join(detection.processed_sheets or detection.matched_sheets),
                    "跳过的Sheet": ",".join(detection.skipped_sheets),
                    "命中的关键字段": ",".join(detection.matched_fields),
                    "缺失字段": ",".join(detection.missing_fields),
                    "影响范围": role,
                    "备注": detection.note,
                }
            )
        elif role in reusable_targets:
            rows.append(
                {
                    "文件角色": role_display_name(role),
                    "是否必需": "是",
                    "识别状态": "matched_from_partial_output",
                    "实际文件名": reusable_targets[role].name,
                    "识别命中的Sheet": "",
                    "实际处理的Sheet": "",
                    "跳过的Sheet": "",
                    "命中的关键字段": "",
                    "缺失字段": "",
                    "影响范围": role,
                    "备注": "使用上次部分填报结果作为基础。",
                }
            )
        else:
            rows.append(
                {
                    "文件角色": role_display_name(role),
                    "是否必需": "是" if role in {"w27_target", "kepule_target"} else "否",
                    "识别状态": "missing",
                    "实际文件名": "",
                    "识别命中的Sheet": "",
                    "实际处理的Sheet": "",
                    "跳过的Sheet": "",
                    "命中的关键字段": "",
                    "缺失字段": "",
                    "影响范围": role,
                    "备注": "未识别到对应文件。",
                }
            )
    return rows


def write_status(output_path: Path, status: str) -> None:
    output_path.write_text(status, encoding="utf-8")


def write_status_details(
    output_path: Path,
    status: str,
    core_fill_status: str,
    has_manual_review: bool,
    manual_review_count: int,
) -> None:
    lines = [
        f"状态={status}",
        f"核心填表状态={core_fill_status}",
        f"是否存在人工复核项={'是' if has_manual_review else '否'}",
        f"人工复核项数量={manual_review_count}",
    ]
    output_path.write_text("\n".join(lines), encoding="utf-8")


def write_delivery_summary(
    output_path: Path,
    status: str,
    core_fill_status: str,
    has_manual_review: bool,
    primary_files: list[str],
    technical_files: list[str],
) -> None:
    lines = [
        "给同事先看这些文件：",
        *[f"- {name}" for name in primary_files],
        "",
        f"本次状态：{status}",
        f"核心填表状态：{core_fill_status}",
        f"是否有建议人工复核项：{'是' if has_manual_review else '否'}",
        "",
        "技术调试文件（内部测试用）：",
        *[f"- {name}" for name in technical_files],
    ]
    output_path.write_text("\n".join(lines), encoding="utf-8")


def write_run_log(
    output_path: Path,
    status: str,
    sheet_results: list[SheetWriteResult],
    issues: list[Issue],
    fill_config: FillConfig | None = None,
    extra_lines: list[str] | None = None,
) -> None:
    lines = [f"status={status}"]
    if fill_config:
        lines.extend(
            [
                f"cycle_type={fill_config.cycle_type} ({fill_config.source_map['cycle_type']})",
                f"cycle_code={fill_config.cycle_code} ({fill_config.source_map['cycle_code']})",
                f"start_date={fill_config.start_date} ({fill_config.source_map['start_date']})",
                f"end_date={fill_config.end_date} ({fill_config.source_map['end_date']})",
                f"inventory_date={fill_config.inventory_date} ({fill_config.source_map['inventory_date']})",
                f"kepule_sales_date={fill_config.kepule_sales_date} ({fill_config.source_map['kepule_sales_date']})",
                f"kepule_inventory_date={fill_config.kepule_inventory_date} ({fill_config.source_map['kepule_inventory_date']})",
            ]
        )
    for result in sheet_results:
        lines.append(
            f"{result.workbook_role}.{result.sheet_name}: "
            f"raw_candidate_count={result.raw_candidate_count}; "
            f"skip_zero_sales_count={result.skip_zero_sales_count}; "
            f"skip_no_effect_count={result.skip_no_effect_count}; "
            f"candidate_after_filter_count={result.candidate_after_filter_count}; "
            f"original_cycle_rows={result.original_cycle_rows}; "
            f"existing_key_count={result.existing_key_count}; "
            f"template_blank_write_count={result.template_row_writes}; "
            f"append_row_count={result.appended_row_writes}; "
            f"backfill_blank_count={result.backfilled_rows}; "
            f"skipped_rows={result.skipped_rows}; "
            f"write_count={result.written_rows}; "
            f"ended_beyond_template={result.ended_beyond_template}"
        )
    counts = {
        "ERROR": sum(1 for issue in issues if issue.level == "ERROR"),
        "WARNING": sum(1 for issue in issues if issue.level == "WARNING"),
        "INFO": sum(1 for issue in issues if issue.level == "INFO"),
    }
    lines.append(f"issues={counts}")
    if extra_lines:
        lines.extend(extra_lines)
    output_path.write_text("\n".join(lines), encoding="utf-8")
