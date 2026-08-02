from __future__ import annotations

from copy import copy
from dataclasses import dataclass

from openpyxl.formula.translate import Translator
from openpyxl.utils import column_index_from_string, get_column_letter


LOGICAL_BUSINESS_COLUMNS = {
    "3_SKU明细": [
        "A", "B", "C", "D", "E", "F", "G", "H", "J", "K", "M", "O", "Q", "R",
        "S", "T", "U", "V", "W", "Y", "AB", "AC", "AD", "AE", "AG", "AH", "AJ", "AN",
    ],
    "7_库存补货": [
        "A", "B", "C", "D", "E", "F", "G", "H", "I", "K", "L", "M", "N", "S",
        "T", "U", "X",
    ],
    "源_销售明细": [
        "A", "E", "F", "G", "H", "I", "J", "K", "L", "M", "N", "O", "P", "Q",
        "T", "U", "V", "W",
    ],
    "源_库存快照": [
        "A", "D", "E", "F", "G", "H", "I", "J", "L", "M", "N", "O",
    ],
}


@dataclass
class SheetContext:
    header_row: int
    columns: dict[str, int]
    data_start_row: int


def find_header_row(ws, required_fields: list[str], search_rows: int = 10) -> int:
    for row_index, row_values in enumerate(
        ws.iter_rows(min_row=1, max_row=min(ws.max_row, search_rows), values_only=True),
        start=1,
    ):
        if all(field in row_values for field in required_fields):
            return row_index
    raise ValueError(f"Could not find header row in sheet {ws.title}")


def build_column_map(ws, header_row: int) -> dict[str, int]:
    mapping = {}
    values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
    for col, value in enumerate(values, start=1):
        if value is not None and str(value).strip():
            mapping[str(value).strip()] = col
    return mapping


def build_sheet_context(ws, required_fields: list[str]) -> SheetContext:
    header_row = find_header_row(ws, required_fields)
    return SheetContext(header_row=header_row, columns=build_column_map(ws, header_row), data_start_row=header_row + 1)


def _business_column_indexes(sheet_name: str) -> list[int]:
    return [column_index_from_string(letter) for letter in LOGICAL_BUSINESS_COLUMNS[sheet_name]]


def is_logical_blank_row(ws, row_index: int) -> bool:
    business_cols = _business_column_indexes(ws.title)
    return all(ws.cell(row_index, col).value in (None, "") for col in business_cols)


def first_logical_blank_row(ws, data_start_row: int) -> int | None:
    for row_index in range(data_start_row, ws.max_row + 1):
        if is_logical_blank_row(ws, row_index):
            return row_index
    return None


def first_non_blank_row_from_bottom(ws, data_start_row: int) -> int:
    for row_index in range(ws.max_row, data_start_row - 1, -1):
        if not is_logical_blank_row(ws, row_index):
            return row_index
    return data_start_row


def copy_template_row(ws, template_row: int, target_row: int) -> None:
    for col in range(1, ws.max_column + 1):
        source = ws.cell(template_row, col)
        dest = ws.cell(target_row, col)
        if source.has_style:
            dest._style = copy(source._style)
        if source.number_format:
            dest.number_format = source.number_format
        if source.font:
            dest.font = copy(source.font)
        if source.fill:
            dest.fill = copy(source.fill)
        if source.border:
            dest.border = copy(source.border)
        if source.alignment:
            dest.alignment = copy(source.alignment)
        if source.protection:
            dest.protection = copy(source.protection)
        if isinstance(source.value, str) and source.value.startswith("="):
            origin = f"{get_column_letter(col)}{template_row}"
            target = f"{get_column_letter(col)}{target_row}"
            dest.value = Translator(source.value, origin=origin).translate_formula(target)


def row_key(ws, row_index: int, columns: dict[str, int], key_fields: list[str]) -> tuple[str, ...]:
    return tuple(str(ws.cell(row_index, columns[field]).value or "") for field in key_fields)
