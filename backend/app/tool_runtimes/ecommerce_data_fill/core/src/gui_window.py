from __future__ import annotations

import os
import shutil
import sys
import tempfile
import traceback
import tkinter as tk
import calendar
from collections import deque
from datetime import datetime
from pathlib import Path
from threading import Thread
from tkinter import filedialog, messagebox, ttk

try:
    import windnd
except Exception:  # pragma: no cover - optional dependency on non-Windows envs
    windnd = None

from app import _detect_files, build_cli_values, run_amazon_inventory_fill, run_ecommerce_fill, run_kepule_fill, run_fill
from src.app_paths import app_base_dir
from src.date_rules import automatic_date_values, current_reporting_period
from openpyxl import load_workbook
from src.file_scanner import scan_excel_files
from src.gui_file_selection import ImportedFileSelection
from src.gui_import_trace import write_import_trace
from src.gui_presenters import (
    build_gui_file_slots,
    build_gui_review_items_from_rows,
    split_gui_file_slots,
    split_preflight_issues,
)
from src.models import Issue
from src.validators import validate_required_roles


WINDOW_TITLE = "电商数据自动填表助手"
ROOT_DIR = app_base_dir()
STEP_TITLES = ["导入文件", "确认条件", "预检查", "查看结果"]
WORKFLOW_ROLE_ORDER = {
    "ecommerce": ("w27_target", "sales_theme_analysis", "sales_30d", "product_archive", "jd_amazon_inventory", "jd_self_weekly_sales"),
    "kepule": ("kepule_target", "domestic_sales_theme_analysis", "domestic_sales_ranking", "cross_border_profit_sku", "product_archive", "jd_amazon_inventory", "jd_self_weekly_sales"),
    "amazon": ("amazon_inventory_target", "amazon_inventory_weekly", "fba_inventory"),
}
WORKFLOW_ROLES = {workflow: set(roles) for workflow, roles in WORKFLOW_ROLE_ORDER.items()}
WORKFLOW_TITLES = {
    "ecommerce": "电商数据分析表填写",
    "kepule": "周月报填写",
    "amazon": "亚马逊库存表填写",
}
WORKFLOW_CLOSE_GUIDANCE = {
    "ecommerce": "• 电商数据分析表目标模板\n• 本次导入的电商来源表\n• 上一次仍打开的电商输出表",
    "kepule": "• 周月报目标模板\n• 本次导入的周月报来源表\n• 上一次仍打开的周月报输出表",
    "amazon": "• 亚马逊库存目标模板\n• 亚马逊库存每周更新\n• FBA 仓库明细\n• 上一次仍打开的亚马逊库存输出表",
}


def decode_process_output(value: bytes | str | None) -> str:
    if value is None:
        return ""
    if isinstance(value, str):
        return value
    for encoding in ("utf-8", "gbk", sys.getfilesystemencoding()):
        try:
            return value.decode(encoding)
        except UnicodeDecodeError:
            continue
    return value.decode("utf-8", errors="replace")


def launch_gui(workflow: str | None = None, session_values: dict[str, object] | None = None) -> None:
    if workflow is None:
        _launch_selector(session_values)
        return
    root = tk.Tk()
    root.title(WORKFLOW_TITLES[workflow])
    root.geometry("1460x960")
    root.minsize(1220, 820)
    root.configure(bg="#F3F7FB")
    try:
        root.state("zoomed")
    except tk.TclError:
        pass

    _build_layout(root, workflow, session_values)
    root.mainloop()


def _launch_selector(session_values: dict[str, object] | None = None) -> None:
    root = tk.Tk()
    root.title(WINDOW_TITLE)
    root.geometry("720x430")
    root.minsize(680, 400)
    root.configure(bg="#F3F7FB")
    frame = ttk.Frame(root, padding=32)
    frame.pack(fill="both", expand=True)
    ttk.Label(frame, text="选择本次要填写的表", font=("Microsoft YaHei UI", 22, "bold")).pack(anchor="w")
    ttk.Label(frame, text="三个任务分开处理，只导入当前任务需要的目标表和来源文件。", font=("Microsoft YaHei UI", 10)).pack(anchor="w", pady=(8, 28))
    for code, title, detail in [
        ("ecommerce", "电商数据分析表", "填写 3_SKU明细、7_库存补货"),
        ("kepule", "周月报", "填写源_销售明细、源_库存快照及 TOP10"),
        ("amazon", "亚马逊库存表", "使用 FBA仓库明细 + 亚马逊库存每周更新"),
    ]:
        card = ttk.Frame(frame, padding=14)
        card.pack(fill="x", pady=6)
        ttk.Button(card, text=title, style="Primary.TButton", command=lambda value=code: (root.destroy(), launch_gui(value, session_values))).pack(side="left")
        ttk.Label(card, text=detail, font=("Microsoft YaHei UI", 10)).pack(side="left", padx=18)
    root.mainloop()


def _build_layout(root: tk.Tk, workflow: str, session_values: dict[str, object] | None = None) -> None:
    style = ttk.Style(root)
    try:
        style.theme_use("clam")
    except tk.TclError:
        pass

    style.configure(".", font=("Microsoft YaHei UI", 10))
    style.configure("Page.TFrame", background="#F4F7FB")
    style.configure("Card.TFrame", background="#FFFFFF", relief="flat")
    style.configure("CardTitle.TLabel", background="#FFFFFF", foreground="#102A43", font=("Microsoft YaHei UI", 12, "bold"))
    style.configure("Muted.TLabel", background="#FFFFFF", foreground="#5B7083")
    style.configure("HeroTitle.TLabel", background="#F3F7FB", foreground="#102A43", font=("Microsoft YaHei UI", 21, "bold"))
    style.configure("HeroSub.TLabel", background="#F3F7FB", foreground="#5B7083", font=("Microsoft YaHei UI", 10))
    style.configure("StepWrap.TFrame", background="#F4F7FB")
    style.configure("StepWrapActive.TFrame", background="#E8F7F4")
    style.configure("StepBadge.TLabel", background="#E6EEF6", foreground="#486581", font=("Microsoft YaHei UI", 9, "bold"), padding=(12, 7))
    style.configure("StepBadgeActive.TLabel", background="#0F766E", foreground="#FFFFFF", font=("Microsoft YaHei UI", 9, "bold"), padding=(12, 7))
    style.configure("StepName.TLabel", background="#F4F7FB", foreground="#243B53", font=("Microsoft YaHei UI", 10, "bold"))
    style.configure("StepNameActive.TLabel", background="#E8F7F4", foreground="#0F766E", font=("Microsoft YaHei UI", 10, "bold"))
    style.configure("Primary.TButton", padding=(22, 12), font=("Microsoft YaHei UI", 10, "bold"), background="#0F766E", foreground="#FFFFFF", borderwidth=0)
    style.map("Primary.TButton", background=[("active", "#115E59"), ("pressed", "#115E59")], foreground=[("active", "#FFFFFF")])
    style.configure("Secondary.TButton", padding=(18, 10), font=("Microsoft YaHei UI", 10), background="#E7EEF6", foreground="#243B53", borderwidth=0)
    style.map("Secondary.TButton", background=[("active", "#D9E8F5"), ("pressed", "#D9E8F5")])
    style.configure("Ghost.TButton", padding=(12, 8), font=("Microsoft YaHei UI", 10), background="#FFFFFF", foreground="#486581", borderwidth=0)
    style.map("Ghost.TButton", background=[("active", "#F0F4F8"), ("pressed", "#F0F4F8")])
    style.configure("Treeview", rowheight=28, font=("Microsoft YaHei UI", 10))
    style.configure("Treeview.Heading", font=("Microsoft YaHei UI", 10, "bold"))

    container = ttk.Frame(root, style="Page.TFrame", padding=20)
    container.pack(fill="both", expand=True)
    container.columnconfigure(0, weight=1)
    container.rowconfigure(3, weight=1)

    imported_file_selection = ImportedFileSelection()
    manual_role_files: dict[str, Path] = {}
    current_slots = []
    current_role_files: dict[str, Path] = {}
    current_step = 0
    is_running = False
    session_values = dict(session_values or {})

    status_var = tk.StringVar(value="当前提示：请先导入文件。")
    import_summary_var = tk.StringVar(value="尚未选择文件")
    selected_files_var = tk.StringVar(value="已选文件：无")
    output_dir_var = tk.StringVar(value=str(session_values.get("_output_dir") or (ROOT_DIR / "output_gui")))
    result_summary_var = tk.StringVar(value="尚未开始运行")
    detail_summary_var = tk.StringVar(value="运行完成后，这里会显示结果摘要。")
    running_var = tk.StringVar(value="")
    precheck_title_var = tk.StringVar(value="先完成文件导入和填写条件，再到这一步检查。")
    precheck_detail_var = tk.StringVar(value="程序会把必须补齐和建议补充的文件分开列出。")
    light_hint_var = tk.StringVar(value="")
    field_vars: dict[str, tk.StringVar] = {}
    date_field_vars: dict[str, tuple[tk.StringVar, tk.StringVar, tk.StringVar]] = {}
    current_period = current_reporting_period()

    output_listbox: tk.Listbox | None = None
    review_tree: ttk.Treeview | None = None
    latest_run_dir: Path | None = None
    source_tree: ttk.Treeview | None = None
    active_drop_zone: tk.Widget | None = None
    pending_drop_batches: deque[list] = deque()

    header = ttk.Frame(container, style="Page.TFrame")
    header.grid(row=0, column=0, sticky="ew")
    header.columnconfigure(0, weight=1)
    ttk.Label(header, text=WORKFLOW_TITLES[workflow], style="HeroTitle.TLabel").grid(row=0, column=0, sticky="w")
    def current_session_values() -> dict[str, object]:
        values: dict[str, object] = {
            label: variable.get().strip()
            for label, variable in field_vars.items()
        }
        values.update({
            label: "-".join(part.get().strip() for part in parts)
            for label, parts in date_field_vars.items()
        })
        try:
            values["_auto_date_overrides"] = list(auto_date_overrides)
        except NameError:
            values["_auto_date_overrides"] = []
        values["_output_dir"] = output_dir_var.get().strip()
        return values

    def choose_next_workflow() -> None:
        if is_running:
            messagebox.showinfo("正在填写", "当前任务仍在生成中，请完成后再选择下一张表。")
            return
        next_session_values = current_session_values()
        root.destroy()
        launch_gui(session_values=next_session_values)

    ttk.Button(header, text="选择下一张表", style="Secondary.TButton", command=choose_next_workflow).grid(row=0, column=1, sticky="e")
    ttk.Label(
        header,
        text="按向导一步一步完成：导入文件、确认条件、预检查、开始填写与查看结果。",
        style="HeroSub.TLabel",
    ).grid(row=1, column=0, sticky="w", pady=(6, 0))

    progress_frame = ttk.Frame(container, style="Page.TFrame", padding=(0, 16, 0, 12))
    progress_frame.grid(row=1, column=0, sticky="ew")
    for index in range(len(STEP_TITLES)):
        progress_frame.columnconfigure(index * 2, weight=1)
        if index < len(STEP_TITLES) - 1:
            progress_frame.columnconfigure(index * 2 + 1, weight=0)

    step_badges: list[ttk.Label] = []
    step_names: list[ttk.Label] = []
    connectors: list[tk.Frame] = []
    step_wrappers: list[ttk.Frame] = []
    for index, title in enumerate(STEP_TITLES):
        wrap = ttk.Frame(progress_frame, style="StepWrap.TFrame", padding=(10, 10))
        wrap.grid(row=0, column=index * 2, sticky="ew")
        wrap.columnconfigure(0, weight=1)
        badge = ttk.Label(wrap, text=f"步骤 {index + 1}", style="StepBadge.TLabel", anchor="center", cursor="hand2")
        badge.grid(row=0, column=0)
        name = ttk.Label(wrap, text=title, style="StepName.TLabel", anchor="center", cursor="hand2")
        name.grid(row=1, column=0, pady=(8, 0))
        step_wrappers.append(wrap)
        step_badges.append(badge)
        step_names.append(name)
        if index < len(STEP_TITLES) - 1:
            line = tk.Frame(progress_frame, bg="#D9E2EC", height=2)
            line.grid(row=0, column=index * 2 + 1, sticky="ew", padx=12, pady=(18, 0))
            connectors.append(line)

    hint_bar = tk.Label(
        container,
        textvariable=light_hint_var,
        bg="#D9F3EE",
        fg="#0F766E",
        font=("Microsoft YaHei UI", 10, "bold"),
        anchor="w",
        padx=14,
        pady=10,
    )
    hint_bar.grid(row=2, column=0, sticky="ew", pady=(0, 10))
    hint_bar.grid_remove()
    hint_job: str | None = None

    card = ttk.Frame(container, style="Card.TFrame", padding=18)
    card.grid(row=3, column=0, sticky="nsew")
    card.columnconfigure(0, weight=1)
    card.rowconfigure(0, weight=1)

    page_container = ttk.Frame(card, style="Card.TFrame")
    page_container.grid(row=0, column=0, sticky="nsew")
    page_container.columnconfigure(0, weight=1)
    page_container.rowconfigure(0, weight=1)

    pages: list[ttk.Frame] = []

    def make_scroll_page(parent: ttk.Frame) -> tuple[ttk.Frame, ttk.Frame, tk.Canvas]:
        page = ttk.Frame(parent, style="Card.TFrame")
        page.columnconfigure(0, weight=1)
        page.rowconfigure(0, weight=1)

        canvas = tk.Canvas(page, bg="#FFFFFF", highlightthickness=0, bd=0)
        scrollbar = ttk.Scrollbar(page, orient="vertical", command=canvas.yview)
        canvas.grid(row=0, column=0, sticky="nsew")
        scrollbar.grid(row=0, column=1, sticky="ns")
        canvas.configure(yscrollcommand=scrollbar.set)

        inner = ttk.Frame(canvas, style="Card.TFrame", padding=4)
        window = canvas.create_window((0, 0), window=inner, anchor="nw")
        inner.columnconfigure(0, weight=1)
        scroll_sync_job: str | None = None
        last_inner_width = 0

        def _sync_scroll(_event=None) -> None:
            nonlocal scroll_sync_job
            if scroll_sync_job is not None:
                canvas.after_cancel(scroll_sync_job)
            scroll_sync_job = canvas.after(40, _apply_scroll_region)

        def _apply_scroll_region() -> None:
            nonlocal scroll_sync_job
            scroll_sync_job = None
            canvas.configure(scrollregion=canvas.bbox("all"))

        def _resize_inner(event) -> None:
            nonlocal last_inner_width
            if event.width == last_inner_width:
                return
            last_inner_width = event.width
            canvas.itemconfigure(window, width=event.width)

        def _bind_wheel(_event=None) -> None:
            canvas.bind_all(
                "<MouseWheel>",
                lambda e, target=canvas: target.yview_scroll(int(-1 * (e.delta / 120)), "units"),
            )

        def _unbind_wheel(_event=None) -> None:
            canvas.unbind_all("<MouseWheel>")

        inner.bind("<Configure>", _sync_scroll)
        canvas.bind("<Configure>", _resize_inner)
        canvas.bind("<Enter>", _bind_wheel)
        inner.bind("<Enter>", _bind_wheel)
        canvas.bind("<Leave>", _unbind_wheel)
        return page, inner, canvas

    step1_page = ttk.Frame(page_container, style="Card.TFrame")
    step1_page.grid(row=0, column=0, sticky="nsew")
    step1_page.columnconfigure(0, weight=1)
    step1_page.rowconfigure(5, weight=1)
    pages.append(step1_page)

    ttk.Label(step1_page, text="第 1 步：导入本周文件", style="CardTitle.TLabel").grid(row=0, column=0, sticky="w")
    ttk.Label(
        step1_page,
        text="可连续批量选择文件或添加多个文件夹；之前已选文件会保留，只有“清空本次已选文件”才会清空。",
        style="Muted.TLabel",
    ).grid(row=1, column=0, sticky="w", pady=(8, 16))

    drop_card = tk.Frame(step1_page, bg="#F8FBFF", bd=0, relief="solid", highlightbackground="#BFD2E6", highlightcolor="#0F766E", highlightthickness=1)
    drop_card.grid(row=2, column=0, sticky="ew", pady=(0, 18))
    drop_card.columnconfigure(0, weight=1)
    active_drop_zone = drop_card

    drop_title = tk.Label(drop_card, text="把 Excel 文件拖到这里", bg="#F8FBFF", fg="#102A43", font=("Microsoft YaHei UI", 18, "bold"))
    drop_title.grid(row=0, column=0, pady=(26, 6))
    drop_sub = tk.Label(drop_card, text="三个来源文件夹可依次添加；已有文件不会被清空，重复文件会自动跳过。", bg="#F8FBFF", fg="#5B7083", font=("Microsoft YaHei UI", 10))
    drop_sub.grid(row=1, column=0, pady=(0, 18))
    drop_actions = ttk.Frame(drop_card, style="Card.TFrame")
    drop_actions.grid(row=2, column=0, pady=(0, 24))

    def append_imported_paths(paths: list[Path]) -> int:
        """Add new files without discarding the user's earlier selections."""
        return imported_file_selection.add(paths)

    def choose_files() -> None:
        selected = filedialog.askopenfilenames(
            title="选择本周 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm *.xls"), ("所有文件", "*.*")],
        )
        if not selected:
            return
        selected_paths = [Path(item) for item in selected]
        added = append_imported_paths(selected_paths)
        trace_path = write_import_trace(
            ROOT_DIR,
            action="files",
            selected_path="; ".join(str(path) for path in selected_paths),
            scanned_count=len(selected_paths),
            accepted_count=added,
        )
        refresh_slots()
        if added == 0:
            messagebox.showwarning(
                "未导入新文件",
                f"本次选择了 {len(selected_paths)} 个文件，但没有新增可用文件。\n"
                f"可能是重复选择、文件已被移动，或路径暂不可访问。\n"
                f"诊断日志：\n{trace_path}",
            )
        show_light_hint(f"已添加 {added} 个文件；之前已选文件仍会保留。")

    def choose_folder() -> None:
        selected = filedialog.askdirectory(title="添加一个本周 Excel 文件夹（可连续添加三个文件夹）")
        if not selected:
            return
        folder = Path(selected)
        try:
            new_paths = [
                entry.path
                for entry in scan_excel_files(folder, folder, recursive=True).input_files
            ]
        except Exception as error:
            trace_path = write_import_trace(
                ROOT_DIR,
                action="folder_scan_error",
                selected_path=str(folder),
                scanned_count=0,
                accepted_count=0,
            )
            messagebox.showerror(
                "读取文件夹失败",
                f"无法读取：\n{folder}\n\n{error}\n\n诊断日志：\n{trace_path}",
            )
            return
        added = append_imported_paths(new_paths)
        trace_path = write_import_trace(
            ROOT_DIR,
            action="folder",
            selected_path=str(folder),
            scanned_count=len(new_paths),
            accepted_count=added,
        )
        refresh_slots()
        if added == 0:
            messagebox.showwarning(
                "未导入新文件",
                f"已选择文件夹：\n{folder}\n\n"
                f"扫描到 {len(new_paths)} 个 Excel 文件，但没有新增文件。\n"
                f"若扫描数为 0，请确认选中的是来源文件夹，而不是程序文件夹。\n"
                f"诊断日志：\n{trace_path}",
            )
        show_light_hint(f"已添加 {added} 个文件；还可以继续添加其他文件夹。")

    ttk.Button(drop_actions, text="批量选择文件", style="Primary.TButton", command=choose_files).pack(side="left")
    ttk.Button(drop_actions, text="添加文件夹", style="Secondary.TButton", command=choose_folder).pack(side="left", padx=(10, 0))
    ttk.Label(step1_page, textvariable=import_summary_var, style="Muted.TLabel").grid(row=3, column=0, sticky="w", pady=(0, 12))
    ttk.Label(step1_page, textvariable=selected_files_var, style="Muted.TLabel", wraplength=1360).grid(row=4, column=0, sticky="w", pady=(0, 12))

    split = ttk.Panedwindow(step1_page, orient="horizontal")
    split.grid(row=5, column=0, sticky="nsew")
    step1_page.rowconfigure(5, weight=1)

    left = ttk.Frame(split, style="Card.TFrame")
    left.columnconfigure(0, weight=1)
    left.rowconfigure(1, weight=1)
    right = ttk.Frame(split, style="Card.TFrame")
    right.columnconfigure(0, weight=1)
    right.rowconfigure(1, weight=1)
    split.add(left, weight=5)
    split.add(right, weight=5)

    ttk.Label(left, text="目标模板", style="CardTitle.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 8))
    target_frame = ttk.Frame(left, style="Card.TFrame")
    target_frame.grid(row=1, column=0, sticky="nsew", padx=(0, 10))
    target_frame.columnconfigure(0, weight=1)
    target_frame.rowconfigure(0, weight=1)

    tree = ttk.Treeview(target_frame, columns=("label", "status", "file"), show="headings", height=8)
    tree.heading("label", text="文件类型")
    tree.heading("status", text="状态")
    tree.heading("file", text="当前文件")
    tree.column("label", width=240, anchor="w")
    tree.column("status", width=120, anchor="center")
    tree.column("file", width=420, anchor="w")
    tree.grid(row=0, column=0, sticky="nsew")
    ttk.Scrollbar(target_frame, orient="vertical", command=tree.yview).grid(row=0, column=1, sticky="ns")
    tree.configure(yscrollcommand=lambda first, last: target_frame.grid_slaves(row=0, column=1)[0].set(first, last))

    ttk.Label(right, text="来源文件", style="CardTitle.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 8))
    source_frame = ttk.Frame(right, style="Card.TFrame")
    source_frame.grid(row=1, column=0, sticky="nsew")
    source_frame.columnconfigure(0, weight=1)
    source_frame.rowconfigure(0, weight=1)
    source_tree = ttk.Treeview(source_frame, columns=("label", "status", "file"), show="headings", height=8)
    source_tree.heading("label", text="文件类型")
    source_tree.heading("status", text="状态")
    source_tree.heading("file", text="当前文件")
    source_tree.column("label", width=240, anchor="w")
    source_tree.column("status", width=120, anchor="center")
    source_tree.column("file", width=420, anchor="w")
    source_tree.grid(row=0, column=0, sticky="nsew")
    ttk.Scrollbar(source_frame, orient="vertical", command=source_tree.yview).grid(row=0, column=1, sticky="ns")
    source_tree.configure(yscrollcommand=lambda first, last: source_frame.grid_slaves(row=0, column=1)[0].set(first, last))

    tool_row = ttk.Frame(step1_page, style="Card.TFrame")
    tool_row.grid(row=6, column=0, sticky="ew", pady=(14, 0))

    def pick_role_file() -> None:
        selection = tree.selection() or (() if source_tree is None else source_tree.selection())
        if not selection:
            messagebox.showinfo("补选文件", "请先选中要补选的文件类型。")
            return
        role = selection[0]
        selected = filedialog.askopenfilename(
            title="为当前文件类型补选 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xlsm *.xls"), ("所有文件", "*.*")],
        )
        if not selected:
            return
        manual_role_files[role] = Path(selected)
        refresh_slots()

    def get_selected_role() -> str | None:
        selection = tree.selection()
        if selection:
            return str(selection[0])
        if source_tree is not None:
            source_selection = source_tree.selection()
            if source_selection:
                return str(source_selection[0])
        return None

    def remove_selected_file() -> None:
        role = get_selected_role()
        if not role:
            messagebox.showinfo("移除文件", "请先选中要移除的文件类型。")
            return
        selected_path = current_role_files.get(role)
        if selected_path is None:
            messagebox.showinfo("移除文件", "当前选中项还没有对应文件，无需移除。")
            return
        manual_role_files.pop(role, None)
        imported_file_selection.remove(selected_path)
        for manual_role, manual_path in list(manual_role_files.items()):
            if manual_path.resolve() == selected_path.resolve():
                manual_role_files.pop(manual_role, None)
        refresh_slots()
        show_light_hint(f"已移除：{selected_path.name}")

    ttk.Button(tool_row, text="为选中项补选文件", style="Secondary.TButton", command=pick_role_file).pack(side="left")
    ttk.Button(tool_row, text="移除选中文件", style="Ghost.TButton", command=remove_selected_file).pack(side="left", padx=(10, 0))
    ttk.Button(tool_row, text="重新识别", style="Ghost.TButton", command=lambda: refresh_slots()).pack(side="left", padx=(10, 0))
    ttk.Button(
        tool_row,
        text="清空本次已选文件",
        style="Ghost.TButton",
        command=lambda: (imported_file_selection.clear(), manual_role_files.clear(), refresh_slots()),
    ).pack(side="left", padx=(10, 0))

    step2_page, step2_inner, _ = make_scroll_page(page_container)
    step2_page.grid(row=0, column=0, sticky="nsew")
    pages.append(step2_page)
    step2_inner.columnconfigure(0, weight=1)
    ttk.Label(step2_inner, text="第 2 步：确认本次填写条件", style="CardTitle.TLabel").grid(row=0, column=0, sticky="w")
    ttk.Label(step2_inner, text="程序已经带出默认值。只在确认确实不同的时候再改。", style="Muted.TLabel").grid(row=1, column=0, sticky="w", pady=(8, 18))

    form_card = ttk.Frame(step2_inner, style="Card.TFrame")
    form_card.grid(row=2, column=0, sticky="ew")
    form_card.columnconfigure(1, weight=1)

    fields = [
        ("周期类型", str(session_values.get("周期类型") or current_period["周期类型"])),
        ("周次编码", str(session_values.get("周次编码") or current_period["周次编码"])),
    ]
    for index, (label_text, default_value) in enumerate(fields):
        ttk.Label(form_card, text=label_text, style="Muted.TLabel").grid(row=index, column=0, sticky="w", pady=8, padx=(0, 10))
        field_var = tk.StringVar(value=default_value)
        field_vars[label_text] = field_var
        ttk.Entry(form_card, textvariable=field_var).grid(row=index, column=1, sticky="ew", pady=8)

    def open_calendar(year_var: tk.StringVar, month_var: tk.StringVar, day_var: tk.StringVar) -> None:
        today = datetime.today()
        try:
            shown_year = int(year_var.get())
            shown_month = int(month_var.get())
            datetime(shown_year, shown_month, 1)
        except ValueError:
            shown_year, shown_month = today.year, today.month

        dialog = tk.Toplevel(root)
        dialog.title("选择日期")
        dialog.transient(root)
        dialog.grab_set()
        state = {"year": shown_year, "month": shown_month}
        body = ttk.Frame(dialog, padding=12)
        body.grid(sticky="nsew")

        def render() -> None:
            for child in body.winfo_children():
                child.destroy()
            ttk.Button(body, text="上月", command=lambda: change_month(-1)).grid(row=0, column=0, padx=2)
            ttk.Label(body, text=f"{state['year']}年{state['month']}月", style="CardTitle.TLabel").grid(row=0, column=1, columnspan=4, pady=(0, 8))
            ttk.Button(body, text="今天", command=lambda: choose(today.day, today.year, today.month)).grid(row=0, column=5, padx=2)
            ttk.Button(body, text="下月", command=lambda: change_month(1)).grid(row=0, column=6, padx=2)
            for column, label in enumerate(("一", "二", "三", "四", "五", "六", "日")):
                ttk.Label(body, text=label, style="Muted.TLabel").grid(row=1, column=column, padx=5, pady=3)
            for row, week in enumerate(calendar.monthcalendar(state["year"], state["month"]), start=2):
                for column, day in enumerate(week):
                    if day:
                        ttk.Button(body, text=str(day), width=4, command=lambda selected=day: choose(selected)).grid(row=row, column=column, padx=2, pady=2)

        def change_month(delta: int) -> None:
            state["month"] += delta
            if state["month"] == 13:
                state["year"], state["month"] = state["year"] + 1, 1
            elif state["month"] == 0:
                state["year"], state["month"] = state["year"] - 1, 12
            render()

        def choose(day: int, year: int | None = None, month: int | None = None) -> None:
            year_var.set(str(year or state["year"]))
            month_var.set(f"{(month or state['month']):02d}")
            day_var.set(f"{day:02d}")
            dialog.destroy()

        render()

    initial_start_date = str(session_values.get("开始日期") or current_period["开始日期"])
    initial_end_date = str(session_values.get("结束日期") or current_period["结束日期"])
    initial_auto_dates = automatic_date_values(initial_start_date, initial_end_date)

    def date_default(label: str) -> tuple[str, str, str]:
        value = str(session_values.get(label) or {
            "开始日期": initial_start_date,
            "结束日期": initial_end_date,
        }.get(label) or initial_auto_dates[label])
        year, month, day = value.split("-")
        return year, month, day

    date_fields = [
        ("开始日期", date_default("开始日期"), "Wxx 电商表周期开始；周月报控制台本周开始"),
        ("结束日期", date_default("结束日期"), "Wxx 电商表周期结束；周月报控制台本周结束"),
        ("库存快照日期", date_default("库存快照日期"), "自动：同电商表开始日期；可手动调整"),
        ("开普乐销售日期", date_default("开普乐销售日期"), "自动：周月报本月周起点；可手动调整"),
        ("开普乐库存日期", date_default("开普乐库存日期"), "自动：同周月报销售日期；可手动调整"),
    ]
    for offset, (label_text, defaults, hint) in enumerate(date_fields, start=len(fields)):
        ttk.Label(form_card, text=label_text, style="Muted.TLabel").grid(row=offset, column=0, sticky="w", pady=8, padx=(0, 10))
        date_row = ttk.Frame(form_card, style="Card.TFrame")
        date_row.grid(row=offset, column=1, sticky="w", pady=8)
        year_var, month_var, day_var = (tk.StringVar(value=value) for value in defaults)
        date_field_vars[label_text] = (year_var, month_var, day_var)
        for variable, width, suffix in ((year_var, 7, "年"), (month_var, 4, "月"), (day_var, 4, "日")):
            ttk.Entry(date_row, textvariable=variable, width=width).pack(side="left")
            ttk.Label(date_row, text=suffix, style="Muted.TLabel").pack(side="left", padx=(3, 9))
        ttk.Button(date_row, text="选择日期", style="Secondary.TButton", command=lambda y=year_var, m=month_var, d=day_var: open_calendar(y, m, d)).pack(side="left", padx=(0, 10))
        ttk.Label(date_row, text=hint, style="Muted.TLabel").pack(side="left")

    auto_date_labels = ("库存快照日期", "开普乐销售日期", "开普乐库存日期")
    auto_date_overrides: set[str] = set(session_values.get("_auto_date_overrides", []))
    automatic_update = False
    kepule_period_hint_var = tk.StringVar(value="")

    def _date_text(label: str) -> str:
        return "-".join(part.get().strip() for part in date_field_vars[label])

    def _set_date_text(label: str, value: str) -> None:
        year, month, day = value.split("-")
        for variable, part in zip(date_field_vars[label], (year, month, day)):
            variable.set(part)

    def refresh_automatic_dates() -> None:
        nonlocal automatic_update
        try:
            auto_values = automatic_date_values(_date_text("开始日期"), _date_text("结束日期"))
        except ValueError:
            return
        automatic_update = True
        try:
            for label, value in auto_values.items():
                if label not in auto_date_overrides:
                    _set_date_text(label, value)
        finally:
            automatic_update = False
        kepule_period_hint_var.set(
            f"周月报控制台：{_date_text('开普乐销售日期')} 至 {_date_text('结束日期')}；源销售明细使用销售日期，源库存快照使用库存日期。"
        )

    def mark_manual_override(label: str, *_args) -> None:
        if not automatic_update:
            auto_date_overrides.add(label)
            refresh_automatic_dates()

    for label in auto_date_labels:
        for variable in date_field_vars[label]:
            variable.trace_add("write", lambda *_args, field=label: mark_manual_override(field))
    for label in ("开始日期", "结束日期"):
        for variable in date_field_vars[label]:
            variable.trace_add("write", lambda *_args: refresh_automatic_dates())

    def restore_automatic_dates() -> None:
        auto_date_overrides.clear()
        refresh_automatic_dates()

    refresh_automatic_dates()

    ttk.Label(form_card, text="日期按年 / 月 / 日填写；程序会统一转换格式。", style="Muted.TLabel").grid(
        row=len(fields) + len(date_fields), column=1, sticky="w", pady=(0, 8)
    )
    ttk.Label(form_card, textvariable=kepule_period_hint_var, style="Muted.TLabel", wraplength=900, justify="left").grid(
        row=len(fields) + len(date_fields) + 1, column=1, sticky="w", pady=(0, 8)
    )
    ttk.Button(form_card, text="恢复自动日期", style="Secondary.TButton", command=restore_automatic_dates).grid(
        row=len(fields) + len(date_fields) + 2, column=1, sticky="w", pady=(0, 8)
    )

    output_row = len(fields) + len(date_fields) + 3
    ttk.Label(form_card, text="结果输出文件夹", style="Muted.TLabel").grid(row=output_row, column=0, sticky="w", pady=8, padx=(0, 10))
    ttk.Entry(form_card, textvariable=output_dir_var).grid(row=output_row, column=1, sticky="ew", pady=8)

    def choose_output_dir() -> None:
        selected = filedialog.askdirectory(title="选择结果输出文件夹")
        if selected:
            output_dir_var.set(selected)

    ttk.Button(form_card, text="选择输出位置", style="Secondary.TButton", command=choose_output_dir).grid(row=output_row + 1, column=1, sticky="w", pady=(8, 0))

    step3_page, step3_inner, _ = make_scroll_page(page_container)
    step3_page.grid(row=0, column=0, sticky="nsew")
    pages.append(step3_page)
    step3_inner.columnconfigure(0, weight=1)
    ttk.Label(step3_inner, text="第 3 步：运行前检查", style="CardTitle.TLabel").grid(row=0, column=0, sticky="w")
    ttk.Label(step3_inner, text="先看哪些文件必须补齐，再决定是否开始填写。", style="Muted.TLabel").grid(row=1, column=0, sticky="w", pady=(8, 18))

    status_card = ttk.Frame(step3_inner, style="Card.TFrame")
    status_card.grid(row=2, column=0, sticky="ew")
    ttk.Label(status_card, textvariable=precheck_title_var, style="CardTitle.TLabel", wraplength=980, justify="left").grid(row=0, column=0, sticky="w")
    ttk.Label(status_card, textvariable=precheck_detail_var, style="Muted.TLabel", wraplength=980, justify="left").grid(row=1, column=0, sticky="w", pady=(10, 0))
    ttk.Label(status_card, textvariable=status_var, style="Muted.TLabel", wraplength=980, justify="left").grid(row=2, column=0, sticky="w", pady=(14, 0))
    ttk.Label(
        status_card,
        text="推荐顺序：\n1. 回到上一步确认日期\n2. 点击“运行前检查”\n3. 如果必须文件都齐了，再点“开始填写”",
        style="Muted.TLabel",
        justify="left",
    ).grid(row=3, column=0, sticky="w", pady=(18, 0))

    action_row = ttk.Frame(step3_inner, style="Card.TFrame")
    action_row.grid(row=3, column=0, sticky="w", pady=(20, 0))
    ttk.Button(action_row, text="运行前检查", style="Secondary.TButton", command=lambda: run_precheck()).pack(side="left")
    ttk.Button(action_row, text="开始填写", style="Primary.TButton", command=lambda: start_run()).pack(side="left", padx=(10, 0))
    ttk.Label(step3_inner, textvariable=running_var, style="Muted.TLabel", wraplength=980, justify="left").grid(row=4, column=0, sticky="w", pady=(18, 0))

    step4_page = ttk.Frame(page_container, style="Card.TFrame")
    step4_page.grid(row=0, column=0, sticky="nsew")
    step4_page.columnconfigure(0, weight=1)
    step4_page.columnconfigure(1, weight=1)
    step4_page.rowconfigure(3, weight=1)
    pages.append(step4_page)

    ttk.Label(step4_page, text="第 4 步：查看结果与人工核对", style="CardTitle.TLabel").grid(row=0, column=0, columnspan=2, sticky="w")
    ttk.Label(step4_page, textvariable=result_summary_var, style="Muted.TLabel", wraplength=1000, justify="left").grid(row=1, column=0, columnspan=2, sticky="w", pady=(8, 0))
    ttk.Label(step4_page, textvariable=detail_summary_var, style="Muted.TLabel", wraplength=1000, justify="left").grid(row=2, column=0, columnspan=2, sticky="w", pady=(8, 16))

    output_box = ttk.Frame(step4_page, style="Card.TFrame")
    output_box.grid(row=3, column=0, sticky="nsew", padx=(0, 10))
    output_box.columnconfigure(0, weight=1)
    output_box.rowconfigure(1, weight=1)
    ttk.Label(output_box, text="已生成文件", style="CardTitle.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 8))
    output_listbox = tk.Listbox(output_box, height=12, bd=0, relief="flat", highlightthickness=1, highlightbackground="#D9E2EC")
    output_listbox.grid(row=1, column=0, sticky="nsew")
    output_scroll = ttk.Scrollbar(output_box, orient="vertical", command=output_listbox.yview)
    output_scroll.grid(row=1, column=1, sticky="ns")
    output_listbox.configure(yscrollcommand=output_scroll.set)

    def open_path(path: Path) -> None:
        if not path.exists():
            messagebox.showwarning(
                "文件未找到",
                f"找不到该文件或文件夹，可能已被移动或删除：\n{path}",
            )
            return
        try:
            os.startfile(path)  # type: ignore[attr-defined]
        except Exception:
            messagebox.showinfo("打开文件", f"请手动打开：\n{path}")

    def open_selected_output() -> None:
        selection = output_listbox.curselection()
        if not selection:
            return
        selected_name = output_listbox.get(selection[0])
        if latest_run_dir is None:
            messagebox.showwarning("文件未找到", "当前没有可打开的输出文件。")
            return
        open_path(latest_run_dir / selected_name)

    output_listbox.bind("<Double-Button-1>", lambda _event: open_selected_output())

    output_actions = ttk.Frame(output_box, style="Card.TFrame")
    output_actions.grid(row=2, column=0, sticky="w", pady=(10, 0))
    ttk.Button(output_actions, text="打开选中文件", style="Secondary.TButton", command=open_selected_output).pack(side="left")
    ttk.Button(
        output_actions,
        text="打开输出文件夹",
        style="Ghost.TButton",
        command=lambda: open_path(latest_run_dir) if latest_run_dir else messagebox.showwarning("文件未找到", "当前没有可打开的输出文件夹。"),
    ).pack(side="left", padx=(10, 0))

    review_box = ttk.Frame(step4_page, style="Card.TFrame")
    review_box.grid(row=3, column=1, sticky="nsew")
    review_box.columnconfigure(0, weight=1)
    review_box.rowconfigure(1, weight=1)
    ttk.Label(review_box, text="建议人工核对", style="CardTitle.TLabel").grid(row=0, column=0, sticky="w", pady=(0, 8))
    review_tree = ttk.Treeview(review_box, columns=("result_location", "source_location", "field", "reason", "action"), show="headings", height=12)
    review_tree.heading("result_location", text="结果位置")
    review_tree.heading("source_location", text="来源位置")
    review_tree.heading("field", text="字段")
    review_tree.heading("reason", text="原因说明")
    review_tree.heading("action", text="建议操作")
    review_tree.column("result_location", width=180, anchor="w")
    review_tree.column("source_location", width=280, anchor="w")
    review_tree.column("field", width=120, anchor="w")
    review_tree.column("reason", width=280, anchor="w")
    review_tree.column("action", width=230, anchor="w")
    review_tree.grid(row=1, column=0, sticky="nsew")
    review_scroll = ttk.Scrollbar(review_box, orient="vertical", command=review_tree.yview)
    review_scroll.grid(row=1, column=1, sticky="ns")
    review_tree.configure(yscrollcommand=review_scroll.set)

    footer = ttk.Frame(container, style="Page.TFrame", padding=(0, 16, 0, 0))
    footer.grid(row=4, column=0, sticky="ew")
    footer.columnconfigure(0, weight=1)
    nav_hint_var = tk.StringVar(value="先把本周文件拖入大区域，再点“下一步”。")
    ttk.Label(footer, textvariable=nav_hint_var, style="HeroSub.TLabel").grid(row=0, column=0, sticky="w")
    nav_buttons = ttk.Frame(footer, style="Page.TFrame")
    nav_buttons.grid(row=0, column=1, sticky="e")
    prev_button = ttk.Button(nav_buttons, text="上一步", style="Secondary.TButton")
    prev_button.pack(side="left")
    next_button = ttk.Button(nav_buttons, text="下一步", style="Primary.TButton")
    next_button.pack(side="left", padx=(10, 0))
    def show_light_hint(message: str) -> None:
        nonlocal hint_job
        light_hint_var.set("提示： " + message)
        hint_bar.grid()
        if hint_job is not None:
            root.after_cancel(hint_job)
        hint_job = root.after(2600, lambda: hint_bar.grid_remove())

    def refresh_slots() -> None:
        nonlocal current_slots, current_role_files
        tree.delete(*tree.get_children())
        if source_tree is not None:
            source_tree.delete(*source_tree.get_children())
        imported_files = imported_file_selection.files
        detections = _detect_files(imported_files)
        role_files = {role: detection.path for role, detection in detections.items() if detection.path}
        role_files.update(manual_role_files)
        current_role_files = dict(role_files)
        workflow_files = {role: role_files[role] for role in WORKFLOW_ROLE_ORDER[workflow] if role in role_files}
        required_roles = set(WORKFLOW_ROLE_ORDER[workflow]) if workflow == "amazon" else {WORKFLOW_ROLE_ORDER[workflow][0]}
        current_slots = build_gui_file_slots(
            workflow_files,
            role_order=WORKFLOW_ROLE_ORDER[workflow],
            required_roles=required_roles,
        )
        import_summary_var.set(
            f"已导入 {len(imported_files)} 个可访问文件，已识别 {len(role_files)} 类；"
            f"另有 {len(imported_file_selection.unavailable_files)} 个路径暂不可访问；"
            f"手动补选 {len(manual_role_files)} 项。"
            if imported_files or imported_file_selection.unavailable_files or manual_role_files
            else "尚未选择文件"
        )
        selected_paths = [*imported_files, *imported_file_selection.unavailable_files]
        if selected_paths:
            displayed_names = [path.name for path in selected_paths]
            preview = "、".join(displayed_names[:6])
            if len(displayed_names) > 6:
                preview += f" 等 {len(displayed_names)} 个"
            if imported_file_selection.unavailable_files:
                unavailable_preview = "、".join(str(path) for path in imported_file_selection.unavailable_files[:2])
                selected_files_var.set(
                    "已选原始文件：" + preview + "；以下路径当前无法访问：" + unavailable_preview
                )
                show_light_hint("已保留选择，但有文件路径暂时无法访问；请检查文件是否仍存在或是否为网络盘。")
            else:
                selected_files_var.set("已选原始文件：" + preview)
        else:
            selected_files_var.set("已选原始文件：无")
        target_slots, source_slots = split_gui_file_slots(current_slots)
        for slot in target_slots:
            tag = "required_missing" if slot.required and slot.status_text != "已识别" else "optional_missing" if slot.status_text != "已识别" else "ready"
            tree.insert("", "end", iid=slot.role, values=(slot.label, slot.status_text, slot.file_name), tags=(tag,))
        if source_tree is not None:
            for slot in source_slots:
                tag = "required_missing" if slot.required and slot.status_text != "已识别" else "optional_missing" if slot.status_text != "已识别" else "ready"
                source_tree.insert("", "end", iid=slot.role, values=(slot.label, slot.status_text, slot.file_name), tags=(tag,))
        tree.tag_configure("required_missing", foreground="#B42318")
        tree.tag_configure("optional_missing", foreground="#A15C07")
        tree.tag_configure("ready", foreground="#0F5132")
        source_tree.tag_configure("required_missing", foreground="#B42318")
        source_tree.tag_configure("optional_missing", foreground="#A15C07")
        source_tree.tag_configure("ready", foreground="#0F5132")

        missing_required = [slot.label for slot in current_slots if slot.required and slot.status_text != "已识别"]
        missing_optional = [slot.label for slot in current_slots if (not slot.required) and slot.status_text != "已识别"]
        if missing_required:
            status_var.set("当前提示：必须补齐 " + "、".join(missing_required))
        elif missing_optional:
            status_var.set("当前提示：可继续运行；建议补充 " + "、".join(missing_optional[:3]))
        else:
            status_var.set("当前提示：文件已识别完成，可以继续下一步。")

    def collect_issues() -> tuple[dict[str, Path], list[Issue]]:
        detections = _detect_files(imported_file_selection.files)
        role_files = {role: detection.path for role, detection in detections.items() if detection.path}
        role_files.update(manual_role_files)
        amazon_roles = {"amazon_inventory_target", "amazon_inventory_weekly", "fba_inventory"}
        if workflow == "amazon":
            return role_files, [
                Issue("ERROR", "validator", role, f"缺少“{role}”", target_table="亚马逊库存")
                for role in sorted(amazon_roles - set(role_files))
            ]
        target_role = "w27_target" if workflow == "ecommerce" else "kepule_target"
        if target_role not in role_files:
            return role_files, [Issue("ERROR", "validator", target_role, f"缺少“{target_role}”")]
        return role_files, []

    def run_precheck() -> None:
        _, issues = collect_issues()
        required_items, optional_items = split_preflight_issues(issues)
        if required_items:
            precheck_title_var.set("还有必须文件没补齐，先处理这些项目。")
            precheck_detail_var.set("必须补齐：\n- " + "\n- ".join(required_items))
            if optional_items:
                status_var.set("建议补充： " + "、".join(optional_items[:5]))
        elif optional_items:
            precheck_title_var.set("主流程已经可以继续，但建议补充以下文件。")
            precheck_detail_var.set("建议补充：\n- " + "\n- ".join(optional_items))
            status_var.set("当前提示：可以开始填写，也可以先补充建议文件。")
        else:
            precheck_title_var.set("检查通过，可以开始填写。")
            precheck_detail_var.set("必须文件和建议文件都已就绪。")
            status_var.set("当前提示：检查通过，可以开始填写。")
        show_light_hint("已更新预检查结果。")

    def read_review_items(output_dir: Path) -> list:
        review_path = output_dir / "建议人工复核.xlsx"
        if not review_path.exists():
            return []
        workbook = load_workbook(review_path, data_only=True)
        ws = workbook[workbook.sheetnames[0]]
        headers = [cell.value for cell in ws[1]]
        rows: list[dict[str, str]] = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not any(row):
                continue
            row_map = {str(header or ""): "" if value is None else str(value) for header, value in zip(headers, row)}
            if not str(row_map.get("SKU", "")).strip():
                row_map["SKU"] = row_map.get("SKU编码", "")
            rows.append(row_map)
        workbook.close()
        return build_gui_review_items_from_rows(rows)

    def go_to_step(step_index: int, *, from_click: bool = False) -> None:
        nonlocal current_step
        if is_running and step_index != 2:
            show_light_hint("正在生成结果，请等待本次填写完成。")
            return
        if step_index == 3 and latest_run_dir is None:
            show_light_hint("还没有生成结果，完成填写后才能查看第 4 步。")
            return
        current_step = max(0, min(step_index, len(pages) - 1))
        for index, page in enumerate(pages):
            if index == current_step:
                page.tkraise()

        for index, badge in enumerate(step_badges):
            badge.configure(style="StepBadgeActive.TLabel" if index <= current_step else "StepBadge.TLabel")
            step_names[index].configure(style="StepNameActive.TLabel" if index == current_step else "StepName.TLabel")
            step_wrappers[index].configure(style="StepWrapActive.TFrame" if index == current_step else "StepWrap.TFrame")
        for index, line in enumerate(connectors):
            line.configure(bg="#0F766E" if index < current_step else "#D9E2EC")

        prev_button.state(["!disabled"] if current_step > 0 else ["disabled"])
        if current_step == 0:
            nav_hint_var.set("先把本周文件拖入大区域，再点“下一步”。")
            next_button.configure(text="下一步", style="Primary.TButton", command=lambda: go_to_step(1))
        elif current_step == 1:
            nav_hint_var.set("确认本周周期和日期。如果没问题，继续下一步。")
            next_button.configure(text="下一步", style="Primary.TButton", command=lambda: go_to_step(2))
        elif current_step == 2:
            nav_hint_var.set("先运行检查，确认无必填缺失后，再开始填写。")
            next_button.configure(text="开始填写", style="Primary.TButton", command=lambda: start_run())
        else:
            nav_hint_var.set("结果已经生成。你可以打开文件，或返回前面重新调整。")
            next_button.configure(text="重新检查", style="Secondary.TButton", command=lambda: go_to_step(2))
        if from_click:
            show_light_hint(f"已切换到“{STEP_TITLES[current_step]}”。")

    def start_run() -> None:
        nonlocal latest_run_dir, is_running
        if is_running:
            show_light_hint("正在生成结果，请等待本次填写完成。")
            return
        role_files, issues = collect_issues()
        required_items, _ = split_preflight_issues(issues)
        if required_items:
            precheck_title_var.set("还有必须文件没补齐，暂时不能开始填写。")
            precheck_detail_var.set("必须补齐：\n- " + "\n- ".join(required_items))
            go_to_step(2)
            messagebox.showwarning("无法开始", "请先补齐必须文件：\n- " + "\n- ".join(required_items))
            return

        if not messagebox.askokcancel(
            "开始前确认",
            (
                "请先关闭本次使用的 Excel 文件：\n"
                f"{WORKFLOW_CLOSE_GUIDANCE[workflow]}\n\n"
                + (
                    "亚马逊提醒：M 列“是否断货”依赖主推款/长尾款分类；新增 SKU 没有既有分类时会留空并提示人工确认。\n\n"
                    if workflow == "amazon" else ""
                )
                + "关闭后点击“确定”开始填写。若文件仍被占用，程序会提示对应文件。"
            ),
        ):
            show_light_hint("已取消开始填写。请关闭相关 Excel 文件后再试。")
            return

        output_dir = Path(output_dir_var.get().strip() or str(ROOT_DIR / "output_gui"))
        output_dir.mkdir(parents=True, exist_ok=True)
        run_dir = output_dir / f"运行结果_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        run_dir.mkdir(parents=True, exist_ok=True)
        staged_input = Path(tempfile.mkdtemp(prefix="excel_fill_gui_input_"))

        copied: set[Path] = set()
        try:
            for path in list(role_files.values()) + imported_file_selection.files:
                if path in copied or not path.exists():
                    continue
                shutil.copy2(path, staged_input / path.name)
                copied.add(path)
        except PermissionError as error:
            shutil.rmtree(staged_input, ignore_errors=True)
            locked_path = Path(error.filename).name if error.filename else "某个 Excel 文件"
            messagebox.showwarning(
                "文件正在被占用",
                f"无法读取或复制：{locked_path}\n\n请关闭该 Excel 文件（以及同目录中可能打开的输出文件）后，再点击“开始填写”。",
            )
            return
        except OSError as error:
            shutil.rmtree(staged_input, ignore_errors=True)
            messagebox.showerror("文件准备失败", f"准备文件时出错：{error}\n\n请确认文件没有被占用，并且路径可访问。")
            return

        for widget in (output_listbox, review_tree):
            if widget is None:
                continue
            if isinstance(widget, tk.Listbox):
                widget.delete(0, tk.END)
            else:
                widget.delete(*widget.get_children())

        latest_run_dir = None
        is_running = True
        result_summary_var.set("正在运行，请稍候……")
        detail_summary_var.set("程序正在识别文件并写入结果。")
        running_var.set("正在识别文件并填写结果，请不要关闭窗口。")
        go_to_step(2)
        prev_button.state(["disabled"])
        next_button.state(["disabled"])

        def worker() -> None:
            gui_values = {label: var.get().strip() for label, var in field_vars.items()}
            gui_values.update(
                {
                    label: "-".join(part.get().strip() for part in parts)
                    for label, parts in date_field_vars.items()
                }
            )
            cli_values = build_cli_values(gui_values)
            try:
                if workflow == "amazon":
                    return_code = run_amazon_inventory_fill(str(staged_input), str(run_dir))
                elif workflow == "ecommerce":
                    return_code = run_ecommerce_fill(str(staged_input), str(run_dir), **cli_values)
                else:
                    return_code = run_kepule_fill(str(staged_input), str(run_dir), **cli_values)
                root.after(0, lambda: handle_run_finished(run_dir, return_code, "", ""))
            except Exception:
                # Capture the traceback while this worker still owns the
                # exception context.  Formatting inside the later Tk callback
                # would otherwise degrade the useful error into “NoneType: None”.
                error_text = traceback.format_exc()
                root.after(0, lambda: handle_run_finished(run_dir, 1, "", error_text))
            finally:
                shutil.rmtree(staged_input, ignore_errors=True)

        Thread(target=worker, daemon=True).start()

    def handle_run_finished(run_dir: Path, return_code: int, stdout: str, stderr: str) -> None:
        nonlocal latest_run_dir, is_running
        is_running = False
        running_var.set("")
        if return_code != 0:
            result_summary_var.set("运行失败，请根据提示补齐文件或调整参数后重试。")
            detail_summary_var.set("本次没有成功生成结果文件。")
            go_to_step(2)
            messagebox.showerror("运行失败", (stderr or stdout or "程序运行失败").strip())
            return

        keep_names = {"建议人工复核.xlsx"}
        for item in list(run_dir.iterdir()):
            if not item.is_file():
                continue
            if item.name in keep_names or item.name.endswith("_已填报.xlsx"):
                continue
            item.unlink(missing_ok=True)

        latest_run_dir = run_dir
        result_summary_var.set(f"已完成填写，结果保存在：{run_dir}")
        generated_files = [item for item in sorted(run_dir.iterdir()) if item.is_file() and item.suffix.lower() == ".xlsx"]
        if output_listbox is not None:
            for item in generated_files:
                output_listbox.insert(tk.END, item.name)
        review_items = read_review_items(run_dir)
        if review_tree is not None:
            for item in review_items:
                review_tree.insert("", "end", values=(item.result_location_text, item.source_location_text, item.field_text, item.reason_text, item.action_text))
        xlsx_count = len([item for item in generated_files if item.suffix.lower() == ".xlsx"])
        threshold_items = [
            item for item in review_items
            if item.field_text == "是否断货" and ("主推款" in item.reason_text or "长尾款" in item.reason_text)
        ]
        detail_summary_var.set(
            f"本次共生成 {xlsx_count} 个 Excel 文件；"
            f"{'没有人工核对项。' if not review_items else f'有 {len(review_items)} 条建议人工核对，请优先查看右侧列表。'}"
            + (f" M 列有 {len(threshold_items)} 个 SKU 需要确认主推款（120天）或长尾款（90天）。" if threshold_items else "")
        )
        go_to_step(3)
        show_light_hint("结果已经生成，可以查看第 4 步。")
        completion_note = (
            f"\n\nM 列提醒：有 {len(threshold_items)} 个 SKU 未知主推/长尾分类，M 列已留空。请在“建议人工复核.xlsx”中查看并手工选择：主推款填 -120，长尾款填 -90。"
            if threshold_items else ""
        )
        messagebox.showinfo("已完成", f"结果文件已生成。\n\n输出目录：\n{run_dir}{completion_note}")

    def _decode_drop_path(item: bytes | str) -> str:
        if isinstance(item, bytes):
            for encoding in ("utf-8", "gbk", sys.getfilesystemencoding()):
                try:
                    return item.decode(encoding)
                except Exception:
                    continue
            return item.decode(errors="ignore")
        return str(item)

    def handle_drop(files) -> None:
        dropped_paths: list[Path] = []
        for item in files:
            text = _decode_drop_path(item).strip().strip("{}").strip('"')
            if not text:
                continue
            path = Path(text)
            if path.is_dir():
                # Dragging the W28 root folder must behave exactly like the
                # "添加文件夹" button: source files may live in subfolders.
                dropped_paths.extend(entry.path for entry in scan_excel_files(path, path, recursive=True).input_files)
            elif path.is_file() and path.suffix.lower() in {".xlsx", ".xlsm", ".xls"} and not path.name.startswith("~$"):
                dropped_paths.append(path)
        if not dropped_paths:
            messagebox.showinfo("拖拽导入", "没有识别到可用的 Excel 文件。")
            return
        added = append_imported_paths(sorted({path.resolve(): path for path in dropped_paths}.values(), key=lambda p: p.name))
        refresh_slots()
        show_light_hint(f"已添加 {added} 个拖入文件；之前已选文件仍会保留。")

    def flush_pending_drop_batches() -> None:
        while pending_drop_batches:
            batch = pending_drop_batches.popleft()
            handle_drop(batch)
        root.after(180, flush_pending_drop_batches)

    if windnd is not None and active_drop_zone is not None:
        def _queue_drop(files) -> None:
            pending_drop_batches.append(list(files))

        windnd.hook_dropfiles(active_drop_zone, func=_queue_drop)
        drop_sub.configure(text="把 Excel 文件拖到这个区域里，程序会自动识别。也可以点击按钮批量选择文件或文件夹。")
    else:
        drop_sub.configure(text="当前环境未启用拖拽支持。你仍然可以点击下方按钮批量选择文件或文件夹。")

    for index, badge in enumerate(step_badges):
        badge.bind("<Button-1>", lambda _e, target=index: go_to_step(target, from_click=True))
        step_names[index].bind("<Button-1>", lambda _e, target=index: go_to_step(target, from_click=True))
        step_wrappers[index].bind("<Button-1>", lambda _e, target=index: go_to_step(target, from_click=True))

    prev_button.configure(command=lambda: go_to_step(current_step - 1))
    refresh_slots()
    run_precheck()
    flush_pending_drop_batches()
    go_to_step(0)
