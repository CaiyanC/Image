from __future__ import annotations

from pathlib import Path

import yaml
from openpyxl import load_workbook

from src.app_paths import config_dir
from src.models import DetectionResult


def _load_role_rules() -> dict:
    config_path = config_dir() / "roles.yaml"
    with config_path.open("r", encoding="utf-8") as fh:
        return yaml.safe_load(fh)


def detect_workbook_role(path: Path) -> DetectionResult:
    rules = _load_role_rules()["roles"]
    wb = load_workbook(path, read_only=True, data_only=True)
    sheetnames = list(wb.sheetnames)
    first_sheet = wb[wb.sheetnames[0]]
    try:
        first_row = next(first_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    except StopIteration:
        first_row = ()
    header_values = [str(value).strip() for value in first_row if value not in (None, "")]
    first_column_samples = []
    for row in first_sheet.iter_rows(min_row=2, max_row=6, values_only=True):
        if row and row[0] not in (None, ""):
            first_column_samples.append(str(row[0]).strip())

    for role, rule in rules.items():
        required = rule.get("required_sheets", [])
        matched = [sheet for sheet in required if sheet in sheetnames]
        if len(matched) == len(required):
            wb.close()
            return DetectionResult(
                role=role,
                status="matched",
                matched_sheets=matched,
                processed_sheets=matched,
                skipped_sheets=[sheet for sheet in sheetnames if sheet not in matched],
                path=path,
            )

    detected = _detect_source_role(sheetnames, header_values, first_column_samples, path)
    wb.close()
    return detected or DetectionResult(role="unknown", status="unmatched", path=path, processed_sheets=[], skipped_sheets=sheetnames)


def _header_contains(header_values: list[str], required_fields: list[str]) -> bool:
    return all(field in header_values for field in required_fields)


def _detection(role: str, path: Path, sheetnames: list[str], processed_sheets: list[str], matched_fields: list[str]) -> DetectionResult:
    return DetectionResult(
        role=role,
        status="matched",
        matched_sheets=processed_sheets,
        processed_sheets=processed_sheets,
        skipped_sheets=[sheet for sheet in sheetnames if sheet not in processed_sheets],
        matched_fields=matched_fields,
        path=path,
    )


def _detect_source_role(sheetnames: list[str], header_values: list[str], first_column_samples: list[str], path: Path) -> DetectionResult | None:
    # Amazon inventory is an independent three-file workflow.  Its weekly
    # workbook has a title in row 1 (and headers in row 2), so identify these
    # two roles by their stable business filename and station-sheet layout
    # before applying the row-1 source-header rules below.
    filename = path.name
    if {"自营京东库存明细", "亚马逊库存"}.issubset(sheetnames):
        return _detection("jd_amazon_inventory", path, sheetnames, ["自营京东库存明细", "亚马逊库存"], [])
    if "亚马逊库存每周更新" in filename and any("亚马逊" in sheet for sheet in sheetnames):
        return _detection("amazon_inventory_weekly", path, sheetnames, sheetnames, ["SKU", "成本", "15天销量", "30天销量", "工厂库存", "FBA仓在途库存"])
    if "亚马逊" in filename and "最新库存明细表" in filename and {"北美站", "日本站"}.issubset(sheetnames):
        return _detection("amazon_inventory_target", path, sheetnames, sheetnames, ["SKU", "成本", "30天销量", "FBA可用库存", "可销售天数"])
    if {"北美", "日本", "欧洲"}.issubset(sheetnames) and _header_contains(header_values, ["店铺", "品名", "SKU", "毛利润", "销量", "销售额"]):
        return _detection("cross_border_profit_sku", path, sheetnames, ["北美", "日本", "欧洲"], ["店铺", "SKU", "销量", "销售额", "退款金额"])
    if {"北美站", "日本站"}.issubset(sheetnames) and _header_contains(header_values, ["SKU", "品名", "成本", "30天销量", "FBA可用库存", "可销售天数"]):
        return _detection("amazon_latest_inventory", path, sheetnames, ["北美站", "日本站", "欧洲站-英国仓", "欧洲站--德国仓"], ["SKU", "品名", "成本", "30天销量", "FBA可用库存", "可销售天数"])
    if ({"北美", "北美站"} & set(sheetnames)) and ({"日本", "日本站"} & set(sheetnames)) and _header_contains(header_values, ["店铺", "SKU", "品名", "FBA可用库存"]):
        return _detection("fba_inventory", path, sheetnames, sheetnames, ["店铺", "SKU", "品名", "FBA可用库存"])
    if _header_contains(header_values, ["时间", "商品名称", "SKU", "69码", "成交商品件数", "成交金额", "匹配成本价"]):
        return _detection("jd_self_weekly_sales", path, sheetnames, [sheetnames[0]], ["69码", "成交商品件数", "成交金额", "匹配成本价", "总可用库存"])
    if _header_contains(header_values, ["商品编码", "商品名", "产品分类", "商品标签", "实际可用数", "7天销量", "月销量"]):
        return _detection("product_archive", path, sheetnames, [sheetnames[0]], ["商品编码", "商品名", "产品分类", "实际可用数", "7天销量", "月销量"])
    if "30天" in filename and _header_contains(header_values, ["店铺", "商品编码", "款式编码", "商品名称", "产品分类", "销售数量"]):
        return _detection("sales_30d", path, sheetnames, [sheetnames[0]], ["店铺", "商品编码", "商品名称", "产品分类", "销售数量"])
    if _header_contains(header_values, ["店铺", "商品编码", "商品名称", "产品分类", "销售数量", "销售金额", "净销量", "净销售额"]):
        store_blob = " ".join(first_column_samples)
        role = "domestic_sales_theme_analysis" if "档口" in store_blob else "sales_theme_analysis"
        return _detection(role, path, sheetnames, [sheetnames[0]], ["店铺", "商品编码", "商品名称", "产品分类", "销售数量", "销售金额"])
    if _header_contains(header_values, ["店铺", "商品编码", "款式编码", "商品名称", "产品分类", "销售数量"]):
        return _detection("sales_30d", path, sheetnames, [sheetnames[0]], ["店铺", "商品编码", "商品名称", "产品分类", "销售数量"])
    if _header_contains(header_values, ["商品编码", "商品名称", "净销量", "净销售额", "净销售成本"]):
        return _detection("domestic_sales_ranking", path, sheetnames, [sheetnames[0]], ["商品编码", "商品名称", "净销量", "净销售额"])
    return None
