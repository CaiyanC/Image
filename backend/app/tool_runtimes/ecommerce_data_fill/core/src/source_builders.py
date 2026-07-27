from __future__ import annotations

import re
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook

from src.excel_utils import build_sheet_context
from src.models import BuildResult, Issue
from src.shop_mapping import resolve_shop


STYLE_SKU_PATTERN = re.compile(
    r"(?<![A-Z0-9])(?:[A-Z]{1,4}\d+[A-Z0-9]*(?:-[A-Z0-9]+)*|[A-Z]{2,}(?:-[A-Z0-9]+)+)(?![A-Z0-9（(])",
    flags=re.IGNORECASE,
)


def parse_number(value):
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", "")
    if text.endswith("%"):
        try:
            return float(text[:-1]) / 100
        except ValueError:
            return None
    for symbol in ["JPY", "£", "€", "$"]:
        text = text.replace(symbol, "").strip()
    try:
        return float(text)
    except ValueError:
        return None


def inventory_age_days(first_inbound, inventory_date: str) -> int | None:
    """Calculate the completed snapshot's inclusive/exclusive age convention."""
    if first_inbound in (None, ""):
        return None
    if isinstance(first_inbound, datetime):
        inbound_day = first_inbound.date()
    elif isinstance(first_inbound, date):
        inbound_day = first_inbound
    else:
        match = re.search(r"(\d{4})[/-](\d{1,2})[/-](\d{1,2})", str(first_inbound))
        if not match:
            return None
        inbound_day = date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    snapshot_day = datetime.strptime(inventory_date, "%Y-%m-%d").date()
    return max((snapshot_day - inbound_day).days - 1, 0)


def inventory_type_from_age(days: int | float | None) -> str | None:
    """Default type; a manually entered 滞销 remains an explicit override."""
    if days is None:
        return None
    return "长库龄" if days > 540 else "正常"


def week_label_from_cycle(cycle_code: str) -> str:
    match = re.fullmatch(r"\d{4}(W\d{1,2})", str(cycle_code or "").strip(), flags=re.IGNORECASE)
    return match.group(1).upper() if match else cycle_code


def load_rows(path: Path, sheet_name: str | None, required_fields: list[str], optional_fields: list[str] | None = None) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]
    context = build_sheet_context(ws, required_fields)
    optional_columns = {field: column for field, column in context.columns.items() if field in (optional_fields or [])}
    rows = []
    for row_number, values in enumerate(ws.iter_rows(min_row=context.data_start_row, values_only=True), start=context.data_start_row):
        row = {}
        has_value = False
        for field, col in context.columns.items():
            value = values[col - 1] if col - 1 < len(values) else None
            row[field] = value
            if value not in (None, ""):
                has_value = True
        for field, col in optional_columns.items():
            value = values[col - 1] if col - 1 < len(values) else None
            row[field] = value
        if has_value:
            row["__row_number"] = row_number
            row["__sheet_name"] = ws.title
            rows.append(row)
    wb.close()
    return rows


def _blank_stats() -> dict[str, int]:
    return {
        "raw_candidate_count": 0,
        "skip_zero_sales_count": 0,
        "skip_no_effect_count": 0,
        "candidate_after_filter_count": 0,
        "product_archive_candidate_count": 0,
        "sales30_candidate_count": 0,
        "jd_inventory_candidate_count": 0,
        "amazon_inventory_candidate_count": 0,
        "shop_sku_fallback_count": 0,
    }


def _first_number(*values):
    for value in values:
        if value is not None:
            return value
    return None


def _numbers_differ(left, right, tolerance: float = 1e-6) -> bool:
    if left is None or right is None:
        return False
    return abs(float(left) - float(right)) > tolerance


def build_domestic_sales_ranking_index(path: Path) -> dict[str, dict]:
    rows = load_rows(
        path,
        None,
        ["商品编码", "商品名称"],
        optional_fields=["净销量", "净销售额", "净销售成本", "销售数量", "销售金额", "销售成本"],
    )
    grouped: dict[str, dict] = {}
    numeric_fields = ["净销量", "净销售额", "净销售成本", "销售数量", "销售金额", "销售成本"]
    for row in rows:
        if not row.get("商品编码"):
            continue
        sku = str(row["商品编码"]).strip()
        bucket = grouped.setdefault(sku, {**row, **{field: 0.0 for field in numeric_fields}})
        for field in numeric_fields:
            value = parse_number(row.get(field))
            if value is not None:
                bucket[field] += value
    return grouped


def _inventory_region(value: str) -> str:
    text = str(value or "").lower()
    if any(token in text for token in ["北美", "kingpool-us", "-us", "美国", "加拿大"]):
        return "north_america"
    if any(token in text for token in ["日本", "-jp", "jp-"]):
        return "japan"
    if any(token in text for token in ["英国", "-uk", "uk-"]):
        return "uk"
    if any(token in text for token in ["德国", "-de", "de-"]):
        return "germany"
    return ""


def build_fba_inventory_index(path: Path) -> dict[tuple[str, str], dict]:
    index: dict[tuple[str, str], dict] = {}
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            try:
                context = build_sheet_context(ws, ["SKU", "品名"])
            except ValueError:
                continue
            for row_number, values in enumerate(ws.iter_rows(min_row=context.data_start_row, values_only=True), start=context.data_start_row):
                sku = values[context.columns["SKU"] - 1] if context.columns["SKU"] - 1 < len(values) else None
                if not sku or "含停售" in sheet_name:
                    continue
                row = {
                    field: values[column - 1] if column - 1 < len(values) else None
                    for field, column in context.columns.items()
                }
                row["__sheet_name"] = sheet_name
                row["__row_number"] = row_number
                region = _inventory_region(sheet_name) or _inventory_region(row.get("店铺", ""))
                if region:
                    index[(region, str(sku).strip())] = row
    finally:
        workbook.close()
    return index


def build_product_archive_index(path: Path) -> dict[str, dict]:
    rows = load_rows(path, None, ["商品编码", "商品名", "产品分类", "实际可用数", "采购在途", "7天销量", "月销量"])
    return {str(row["商品编码"]).strip(): row for row in rows if row.get("商品编码")}


def build_domestic_ad_allocation_index(path: Path) -> dict[tuple[str, str], float]:
    """Read the optional two-column allocation block after 销售成本.

    Some processed weekly exports add two intentionally headerless columns:
    sales-share and allocated advertising spend. The legacy W27 export has no
    such gap, so it safely yields an empty index.
    """
    index: dict[tuple[str, str], float] = {}
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            try:
                context = build_sheet_context(ws, ["店铺", "商品编码", "销售成本"])
            except ValueError:
                continue
            cost_column = context.columns["销售成本"]
            allocated_column = cost_column + 2
            if allocated_column > ws.max_column:
                continue
            # Both allocation columns must be structurally headerless. This
            # prevents reading ordinary named business fields by position.
            if ws.cell(context.header_row, cost_column + 1).value not in (None, ""):
                continue
            if ws.cell(context.header_row, allocated_column).value not in (None, ""):
                continue
            for row_number in range(context.data_start_row, ws.max_row + 1):
                store = str(ws.cell(row_number, context.columns["店铺"]).value or "").strip()
                sku = str(ws.cell(row_number, context.columns["商品编码"]).value or "").strip()
                allocated = parse_number(ws.cell(row_number, allocated_column).value)
                if store and sku and allocated is not None:
                    index[(store, sku)] = allocated
    finally:
        workbook.close()
    return index


def build_jd_refund_allocation_index(path: Path) -> dict[str, float]:
    """Read the optional headerless JD refund allocation result column."""
    index: dict[str, float] = {}
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            try:
                context = build_sheet_context(ws, ["69码", "成交金额"])
            except ValueError:
                continue
            base_column = context.columns.get("全国近30天出库销量") or context.columns.get("30天")
            if base_column is None or base_column + 2 > ws.max_column:
                continue
            if ws.cell(context.header_row, base_column + 1).value not in (None, ""):
                continue
            if ws.cell(context.header_row, base_column + 2).value not in (None, ""):
                continue
            for row_number in range(context.data_start_row, ws.max_row + 1):
                sku = str(ws.cell(row_number, context.columns["69码"]).value or "").strip()
                allocated = parse_number(ws.cell(row_number, base_column + 2).value)
                if sku and allocated is not None:
                    index[sku] = allocated
    finally:
        workbook.close()
    return index


def _canonical_product_category(value: object) -> str:
    """Normalize archive-only category labels to the target's category vocabulary."""
    text = str(value or "").strip()
    if text in {"分类-未设置", "虚拟分类-未设置", "未设置"}:
        return ""
    if text == "套锅/户外炊具":
        return "锅具"
    return text


def build_product_category_index(path: Path, style_source_path: Path | None = None) -> dict[str, str]:
    """Index only deterministic product-code/style-code category matches.

    Cross-border sales exports use style SKUs while the product archive often
    stores an Amazon-specific product code.  A style code is safe to use only
    when every archive row carrying it has the same normalized category.
    """
    rows = load_rows(path, None, ["商品编码", "商品名", "产品分类", "实际可用数", "采购在途", "7天销量", "月销量"])
    candidates: dict[str, set[str]] = defaultdict(set)
    for row in rows:
        category = _canonical_product_category(row.get("产品分类"))
        if not category:
            continue
        keys = [str(row.get(field) or "").strip().upper() for field in ("商品编码", "款式编码")]
        product_name = str(row.get("商品名") or "")
        compact_version_name = re.sub(
            r"(?<=[A-Z0-9-])\s+(?=(?:PRO|PLUS)\b)",
            "",
            product_name,
            flags=re.IGNORECASE,
        )
        # Amazon archive product codes are often opaque ASIN-like values, but
        # their standard product names contain the actual style SKU used by
        # the order-profit source. Extract only explicit SKU-shaped spans and
        # still require category uniqueness before exposing the alias.
        keys.extend(match.upper() for match in STYLE_SKU_PATTERN.findall(compact_version_name))
        for key in keys:
            if key:
                candidates[key].add(category)
    # The product archive usually identifies Amazon products with a platform
    # code, while sales exports use the style SKU.  The domestic 30-day export
    # contains the deterministic bridge 款式编码 -> 商品编码.  Only retain a
    # style when every linked archive product agrees on the category.
    if style_source_path is not None:
        for row in load_rows(style_source_path, None, ["商品编码", "商品名称"], optional_fields=["款式编码"]):
            style_key = str(row.get("款式编码") or "").strip().upper()
            product_key = str(row.get("商品编码") or "").strip().upper()
            if not style_key or not product_key:
                continue
            for category in candidates.get(product_key, set()):
                candidates[style_key].add(category)
    return {key: next(iter(categories)) for key, categories in candidates.items() if len(categories) == 1}


def resolve_product_category(category_index: dict[str, str], sku: object) -> str:
    """Resolve an exact style first, then a uniquely categorized base style.

    Cross-border files often add a regional/version suffix (for example
    ``CS-B14-DFB``) while the archive records the base model.  A fallback is
    safe only when every matching base/prefix alias points to one category.
    """
    key = str(sku or "").strip().upper()
    if not key:
        return ""
    if key in category_index:
        return category_index[key]

    categories: set[str] = set()
    parts = key.split("-")
    for length in range(len(parts) - 1, 1, -1):
        base = "-".join(parts[:length])
        if base in category_index:
            categories.add(category_index[base])
    prefix = f"{key}-"
    categories.update(category for style, category in category_index.items() if style.startswith(prefix))
    return next(iter(categories)) if len(categories) == 1 else ""


def build_domestic_style_sku_index(path: Path) -> tuple[dict[str, str], set[str]]:
    """Read the optional style-code column from the domestic 30-day sales export."""
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    context = build_sheet_context(ws, ["商品编码", "商品名称"])
    style_column = context.columns.get("款式编码")
    if style_column is None:
        wb.close()
        return {}, set()

    styles_by_code: dict[str, set[str]] = {}
    known_styles: set[str] = set()
    for values in ws.iter_rows(min_row=context.data_start_row, values_only=True):
        product_code = str(values[context.columns["商品编码"] - 1] or "").strip()
        style_sku = str(values[style_column - 1] or "").strip().upper()
        if not product_code or not style_sku:
            continue
        styles_by_code.setdefault(product_code, set()).add(style_sku)
        known_styles.add(style_sku)
    wb.close()

    return (
        {product_code: next(iter(styles)) for product_code, styles in styles_by_code.items() if len(styles) == 1},
        known_styles,
    )


def build_domestic_supplier_index(path: Path) -> dict[str, str]:
    """Map a product code to its supplier from the domestic sales export."""
    rows = load_rows(path, None, ["商品编码", "商品名称"])
    suppliers: dict[str, str] = {}
    for row in rows:
        product_code = str(row.get("商品编码") or "").strip()
        supplier = str(row.get("供应商") or "").strip()
        if product_code and supplier and product_code not in suppliers:
            suppliers[product_code] = supplier
    return suppliers


def resolve_domestic_inventory_sku(
    product_code: str,
    product_name: object,
    style_by_product_code: dict[str, str],
    known_styles: set[str],
) -> tuple[str | None, str]:
    """Resolve a domestic inventory SKU, retaining a barcode only when no unique SKU is available."""
    direct_style = style_by_product_code.get(product_code)
    if direct_style:
        return direct_style, "款式编码"

    name = re.sub(r"-\d+(?:\.\d+)?L", "", str(product_name or ""), flags=re.IGNORECASE)
    candidates = {match.group(0).upper() for match in STYLE_SKU_PATTERN.finditer(name)}
    verified_candidates = candidates & known_styles
    if len(verified_candidates) == 1:
        return next(iter(verified_candidates)), "商品名匹配款式编码"
    if len(candidates) == 1:
        return next(iter(candidates)), "商品名提取"
    return None, ""


def build_special_sku_audit(role_files: dict[str, Path], sku: str) -> dict[str, str]:
    result = {
        "SKU编码": sku,
        "来源文件": "",
        "来源Sheet": "",
        "来源行号": "",
        "店铺": "",
        "商品编码": sku,
        "商品名称": "",
        "销售数量": "",
        "净销量": "",
        "销售金额": "",
        "净销售额": "",
        "销售成本": "",
        "净销售成本": "",
        "成本价": "",
        "最终写入销售额取自哪一列": "销售金额",
        "最终写入成交件数取自哪一列": "销售数量",
        "最终写入单件成本取自哪一列": "销售成本/销售数量",
        "判断": "需要人工复核",
    }
    if "sales_theme_analysis" not in role_files:
        return result

    source_path = role_files["sales_theme_analysis"]
    for row in load_rows(source_path, None, ["店铺", "商品编码", "商品名称", "销售数量", "净销量", "销售金额", "净销售额", "销售成本", "净销售成本", "成本价"]):
        if str(row.get("商品编码") or "").strip() != sku:
            continue
        result.update(
            {
                "来源文件": source_path.name,
                "来源Sheet": row.get("__sheet_name", "Sheet1"),
                "来源行号": str(row.get("__row_number", "")),
                "店铺": str(row.get("店铺") or ""),
                "商品名称": str(row.get("商品名称") or ""),
                "销售数量": "" if row.get("销售数量") is None else str(row.get("销售数量")),
                "净销量": "" if row.get("净销量") is None else str(row.get("净销量")),
                "销售金额": "" if row.get("销售金额") is None else str(row.get("销售金额")),
                "净销售额": "" if row.get("净销售额") is None else str(row.get("净销售额")),
                "销售成本": "" if row.get("销售成本") is None else str(row.get("销售成本")),
                "净销售成本": "" if row.get("净销售成本") is None else str(row.get("净销售成本")),
                "成本价": "" if row.get("成本价") is None else str(row.get("成本价")),
            }
        )
        sales_amount = parse_number(row.get("销售金额"))
        sales_qty = parse_number(row.get("销售数量"))
        sales_cost = parse_number(row.get("销售成本"))
        if sales_qty not in (None, 0) and sales_cost is not None:
            implied_unit_cost = sales_cost / sales_qty
            result["成本价"] = str(implied_unit_cost)
        if sales_amount not in (None, 0) and sales_qty not in (None, 0) and sales_cost not in (None, 0):
            avg_price = sales_amount / sales_qty
            unit_cost = sales_cost / sales_qty
            result["判断"] = "源表真实数据" if avg_price < unit_cost else "需要人工复核"
        return result
    return result


def build_amazon_cost_index(path: Path) -> dict[str, float]:
    wb = load_workbook(path, read_only=True, data_only=True)
    result = {}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        context = build_sheet_context(ws, ["SKU", "品名", "成本", "30天销量", "FBA可用库存"])
        for row_index in range(context.data_start_row, ws.max_row + 1):
            sku = ws.cell(row_index, context.columns["SKU"]).value
            if not sku:
                continue
            cost = parse_number(ws.cell(row_index, context.columns["成本"]).value)
            if cost is not None:
                result[str(sku).strip()] = cost
    wb.close()
    return result


def build_jd_inventory_index(path: Path) -> dict[str, dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["自营京东库存明细"]
    context = build_sheet_context(ws, ["条码", "商品简称", "全国近7天出库销量", "全国近30天出库销量", "总可用库存", "采购在途", "成本单价"])
    result = {}
    for row_index in range(context.data_start_row, ws.max_row + 1):
        sku = ws.cell(row_index, context.columns["条码"]).value
        if not sku:
            continue
        sku = str(sku).strip()
        result[sku] = {
            "商品名称": ws.cell(row_index, context.columns["商品简称"]).value,
            "在库库存": parse_number(ws.cell(row_index, context.columns["总可用库存"]).value),
            "在途库存": parse_number(ws.cell(row_index, context.columns["采购在途"]).value),
            "近7天销量": parse_number(ws.cell(row_index, context.columns["全国近7天出库销量"]).value),
            "近30天销量": parse_number(ws.cell(row_index, context.columns["全国近30天出库销量"]).value),
            "单件成本": parse_number(ws.cell(row_index, context.columns["成本单价"]).value),
            "sheet": ws.title,
            "file": path.name,
        }
    wb.close()
    return result


def build_w27_sku_rows(role_files: dict[str, Path], issues: list[Issue], cycle_code: str, shop_config: dict) -> BuildResult:
    result = BuildResult(stats=_blank_stats())
    archive_index = build_product_archive_index(role_files["product_archive"]) if "product_archive" in role_files else {}
    allowed_platforms = {"天猫", "京东POP", "拼多多"}

    if "sales_theme_analysis" in role_files:
        source_path = role_files["sales_theme_analysis"]
        for row in load_rows(
            source_path,
            None,
            [
                "店铺", "商品编码", "商品名称", "产品分类", "销售数量", "销售金额", "销售成本",
                "退货金额",
            ],
            optional_fields=["退货数量", "退货单数"],
        ):
            result.stats["raw_candidate_count"] += 1
            store = str(row["店铺"]).strip()
            normalized = resolve_shop(store, shop_config)
            if normalized["platform"] not in allowed_platforms:
                continue
            sku = str(row["商品编码"]).strip()
            sales_amount = parse_number(row.get("销售金额"))
            sales_qty = parse_number(row.get("销售数量"))
            # W27's completed workbook excludes rows with no monetary sales
            # impact even when an operational quantity is present.
            if sales_amount in (None, 0):
                result.stats["skip_zero_sales_count"] += 1
                issues.append(Issue("INFO", "builder", "sales_theme_analysis", "本周期无销售，未写入 3_SKU明细", source_path.name, "Sheet1", "", sku, "", "W27", "3_SKU明细", "保留库存补货候选即可。"))
                result.audit_rows.append(
                    {
                        "目标Sheet": "3_SKU明细",
                        "平台": normalized["platform"],
                        "店铺名称": normalized["target_store_name"],
                        "SKU编码": sku,
                        "商品名称": row.get("商品名称") or sku,
                        "原目标表是否已有": "",
                        "本次是否写入": "否",
                        "差异类型": "被过滤跳过",
                        "来源文件": source_path.name,
                        "来源Sheet": row.get("__sheet_name", "Sheet1"),
                        "来源字段": "销售金额/销售数量",
                        "关键值": f"销售金额={sales_amount}; 成交件数={sales_qty}",
                        "说明": "本周期无销售，未写入 3_SKU明细",
                    }
                )
                continue
            archive = archive_index.get(sku, {})
            sales_cost = parse_number(row.get("销售成本"))
            # Detail-sheet unit cost follows the actual sales source. The
            # product archive only provides a fallback when sales cost is
            # absent, because promotional/allocation cost is reflected here.
            unit_cost = parse_number(row.get("成本价"))
            if unit_cost is None and sales_qty not in (None, 0) and sales_cost is not None:
                unit_cost = sales_cost / sales_qty
            if unit_cost is None:
                unit_cost = parse_number(archive.get("成本价"))
            category_from_sales = str(row.get("产品分类") or "").strip()
            category_from_archive = str(archive.get("产品分类") or "").strip()
            product_name = row.get("商品名称") or archive.get("商品名") or sku
            product_category = category_from_sales or category_from_archive or ""
            if not category_from_sales and not category_from_archive:
                issues.append(Issue("WARNING", "builder", "sales_theme_analysis", "商品分类缺失", source_path.name, "Sheet1", "", sku, "商品分类", "W27", "商品分类", "补充商品档案或源表分类字段。"))
            if unit_cost is None:
                issues.append(Issue("WARNING", "builder", "sales_theme_analysis", "单件成本缺失", source_path.name, "Sheet1", "", sku, "单件成本", "W27", "单件成本", "补充成本来源后重跑。"))
            if sales_amount not in (None, 0) and sales_qty not in (None, 0) and unit_cost not in (None, 0):
                avg_price = sales_amount / sales_qty
                if avg_price < unit_cost:
                    issues.append(Issue("WARNING", "builder", "sales_theme_analysis", "销售额/成交件数/单件成本组合异常，请人工复核", source_path.name, row.get("__sheet_name", "Sheet1"), str(row.get("__row_number", "")), sku, "销售额/成交件数/单件成本", "W27", "3_SKU明细", "检查源表是否为真实异常或字段口径异常。"))
            return_quantity = parse_number(row.get("退货数量"))
            has_real_return_order = (
                return_quantity is not None
                and return_quantity >= 1
                and abs(return_quantity - round(return_quantity)) < 1e-9
            )
            payload = {
                "平台": normalized["platform"],
                "店铺名称": normalized["target_store_name"],
                "商品名称": product_name,
                "SKU编码": sku,
                "商品分类": product_category,
                "销售额": sales_amount,
                "成交件数": sales_qty,
                # W28's completed workbook establishes the rule: small
                # monetary adjustments with zero return quantity are not a
                # refund order and must remain blank in SKU detail.
                # Preserve the legacy value when a source schema does not
                # provide return quantity at all.  When it does, apply the
                # W28-verified order gate.
                "退款金额": (
                    parse_number(row.get("退货金额"))
                    if return_quantity is None or has_real_return_order
                    else None
                ),
                "单件成本": unit_cost,
            }
            return_order_count = parse_number(row.get("退货单数"))
            if has_real_return_order and return_order_count is not None:
                payload["退款订单数"] = return_order_count
            result.rows.append(payload)
            result.source_lookup[(cycle_code, payload["平台"], payload["店铺名称"], sku)] = (source_path.name, row.get("__sheet_name", "Sheet1"))
            if not category_from_sales and category_from_archive:
                result.audit_rows.append(
                    {
                        "目标Sheet": "3_SKU明细",
                        "平台": payload["平台"],
                        "店铺名称": payload["店铺名称"],
                        "SKU编码": sku,
                        "商品名称": product_name,
                        "原目标表是否已有": "",
                        "本次是否写入": "是",
                        "差异类型": "字段使用补充来源",
                        "来源文件": role_files.get("product_archive", Path("")).name,
                        "来源Sheet": archive.get("__sheet_name", "Sheet1"),
                        "来源字段": "商品分类优先取销售源；当销售源分类缺失且商品档案存在同SKU记录时，使用商品档案分类补充",
                        "关键值": product_category,
                        "说明": "商品档案分类补充来源。",
                    }
                )

    if "jd_self_weekly_sales" in role_files:
        source_path = role_files["jd_self_weekly_sales"]
        for row in load_rows(source_path, None, ["商品名称", "69码", "成交商品件数", "成交金额", "匹配成本价"], optional_fields=["成本总价", "供货总价"]):
            result.stats["raw_candidate_count"] += 1
            sku = str(row["69码"]).strip()
            # The completed W27/W28 reports use the JD self-operated supply
            # total as the finance sales amount.  The consumer transaction
            # amount remains a fallback only when that source field is absent.
            sales_amount = _first_number(
                parse_number(row.get("供货总价")),
                parse_number(row.get("成交金额")),
            )
            sales_qty = parse_number(row.get("成交商品件数"))
            if (sales_amount in (None, 0)) and (sales_qty in (None, 0)):
                result.stats["skip_zero_sales_count"] += 1
                issues.append(Issue("INFO", "builder", "jd_self_weekly_sales", "本周期无销售，未写入 3_SKU明细", source_path.name, "经营状况-商品明细周报", "", sku, "", "W27", "3_SKU明细", "保留库存补货候选即可。"))
                result.audit_rows.append(
                    {
                        "目标Sheet": "3_SKU明细",
                        "平台": "京东自营",
                        "店铺名称": "京东自营",
                        "SKU编码": sku,
                        "商品名称": row.get("商品名称") or sku,
                        "原目标表是否已有": "",
                        "本次是否写入": "否",
                        "差异类型": "被过滤跳过",
                        "来源文件": source_path.name,
                        "来源Sheet": row.get("__sheet_name", "经营状况-商品明细周报"),
                        "来源字段": "成交金额/成交商品件数",
                        "关键值": f"销售金额={sales_amount}; 成交件数={sales_qty}",
                        "说明": "本周期无销售，未写入 3_SKU明细",
                    }
                )
                continue
            archive = archive_index.get(sku, {})
            matched_unit_cost = parse_number(row.get("匹配成本价"))
            total_cost = parse_number(row.get("成本总价"))
            unit_cost = matched_unit_cost
            if unit_cost is None and sales_qty not in (None, 0) and total_cost is not None:
                unit_cost = total_cost / sales_qty
            payload = {
                "平台": "京东自营",
                "店铺名称": "京东自营",
                # JD's weekly export uses consumer-facing long titles.  The
                # W27 completed workbook uses the product-archive standard
                # name instead, keeping this column stable across platforms.
                "商品名称": archive.get("商品名") or row.get("商品名称") or sku,
                "SKU编码": sku,
                "商品分类": archive.get("产品分类") or "",
                "销售额": sales_amount,
                "成交件数": sales_qty,
                "退款金额": None,
                "单件成本": unit_cost,
            }
            if not payload["商品分类"]:
                issues.append(Issue("WARNING", "builder", "jd_self_weekly_sales", "商品分类缺失", source_path.name, "经营状况-商品明细周报", "", sku, "商品分类", "W27", "商品分类", "补充商品档案后重跑。"))
            if payload["单件成本"] is None:
                issues.append(Issue("WARNING", "builder", "jd_self_weekly_sales", "单件成本缺失", source_path.name, "经营状况-商品明细周报", "", sku, "单件成本", "W27", "单件成本", "补充成本来源后重跑。"))
            result.rows.append(payload)
            result.source_lookup[(cycle_code, payload["平台"], payload["店铺名称"], sku)] = (source_path.name, row.get("__sheet_name", "经营状况-商品明细周报"))
            if not archive.get("产品分类"):
                continue
            if not row.get("商品名称") and archive.get("商品名"):
                result.audit_rows.append(
                    {
                        "目标Sheet": "3_SKU明细",
                        "平台": payload["平台"],
                        "店铺名称": payload["店铺名称"],
                        "SKU编码": sku,
                        "商品名称": payload["商品名称"],
                        "原目标表是否已有": "",
                        "本次是否写入": "是",
                        "差异类型": "字段使用补充来源",
                        "来源文件": role_files.get("product_archive", Path("")).name,
                        "来源Sheet": archive.get("__sheet_name", "Sheet1"),
                        "来源字段": "商品名称=当前来源字段优先；若缺失，则按同SKU从补充来源补齐",
                        "关键值": payload["商品名称"],
                        "说明": "商品名称来自商品档案补充来源。",
                    }
                )

    result.stats["candidate_after_filter_count"] = len(result.rows)
    return result


def build_w27_inventory_rows(
    role_files: dict[str, Path],
    issues: list[Issue],
    cycle_code: str,
    shop_config: dict,
    existing_target_rows: dict[tuple[str, ...], dict[str, str]] | None = None,
    allowed_sku_keys: set[tuple[str, str, str]] | None = None,
) -> BuildResult:
    archive_index = build_product_archive_index(role_files["product_archive"]) if "product_archive" in role_files else {}
    sales30_index: dict[tuple[str, str], dict] = {}
    sales30_by_sku: dict[str, set[str]] = {}
    sales30_total_by_sku: dict[str, float] = {}
    weekly_sales_by_shop_sku: dict[tuple[str, str], float] = {}
    jd_index: dict[str, dict] = {}
    amazon_index: dict[str, dict] = {}
    jd_inventory_index: dict[str, dict] = {}
    result = BuildResult(stats=_blank_stats())
    candidates: dict[tuple[str, str, str], dict] = {}
    existing_target_keys = {
        (key[1], key[2], key[3])
        for key in (existing_target_rows or {}).keys()
        if key and key[0] == cycle_code
    }

    def ensure_candidate(platform: str, store_name: str, sku: str, product_name: str, source_tag: str) -> None:
        key = (platform, store_name, sku)
        if allowed_sku_keys is not None and key not in allowed_sku_keys:
            return
        candidate = candidates.setdefault(
            key,
            {
                "平台": platform,
                "店铺名称": store_name,
                "SKU编码": sku,
                "商品名称": product_name,
                "sources": set(),
            },
        )
        if product_name and not candidate["商品名称"]:
            candidate["商品名称"] = product_name
        candidate["sources"].add(source_tag)

    if "sales_30d" in role_files:
        source_path = role_files["sales_30d"]
        for row in load_rows(
            source_path,
            None,
            ["店铺", "商品编码", "商品名称", "销售数量"],
            optional_fields=["净销量"],
        ):
            raw_store = str(row.get("店铺") or "").strip()
            normalized = resolve_shop(raw_store, shop_config)
            sku = str(row["商品编码"]).strip()
            key = (normalized["target_store_name"], sku)
            bucket = sales30_index.setdefault(
                key,
                {
                    "店铺名称": normalized["target_store_name"],
                    "平台": normalized["platform"],
                    "商品名称": row.get("商品名称") or "",
                    "净销量": 0.0,
                    "销售数量": 0.0,
                    "has_net": False,
                    "has_sales_qty": False,
                    "30天销量": None,
                    "has_direct_30d": "30天销量" in row,
                    "实际可用数": None,
                    "采购在途": None,
                    "has_actual_available": "实际可用数" in row,
                    "has_purchase_transit": "采购在途" in row,
                    "__sheet_name": row.get("__sheet_name", "Sheet1"),
                },
            )
            net_qty = parse_number(row.get("净销量"))
            sales_qty = parse_number(row.get("销售数量"))
            direct_30d = parse_number(row.get("30天销量"))
            direct_available = parse_number(row.get("实际可用数"))
            direct_transit = parse_number(row.get("采购在途"))
            if net_qty is not None:
                bucket["净销量"] += net_qty
                bucket["has_net"] = True
            if sales_qty is not None:
                bucket["销售数量"] += sales_qty
                bucket["has_sales_qty"] = True
                sales30_total_by_sku[sku] = sales30_total_by_sku.get(sku, 0.0) + sales_qty
            if direct_30d is not None:
                bucket["30天销量"] = direct_30d
            if bucket["has_actual_available"]:
                bucket["实际可用数"] = direct_available
            if bucket["has_purchase_transit"]:
                bucket["采购在途"] = direct_transit
            sales30_by_sku.setdefault(sku, set()).add(normalized["target_store_name"])
            if existing_target_rows is None:
                ensure_candidate(normalized["platform"], normalized["target_store_name"], sku, row.get("商品名称") or "", "sales30")

    if "sales_theme_analysis" in role_files:
        source_path = role_files["sales_theme_analysis"]
        for row in load_rows(source_path, None, ["店铺", "商品编码", "商品名称"], optional_fields=["销售数量"]):
            normalized = resolve_shop(str(row.get("店铺") or "").strip(), shop_config)
            sku = str(row["商品编码"]).strip()
            ensure_candidate(normalized["platform"], normalized["target_store_name"], sku, row.get("商品名称") or "", "sales_source")
            weekly_quantity = parse_number(row.get("销售数量"))
            if weekly_quantity is not None:
                weekly_key = (normalized["target_store_name"], sku)
                weekly_sales_by_shop_sku[weekly_key] = weekly_sales_by_shop_sku.get(weekly_key, 0.0) + weekly_quantity

    if "jd_self_weekly_sales" in role_files:
        source_path = role_files["jd_self_weekly_sales"]
        for row in load_rows(
            source_path,
            None,
            ["69码", "总可用库存"],
            optional_fields=["30天", "全国近30天出库销量", "成交商品件数"],
        ):
            sku = str(row["69码"]).strip()
            jd_index[sku] = row
            ensure_candidate("京东自营", "京东自营", sku, row.get("商品名称") or "", "jd_weekly")

    if existing_target_rows:
        for key, row in existing_target_rows.items():
            if not key or key[0] != cycle_code:
                continue
            ensure_candidate(key[1], key[2], key[3], row.get("商品名称", ""), "existing_target")

    if "jd_amazon_inventory" in role_files:
        jd_inventory_index = build_jd_inventory_index(role_files["jd_amazon_inventory"])
        for sku, row in jd_inventory_index.items():
            ensure_candidate("京东自营", "京东自营", sku, row.get("商品名称") or "", "jd_inventory")

    if "amazon_latest_inventory" in role_files:
        source_path = role_files["amazon_latest_inventory"]
        wb = load_workbook(source_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            context = build_sheet_context(ws, ["SKU", "品名", "30天销量", "FBA可用库存", "FBA仓在途库存", "可用的总库存"])
            for row_index in range(context.data_start_row, ws.max_row + 1):
                sku = ws.cell(row_index, context.columns["SKU"]).value
                if not sku:
                    continue
                amazon_index[str(sku).strip()] = {
                    "商品名称": ws.cell(row_index, context.columns["品名"]).value,
                    "在库库存": parse_number(ws.cell(row_index, context.columns["可用的总库存"]).value),
                    "在途库存": parse_number(ws.cell(row_index, context.columns["FBA仓在途库存"]).value),
                    "近30天销量": parse_number(ws.cell(row_index, context.columns["30天销量"]).value),
                    "sheet": sheet_name,
                    "file": source_path.name,
                }
        wb.close()

    result.stats["raw_candidate_count"] = len(candidates)
    result.stats["sales30_candidate_count"] = len(
        [
            1
            for candidate in candidates.values()
            if (candidate["店铺名称"], candidate["SKU编码"]) in sales30_index
        ]
    )
    result.stats["jd_inventory_candidate_count"] = len([1 for candidate in candidates.values() if candidate["sources"] & {"jd_inventory", "jd_weekly"}])

    for key, candidate in candidates.items():
        sku = candidate["SKU编码"]
        archive = archive_index.get(sku, {})
        sale30 = sales30_index.get((candidate["店铺名称"], sku), {})
        jd = jd_index.get(sku, {})
        amz = amazon_index.get(sku, {})
        jd_inventory = jd_inventory_index.get(sku, {})
        is_jd_self = candidate["平台"] == "京东自营"
        if sale30.get("has_actual_available") and not is_jd_self:
            in_stock = sale30.get("实际可用数")
        elif is_jd_self:
            in_stock = _first_number(
                parse_number(jd_inventory.get("在库库存")),
                parse_number(jd.get("总可用库存")),
                parse_number(archive.get("实际可用数")),
                parse_number(amz.get("在库库存")),
            )
        else:
            # Domestic marketplace rows share the central product archive.
            # A JD warehouse record with the same barcode is JD-self stock,
            # not the stock balance for a Tmall/PDD/JD-POP row.
            in_stock = _first_number(
                parse_number(archive.get("实际可用数")),
                parse_number(amz.get("在库库存")),
            )
        if is_jd_self:
            in_transit = _first_number(parse_number(jd_inventory.get("在途库存")), parse_number(jd.get("采购在途")), 0)
        elif sale30.get("has_purchase_transit"):
            in_transit = sale30.get("采购在途")
        else:
            # The W27 target column is "在途库存" and its matching source
            # field is the archive's "采购在途".  Reservation/transfer
            # quantities are separate operational states and must not be
            # added into this field.
            in_transit = parse_number(archive.get("采购在途"))
            if in_transit is None:
                in_transit = _first_number(parse_number(amz.get("在途库存")), 0)
        sales_7 = _first_number(
            parse_number(jd.get("成交商品件数")) if is_jd_self else None,
            parse_number(jd_inventory.get("近7天销量")) if is_jd_self else None,
            weekly_sales_by_shop_sku.get((candidate["店铺名称"], sku)),
            parse_number(archive.get("7天销量")),
        )
        sales_30 = sale30.get("30天销量") if sale30.get("has_direct_30d") and not is_jd_self else None
        source_field = ""
        source_file = ""
        source_sheet = ""
        raw_value = ""
        fallback_used = "否"

        if sale30.get("has_direct_30d") and sales_30 is not None and not is_jd_self:
            source_field = "30天销量"
            source_file = role_files.get("sales_30d", Path("")).name
            source_sheet = sale30.get("__sheet_name", "Sheet1")
            raw_value = str(sale30.get("30天销量"))

        # The designated 30-day sales-theme export is the business source of
        # truth for domestic 30-day quantity.  Its net quantity naturally
        # changes as returns are recorded after an earlier export, so it must
        # not be replaced by this week's gross sales or shipment quantity.
        if not source_field and not is_jd_self and sale30.get("has_net"):
            sales_30 = sale30.get("净销量")
            source_field = "净销量"
            source_file = role_files.get("sales_30d", Path("")).name
            source_sheet = sale30.get("__sheet_name", "Sheet1")
            raw_value = str(sale30.get("净销量"))

        fallback_candidates = []
        if not source_field and is_jd_self:
            fallback_candidates = [
            (
                _first_number(
                    parse_number(jd.get("全国近30天出库销量")),
                    parse_number(jd.get("30天")),
                ) if is_jd_self else None,
                "全国近30天出库销量" if parse_number(jd.get("全国近30天出库销量")) is not None else "30天",
                role_files.get("jd_self_weekly_sales", Path("")).name,
                jd.get("__sheet_name", "Sheet1"),
                "否",
            ),
            (parse_number(jd_inventory.get("近30天销量")) if is_jd_self else None, "全国近30天出库销量", jd_inventory.get("file", ""), jd_inventory.get("sheet", ""), "否"),
            ]
        elif not source_field:
            issues.append(
                Issue(
                    "WARNING",
                    "builder",
                    "inventory_rollup",
                    "国内近30天销量缺少明确来源字段，未使用销售数量/净销量/月销量替代",
                    role_files.get("sales_30d", Path("")).name,
                    sale30.get("__sheet_name", "Sheet1"),
                    "",
                    sku,
                    "近30天销量",
                    "W27",
                    "7_库存补货",
                    "请人工确认并复核，并补充含“30天销量”字段的来源表后重跑。",
                )
            )

        for candidate_value, candidate_field, candidate_file, candidate_sheet, candidate_fallback in fallback_candidates:
            if candidate_value not in (None, 0):
                sales_30 = candidate_value
                source_field = candidate_field
                source_file = candidate_file
                source_sheet = candidate_sheet
                fallback_used = candidate_fallback
                if candidate_field == "净销量":
                    raw_value = str(sale30.get("净销量"))
                elif candidate_field == "销售数量":
                    raw_value = str(sale30.get("销售数量"))
                elif candidate_field == "月销量":
                    raw_value = str(archive.get("月销量"))
                elif candidate_field == "30天":
                    raw_value = str(jd.get("30天"))
                elif candidate_field == "全国近30天出库销量":
                    raw_value = str(jd_inventory.get("近30天销量"))
                else:
                    raw_value = str(amz.get("近30天销量"))
                break
        if sales_30 is not None and source_field != "净销量" and sku in sales30_by_sku and candidate["店铺名称"] not in sales30_by_sku[sku]:
            result.stats["shop_sku_fallback_count"] += 1
        # The inventory sheet is an inventory balance, not a sales candidate
        # list. Rows with no stock and no inbound quantity are excluded even
        # when a sales source happens to contain a quantity for the SKU.
        if in_stock in (None, 0) and in_transit in (None, 0):
            result.stats["skip_no_effect_count"] += 1
            issues.append(Issue("INFO", "builder", "inventory_rollup", "无有效库存/销量信息，未写入 7_库存补货", "", "", "", sku, "", "W27", "7_库存补货", "补充库存或销量来源后重跑。"))
            result.audit_rows.append(
                {
                    "目标Sheet": "7_库存补货",
                    "平台": candidate["平台"],
                    "店铺名称": candidate["店铺名称"],
                    "SKU编码": sku,
                    "商品名称": candidate["商品名称"] or archive.get("商品名") or sku,
                    "原目标表是否已有": "",
                    "本次是否写入": "否",
                    "差异类型": "被过滤跳过",
                    "来源文件": source_file,
                    "来源Sheet": source_sheet,
                    "来源字段": source_field or "近30天销量/库存字段",
                    "关键值": f"在库库存={in_stock}; 在途库存={in_transit}; 近7天销量={sales_7}; 近30天销量={sales_30}",
                    "说明": "无有效库存/销量信息，未写入 7_库存补货",
                }
            )
            continue
        payload = {
            "平台": candidate["平台"],
            "店铺名称": candidate["店铺名称"],
            # Keep the inventory sheet's domestic display name aligned with
            # the same standard product master used in 3_SKU明细.
            "商品名称": archive.get("商品名") or candidate["商品名称"] or amz.get("商品名称") or jd_inventory.get("商品名称") or sku,
            "SKU编码": sku,
            "在库库存": in_stock,
            "可售库存": in_stock,
            "在途库存": in_transit,
            "近7天销量": sales_7,
            "近30天销量": sales_30,
        }
        result.rows.append(payload)
        if archive:
            result.stats["product_archive_candidate_count"] += 1
        if amz:
            result.stats["amazon_inventory_candidate_count"] += 1
        if fallback_used == "是" and sales_30 not in (None, 0):
            issues.append(Issue("INFO", "builder", "inventory_rollup", "近30天销量使用补充来源", source_file, source_sheet, "", sku, "近30天销量", "W27", "7_库存补货", "已使用补充来源填充近30天销量。"))
        result.audit_rows.append(
            {
                "平台": candidate["平台"],
                "店铺名称": candidate["店铺名称"],
                "SKU编码": sku,
                "商品名称": payload["商品名称"],
                "近30天销量值": sales_30,
                "近30天销量来源文件": source_file,
                "近30天销量来源Sheet": source_sheet,
                "近30天销量来源字段": source_field,
                "近30天销量来源原始值": raw_value,
                "近30天销量是否补充来源": fallback_used,
            }
        )
        primary_source_file = source_file or jd_inventory.get("file") or amz.get("file") or role_files.get("product_archive", Path("")).name
        primary_source_sheet = source_sheet or jd_inventory.get("sheet") or amz.get("sheet") or "Sheet1"
        result.source_lookup[(cycle_code, payload["平台"], payload["店铺名称"], sku)] = (
            primary_source_file,
            primary_source_sheet,
        )
    result.stats["candidate_after_filter_count"] = len(result.rows)
    return result


def build_kepule_sales_rows(
    role_files: dict[str, Path],
    issues: list[Issue],
    sales_date: str,
    shop_config: dict,
) -> BuildResult:
    result = BuildResult(stats=_blank_stats())
    archive_index = build_product_archive_index(role_files["product_archive"]) if "product_archive" in role_files else {}
    category_index = (
        build_product_category_index(role_files["product_archive"], role_files.get("sales_30d"))
        if "product_archive" in role_files
        else {}
    )
    amazon_cost_index = build_amazon_cost_index(role_files["amazon_latest_inventory"]) if "amazon_latest_inventory" in role_files else {}
    domestic_ranking_index = build_domestic_sales_ranking_index(role_files["domestic_sales_ranking"]) if "domestic_sales_ranking" in role_files else {}
    domestic_ad_index = build_domestic_ad_allocation_index(role_files["sales_theme_analysis"]) if "sales_theme_analysis" in role_files else {}
    jd_refund_index = build_jd_refund_allocation_index(role_files["jd_self_weekly_sales"]) if "jd_self_weekly_sales" in role_files else {}
    domestic_payloads: list[dict] = []
    jd_self_payloads: list[dict] = []
    cross_border_payloads: list[dict] = []

    if "domestic_sales_theme_analysis" in role_files:
        source_path = role_files["domestic_sales_theme_analysis"]
        for row in load_rows(
            source_path,
            None,
            ["店铺", "商品编码", "商品名称", "产品分类", "销售数量", "销售金额", "销售成本", "净销量", "净销售额", "净销售成本", "净销售毛利"],
            optional_fields=["退货数量", "退货金额", "退货成本", "是否新品", "是否清库", "是否B2B大单"],
        ):
            result.stats["raw_candidate_count"] += 1
            normalized = resolve_shop(str(row["店铺"]).strip(), shop_config)
            net_qty = parse_number(row.get("净销量"))
            net_sales = parse_number(row.get("净销售额"))
            net_cost = parse_number(row.get("净销售成本"))
            net_profit = parse_number(row.get("净销售毛利"))
            gross_qty = parse_number(row.get("销售数量"))
            gross_sales = parse_number(row.get("销售金额"))
            gross_cost = parse_number(row.get("销售成本"))
            return_qty = parse_number(row.get("退货数量"))
            refund_amount = parse_number(row.get("退货金额"))
            has_real_return_order = (
                return_qty is not None
                and return_qty >= 1
                and abs(return_qty - round(return_qty)) < 1e-9
            )

            # The Kepule source-detail contract is gross sales plus a separate
            # refund column. W27/W28 both keep 销售数量/销售金额/销售成本 here;
            # using the 净* columns double-applies returns and understates the
            # rows that have a refund.
            qty = gross_qty if gross_qty is not None else net_qty
            sales_amount = gross_sales if gross_sales is not None else net_sales
            sales_cost = gross_cost if gross_cost is not None else net_cost
            profit = (sales_amount - sales_cost) if (sales_amount is not None and sales_cost is not None) else net_profit
            payload = {
                "渠道_标准": normalized["channel_standard"],
                # The completed Kepule template uses the same normalized
                # business-channel label in all three channel dimensions.
                # The source shop name is retained in the audit report rather
                # than becoming a second, incompatible business key.
                "渠道_原始": normalized["channel_standard"],
                "国家/站点": "国内",
                "店铺/客户": normalized["channel_standard"],
                "SKU": str(row["商品编码"]).strip(),
                "产品名称": row.get("商品名称") or str(row["商品编码"]).strip(),
                # 商品档案是产品分类的主数据。销售导出里的分类会随着平台
                # 口径变化，不能覆盖商品档案；仅在档案无法唯一解析时才回退
                # 到销售导出的分类，避免把来源表的临时口径写进周月报。
                "产品分类": _canonical_product_category(
                    resolve_product_category(category_index, row["商品编码"])
                    or archive_index.get(str(row["商品编码"]).strip(), {}).get("产品分类", "")
                    or row.get("产品分类")
                ),
                "是否新品": row.get("是否新品"),
                "是否清库": row.get("是否清库"),
                "是否B2B大单": row.get("是否B2B大单"),
                "销量": qty,
                "销售额_元": sales_amount,
                "销售成本_元": sales_cost,
                # The domestic source's 退货金额 is a refund amount. It is not
                # advertising spend: W28 completed Tmall rows independently
                # confirm the same values in 退款金额_元. There is no domestic
                # advertising-spend source column in this input contract, so
                # keep 广告费_元 blank rather than relabel a refund.
                "平台利润_元": None,
                "退款金额_元": refund_amount if has_real_return_order and refund_amount not in (None, 0) else None,
                "广告费_元": domestic_ad_index.get((str(row.get("店铺") or "").strip(), str(row["商品编码"]).strip())),
                # Business remarks are not a technical provenance channel.
                # The source basis is recorded in the audit report instead.
                "备注": "",
            }
            # Keep quantity-bearing zero-revenue rows: W28 retains the same
            # structure for several channels. Only a row with neither quantity
            # nor revenue is a fee/residual row rather than sales detail.
            if all(payload[field] in (None, 0) for field in ["销量", "销售额_元"]):
                result.stats["skip_no_effect_count"] += 1
                issues.append(Issue("INFO", "builder", "domestic_sales_theme_analysis", "销量和销售额均为零，未写入 源_销售明细；成本或退款残值不单独形成销售行", source_path.name, "Sheet1", "", payload["SKU"], "销量/销售额", "开普乐", "源_销售明细", "保留在来源审计中，不新增销售明细行。"))
                continue
            if not payload["产品分类"]:
                issues.append(Issue("WARNING", "builder", "domestic_sales_theme_analysis", "商品分类缺失", source_path.name, "Sheet1", "", payload["SKU"], "产品分类", "开普乐", "产品分类", "补充商品档案后重跑。"))
            domestic_payloads.append(payload)
            result.source_lookup[(sales_date, payload["渠道_标准"], payload["店铺/客户"], payload["SKU"])] = (source_path.name, row.get("__sheet_name", "Sheet1"))

    if "domestic_sales_ranking" in role_files:
        domestic_totals: dict[str, dict[str, float]] = {}
        for payload in domestic_payloads:
            totals = domestic_totals.setdefault(payload["SKU"], {"销量": 0.0, "销售额_元": 0.0, "销售成本_元": 0.0})
            for field in totals:
                value = parse_number(payload.get(field))
                if value is not None:
                    totals[field] += value
        for sku, totals in domestic_totals.items():
            ranking_row = domestic_ranking_index.get(sku)
            if not ranking_row:
                continue
            ranking_qty = _first_number(parse_number(ranking_row.get("净销量")), parse_number(ranking_row.get("销售数量")))
            ranking_sales = _first_number(parse_number(ranking_row.get("净销售额")), parse_number(ranking_row.get("销售金额")))
            ranking_cost = _first_number(parse_number(ranking_row.get("净销售成本")), parse_number(ranking_row.get("销售成本")))
            mismatches = []
            for field, ranking_value in [("销量", ranking_qty), ("销售额_元", ranking_sales), ("销售成本_元", ranking_cost)]:
                if _numbers_differ(totals[field], ranking_value):
                    mismatches.append(f"{field}={totals[field]} vs 排行表={ranking_value}")
            if mismatches:
                issues.append(
                    Issue(
                        "WARNING", "builder", "domestic_sales_ranking",
                        f"与国内销售排行数据不一致，建议人工复核：{'；'.join(mismatches)}",
                        role_files["domestic_sales_ranking"].name, ranking_row.get("__sheet_name", "Sheet1"),
                        str(ranking_row.get("__row_number", "")), sku, "净销量/净销售额/净销售成本",
                        "开普乐", "源_销售明细", "确认国内销售主题分析与排行表口径后再交付。",
                    )
                )

    if "jd_self_weekly_sales" in role_files:
        source_path = role_files["jd_self_weekly_sales"]
        for row in load_rows(source_path, None, ["商品名称", "69码", "成交商品件数", "成交金额", "匹配成本价"]):
            result.stats["raw_candidate_count"] += 1
            sku = str(row["69码"]).strip()
            qty = parse_number(row.get("成交商品件数"))
            cost = parse_number(row.get("匹配成本价"))
            payload = {
                "渠道_标准": "京东自营",
                "渠道_原始": "京东自营",
                "国家/站点": "国内",
                "店铺/客户": "京东自营",
                "SKU": sku,
                "产品名称": row.get("商品名称") or sku,
                "产品分类": _canonical_product_category(resolve_product_category(category_index, sku) or archive_index.get(sku, {}).get("产品分类", "")),
                "是否新品": None,
                "是否清库": None,
                "是否B2B大单": None,
                "销量": qty,
                "销售额_元": parse_number(row.get("成交金额")),
                "销售成本_元": (cost * qty) if (cost is not None and qty is not None) else None,
                "平台利润_元": None,
                "退款金额_元": jd_refund_index.get(sku),
                "广告费_元": None,
                "备注": "",
            }
            if all(payload[field] in (None, 0) for field in ["销量", "销售额_元", "销售成本_元", "平台利润_元", "退款金额_元", "广告费_元"]):
                result.stats["skip_no_effect_count"] += 1
                issues.append(Issue("INFO", "builder", "jd_self_weekly_sales", "无销售/成本/退款/广告/利润影响，未写入 源_销售明细", source_path.name, "经营状况-商品明细周报", "", sku, "", "开普乐", "源_销售明细", "保留为空即可。"))
                continue
            jd_self_payloads.append(payload)
            result.source_lookup[(sales_date, payload["渠道_标准"], payload["店铺/客户"], payload["SKU"])] = (source_path.name, row.get("__sheet_name", "经营状况-商品明细周报"))

    if "cross_border_profit_sku" in role_files:
        source_path = role_files["cross_border_profit_sku"]
        channel_map = {
            "北美": ("亚马逊北美", "北美站"),
            "日本": ("亚马逊日本", "日本站"),
            "欧洲": ("亚马逊欧洲", "欧洲站"),
        }
        wb = load_workbook(source_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            context = build_sheet_context(ws, ["店铺", "品名", "SKU", "毛利润", "销量", "销售额", "退款金额", "广告花费", "合计成本"])
            channel_standard, country = channel_map.get(sheet_name, ("未识别渠道", sheet_name))
            for row_index in range(context.data_start_row, ws.max_row + 1):
                result.stats["raw_candidate_count"] += 1
                sku = ws.cell(row_index, context.columns["SKU"]).value
                if not sku:
                    continue
                sku = str(sku).strip()
                qty = parse_number(ws.cell(row_index, context.columns["销量"]).value)
                total_cost = parse_number(ws.cell(row_index, context.columns["合计成本"]).value)
                if total_cost is not None:
                    sales_cost = abs(total_cost)
                    cost_note = "跨境销售成本来源=订单利润表合计成本"
                else:
                    cost_parts = []
                    for field in ["采购成本", "头程成本", "其他成本", "平台费", "FBA发货费"]:
                        column = context.columns.get(field)
                        if column is not None:
                            part = parse_number(ws.cell(row_index, column).value)
                            if part is not None:
                                cost_parts.append(abs(part))
                    sales_cost = sum(cost_parts) if cost_parts else None
                    cost_note = "跨境销售成本来源=订单利润表成本项回退"
                payload = {
                    "渠道_标准": channel_standard,
                    # The target's source-channel column is a normalized
                    # business label, matching the channel key used by its
                    # summary formulas. Sheet names such as “北美” are only
                    # provenance and belong in the audit report.
                    "渠道_原始": channel_standard,
                    "国家/站点": country,
                    # The template keys cross-border rows by its normalized
                    # channel label, not by a display-only region name.
                    "店铺/客户": channel_standard,
                    "SKU": sku,
                    "产品名称": ws.cell(row_index, context.columns["品名"]).value,
                    "产品分类": _canonical_product_category(resolve_product_category(category_index, sku) or archive_index.get(sku, {}).get("产品分类", "")),
                    "是否新品": None,
                    "是否清库": None,
                    "是否B2B大单": None,
                    "销量": qty,
                    "销售额_元": parse_number(ws.cell(row_index, context.columns["销售额"]).value),
                    "销售成本_元": sales_cost,
                    "平台利润_元": parse_number(ws.cell(row_index, context.columns["毛利润"]).value),
                    # A zero in the cross-border profit export means there is
                    # no refund to report for this SKU.  The target's completed
                    # weekly sheet represents that state as blank, not 0; keep
                    # a non-zero source refund (including its sign-normalised
                    # value) but do not manufacture a zero-value refund.
                    "退款金额_元": (
                        abs(refund_amount)
                        if (refund_amount := parse_number(ws.cell(row_index, context.columns["退款金额"]).value)) not in (None, 0)
                        else None
                    ),
                    "广告费_元": abs(parse_number(ws.cell(row_index, context.columns.get("广告花费", 0)).value) or 0) if context.columns.get("广告花费") else None,
                    "备注": "",
                }
                # The completed source sheet is SKU sales detail.  A row
                # whose sales basis is all zero is not an individual sale,
                # even if the profit report contains fee-only adjustments.
                # Those adjustments remain traceable in the source/audit but
                # must not create an extra SKU detail row.
                if all(payload[field] in (None, 0) for field in ["销量", "销售额_元"]):
                    result.stats["skip_no_effect_count"] += 1
                    issues.append(Issue("INFO", "builder", "cross_border_profit_sku", "销量和销售额均为零，未写入 源_销售明细；费用或退款残值不单独形成销售行", source_path.name, sheet_name, "", sku, "销量/销售额", "开普乐", "源_销售明细", "保留在来源审计中，不新增销售明细行。"))
                    continue
                if channel_standard == "未识别渠道":
                    issues.append(Issue("WARNING", "builder", "cross_border_profit_sku", "未识别店铺/渠道/站点", source_path.name, sheet_name, "", sku, "渠道_标准", "开普乐", "渠道_标准", "补充渠道映射。"))
                cross_border_payloads.append(payload)
                result.source_lookup[(sales_date, payload["渠道_标准"], payload["店铺/客户"], payload["SKU"])] = (source_path.name, sheet_name)
        wb.close()

    # Preserve source-family order rather than imposing an unrelated fixed
    # channel alphabet: cross-border sheets first, then domestic rows grouped
    # by each channel's first appearance, then the independent JD-self file.
    # Python's stable sort keeps the source row order inside each channel.
    domestic_channel_rank: dict[str, int] = {}
    for payload in domestic_payloads:
        domestic_channel_rank.setdefault(str(payload.get("渠道_标准") or ""), len(domestic_channel_rank))
    domestic_payloads.sort(key=lambda payload: domestic_channel_rank[str(payload.get("渠道_标准") or "")])
    result.rows = [*cross_border_payloads, *domestic_payloads, *jd_self_payloads]
    for field in ("是否新品", "是否清库", "是否B2B大单"):
        missing_count = sum(1 for payload in result.rows if payload.get(field) in (None, ""))
        if missing_count:
            issues.append(
                Issue(
                    "WARNING",
                    "builder",
                    "operational_classification",
                    f"{field}在当前来源中缺失，共{missing_count}行未自动填写",
                    "",
                    "",
                    "",
                    "",
                    field,
                    "开普乐",
                    "源_销售明细",
                    f"请业务部门提供包含“{field}”的明细来源，或在输出表中人工确认；程序不会用商品标签猜测。",
                )
            )
    result.stats["candidate_after_filter_count"] = len(result.rows)
    return result


def build_kepule_inventory_rows(
    role_files: dict[str, Path],
    issues: list[Issue],
    inventory_date: str,
    *,
    include_domestic_archive: bool = True,
    previous_supplier_by_display_name: dict[str, str] | None = None,
) -> BuildResult:
    result = BuildResult(stats=_blank_stats())
    previous_supplier_by_display_name = previous_supplier_by_display_name or {}
    fba_inventory_index = build_fba_inventory_index(role_files["fba_inventory"]) if "fba_inventory" in role_files else {}
    domestic_style_by_product_code, known_domestic_styles = (
        build_domestic_style_sku_index(role_files["sales_30d"])
        if "sales_30d" in role_files
        else ({}, set())
    )
    domestic_supplier_by_product_code = (
        build_domestic_supplier_index(role_files["sales_30d"])
        if "sales_30d" in role_files
        else {}
    )
    if "amazon_latest_inventory" in role_files:
        source_path = role_files["amazon_latest_inventory"]
        channel_map = {
            "北美站": "亚马逊北美",
            "日本站": "亚马逊日本",
            "欧洲站-英国仓": "亚马逊欧洲",
            "欧洲站--德国仓": "亚马逊欧洲",
        }
        wb = load_workbook(source_path, read_only=True, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            context = build_sheet_context(ws, ["SKU", "品名", "成本", "可用的总库存", "可销售天数"])
            for row_index in range(context.data_start_row, ws.max_row + 1):
                result.stats["raw_candidate_count"] += 1
                sku = ws.cell(row_index, context.columns["SKU"]).value
                if not sku:
                    continue
                sku = str(sku).strip()
                qty = parse_number(ws.cell(row_index, context.columns["可用的总库存"]).value)
                cost = parse_number(ws.cell(row_index, context.columns["成本"]).value)
                fba_available_qty = (
                    parse_number(ws.cell(row_index, context.columns["FBA可用库存"]).value)
                    if context.columns.get("FBA可用库存")
                    else None
                )
                payload = {
                    "SKU/分类": sku,
                    "仓库/区域": sheet_name,
                    "渠道_标准": channel_map.get(sheet_name, "未识别渠道"),
                    "库存数量": qty,
                    "库存金额_元": (qty * cost) if (qty is not None and cost is not None) else None,
                    "可售天数": parse_number(ws.cell(row_index, context.columns["可销售天数"]).value),
                    # 类型 is an operating judgement (正常/滞销/长库龄), not
                    # an inventory-source label. Region/channel already records
                    # that this is FBA, so keep 类型 empty unless a real rule is
                    # supplied.
                    "类型": None,
                    "备注": f"可售天数={parse_number(ws.cell(row_index, context.columns['可销售天数']).value)}",
                }
                fba_row = fba_inventory_index.get((_inventory_region(sheet_name), sku))
                if fba_row:
                    fba_qty = _first_number(
                        parse_number(fba_row.get("FBA可用库存")),
                        parse_number(fba_row.get("可用库存")),
                        parse_number(fba_row.get("总可用库存")),
                        parse_number(fba_row.get("库存")),
                        parse_number(fba_row.get("FBA库存")),
                    )
                    comparison_qty = fba_available_qty if fba_available_qty is not None else payload["库存数量"]
                    if _numbers_differ(comparison_qty, fba_qty):
                        issues.append(
                            Issue(
                                "WARNING",
                                "builder",
                                "fba_inventory",
                                f"与FBA库存表数量不一致，建议人工复核：FBA可用库存={comparison_qty} vs FBA库存表={fba_qty}",
                                role_files["fba_inventory"].name,
                                fba_row.get("__sheet_name", "Sheet1"),
                                str(fba_row.get("__row_number", "")),
                                sku,
                                "FBA可用库存",
                                "开普乐",
                                "源_库存快照",
                                "确认亚马逊最新库存与FBA库存表口径后再交付。",
                            )
                        )
                # 亚马逊库存是 7_库存补货 的来源核对数据，不属于开普乐
                # “源_库存快照”的国内总仓写入范围。仍逐行读取并与 FBA
                # 库存表核验，避免把来源文件误判为未参与。
        wb.close()

    if "product_archive" in role_files and include_domestic_archive:
        source_path = role_files["product_archive"]
        for row in load_rows(source_path, None, ["商品编码", "商品名", "实际可用数", "产品分类"]):
            result.stats["raw_candidate_count"] += 1
            product_code = str(row["商品编码"]).strip()
            cost = parse_number(row.get("成本价"))
            available_qty = parse_number(row.get("实际可用数"))
            purchase_transit = parse_number(row.get("采购在途"))
            # The source's available quantity excludes purchase-in-transit.
            # The completed domestic snapshot reports total owned inventory.
            qty = (available_qty or 0) + (purchase_transit or 0)
            # The completed week/month workbook establishes the contract for
            # this domestic inventory field: its label is “SKU/分类”, but the
            # actual business identifier is the product-archive display name.
            # Do not replace it with a barcode or a style code.
            display_name = str(row.get("商品名") or "").strip() or product_code
            product_tag = str(row.get("商品标签") or "").strip()
            if product_tag == "E类品":
                result.stats["skip_no_effect_count"] += 1
                issues.append(Issue("INFO", "builder", "product_archive", "商品标签为E类品，不写入国内库存快照", source_path.name, row.get("__sheet_name", "Sheet1"), str(row.get("__row_number", "")), product_code, "商品标签", "开普乐", "源_库存快照", "按档案商品标签排除。"))
                continue
            if (qty is None or qty <= 0) and (cost is None or cost <= 0):
                result.stats["skip_no_effect_count"] += 1
                issues.append(Issue("INFO", "builder", "product_archive", "库存数量为零且无正成本，不写入国内库存快照", source_path.name, row.get("__sheet_name", "Sheet1"), str(row.get("__row_number", "")), product_code, "实际可用数/采购在途/成本价", "开普乐", "源_库存快照", "保留为空即可。"))
                continue
            age_days = inventory_age_days(row.get("首次入库时间"), inventory_date)
            supplier = (
                str(row.get("供应商") or "").strip()
                or domestic_supplier_by_product_code.get(product_code, "")
                or previous_supplier_by_display_name.get(display_name, "")
            )
            if not supplier:
                issues.append(
                    Issue(
                        "WARNING",
                        "builder",
                        "product_archive",
                        "国内库存供应商缺失，未伪造供应商",
                        source_path.name,
                        row.get("__sheet_name", "Sheet1"),
                        str(row.get("__row_number", "")),
                        product_code,
                        "供应商",
                        "开普乐",
                        "源_库存快照",
                        "补充供应商来源或保留上一期同商品的唯一供应商后重跑；当前不写入伪造渠道。",
                    )
                )
            payload = {
                "SKU/分类": display_name,
                "仓库/区域": "国内总仓",
                # The completed monthly/weekly workbook uses the supplier as
                # the domestic inventory channel.  A missing source must not
                # be replaced with a fabricated “国内仓” business channel.
                "渠道_标准": supplier,
                "库存数量": qty,
                # The completed snapshot's 成本价 is the product archive's
                # unit cost; 库存金额 is the corresponding quantity × unit cost.
                "成本价": cost,
                "库存金额_元": (qty * cost) if (qty is not None and cost is not None) else None,
                "库龄天数": age_days,
                # Despite its archive label, the completed W27 workbook maps
                # 库存周转天数 directly into the target's 可售天数 column.
                "可售天数": parse_number(row.get("库存周转天数")),
                # 库龄大于 18 个月（540 天）默认长库龄，其余有明确库龄
                # 的默认正常；滞销保留给业务同事在结果表手工判断。
                "类型": inventory_type_from_age(age_days),
                # Provenance is retained in the audit report. The business
                # template keeps this column available for manual notes.
                "备注": "",
            }
            result.rows.append(payload)
            result.source_lookup[(inventory_date, payload["SKU/分类"], payload["仓库/区域"], payload["渠道_标准"])] = (source_path.name, "Sheet1")

    missing_type_count = sum(1 for payload in result.rows if payload.get("类型") in (None, ""))
    if missing_type_count:
        issues.append(
            Issue(
                "WARNING",
                "builder",
                "operational_classification",
                f"库存类型在当前来源中缺失，共{missing_type_count}行未自动填写",
                "",
                "",
                "",
                "",
                "类型",
                "开普乐",
                "源_库存快照",
                "请业务部门人工确认：首次入库时间缺失，无法按库龄自动写入正常或长库龄；滞销始终由业务部门手工判断。",
            )
        )
    result.stats["candidate_after_filter_count"] = len(result.rows)
    return result
