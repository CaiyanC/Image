from __future__ import annotations

from datetime import date, datetime
from collections import defaultdict
import re

from openpyxl import load_workbook

from src.excel_utils import (
    build_sheet_context,
    copy_template_row,
    first_non_blank_row_from_bottom,
    is_logical_blank_row,
    row_key,
)
from src.models import RowWriteEvent, SheetWriteResult


DOMESTIC_CHANNELS = {"天猫", "京东POP", "京东自营", "拼多多", "抖音2店", "小红书", "得物", "视频号", "商务部", "达人带货合计"}
CROSS_BORDER_CHANNELS = {"亚马逊北美", "亚马逊日本", "亚马逊欧洲"}
KEPULE_SALES_FORCE_WRITE_FIELDS = {
    "日期",
    "周次",
    "年月",
    "事业部",
    "渠道_标准",
    "渠道_原始",
    "国家/站点",
    "店铺/客户",
    "SKU",
    "产品名称",
    "产品分类",
    # These are operational classifications, not formulas.  A reusable
    # template row may still contain the previous row's values, so new rows
    # must explicitly replace them (including with blank when the source does
    # not provide the field). Existing same-period rows are preserved by the
    # upsert branch above, allowing reviewed manual values to survive reruns.
    "是否新品",
    "是否清库",
    "是否B2B大单",
    "销量",
    "销售额_元",
    "销售成本_元",
    "毛利_元",
    "毛利率",
    "退款金额_元",
    "广告费_元",
    "备注",
}
KEPULE_INVENTORY_FORCE_WRITE_FIELDS = {
    "日期",
    "周次",
    "年月",
    "SKU/分类",
    "仓库/区域",
    "渠道_标准",
    "库存数量",
    "库存金额_元",
    "库龄天数",
    "可售天数",
    "库龄段",
    "类型",
    "状态",
    "处置进度",
    "备注",
}
PERIOD_METADATA_FIELDS = {"日期", "周次", "年月"}

# These are the channel-level formulas already used in the completed weekly
# template.  Some blank template rows do not carry the formula in column T,
# so formula preservation alone leaves a business calculation missing.
DOMESTIC_PLATFORM_FEE_RATES = {
    "天猫": "5.5%",
    "京东POP": "5.5%",
    "抖音2店": "5.5%",
    "小红书": "5.5%",
    "视频号": "5.5%",
    "拼多多": "0.5%",
    "得物": "18%",
    "京东自营": "24%",
}


def _year_month(date_value: str) -> str:
    return datetime.strptime(date_value, "%Y-%m-%d").strftime("%Y-%m")


def _excel_date(date_value: str):
    return datetime.strptime(date_value, "%Y-%m-%d").date()


def _domestic_platform_profit_formula(channel: str, row_index: int) -> str | None:
    if str(channel or "").strip() == "商务部":
        return f"=P{row_index}-Q{row_index}"
    rate = DOMESTIC_PLATFORM_FEE_RATES.get(str(channel or "").strip())
    if rate is None:
        return None
    return f"=R{row_index}-(P{row_index}*{rate})-V{row_index}"


def _formula_shape(formula: str) -> str:
    """Compare copied Excel formulas without treating their row number as logic."""
    return re.sub(r"(\$?[A-Z]{1,3})\$?\d+", r"\1#", formula.upper())


def _fill_missing_domestic_platform_profit_formulas(ws, date_value: str) -> None:
    """Apply the current row's domestic-profit formula to blank or stale template formulas."""
    context = build_sheet_context(ws, ["日期", "渠道_标准", "SKU", "销售额_元", "平台利润_元"])
    date_column = context.columns["日期"]
    channel_column = context.columns["渠道_标准"]
    sku_column = context.columns["SKU"]
    sales_column = context.columns["销售额_元"]
    profit_column = context.columns["平台利润_元"]
    period_key = _key_value(_excel_date(date_value))
    for row_index in range(context.data_start_row, ws.max_row + 1):
        if _key_value(ws.cell(row_index, date_column).value) != period_key:
            continue
        if not _key_value(ws.cell(row_index, sku_column).value):
            continue
        if ws.cell(row_index, sales_column).value in (None, ""):
            continue
        formula = _domestic_platform_profit_formula(ws.cell(row_index, channel_column).value, row_index)
        if not formula:
            continue
        profit_cell = ws.cell(row_index, profit_column)
        current = profit_cell.value
        if current in (None, ""):
            profit_cell.value = formula
            continue
        # A copied template row can carry a valid-looking formula for a
        # different channel (for example a 5.5% shop formula on a 商务部 row).
        # Keep reviewed literal values untouched, but replace only a stale
        # formula with the channel-specific formula for this row.
        if isinstance(current, str) and current.startswith("=") and _formula_shape(current) != _formula_shape(formula):
            profit_cell.value = formula


def _key_value(value) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value or "")


def _business_unit(channel_standard: str, channel_raw: str = "", country_site: str = "") -> str:
    channel_standard = (channel_standard or "").strip()
    channel_raw = (channel_raw or "").strip()
    country_site = (country_site or "").strip()
    if channel_standard in DOMESTIC_CHANNELS:
        return "国内事业部"
    if channel_standard in CROSS_BORDER_CHANNELS:
        return "跨境事业部"
    if any(keyword in channel_standard for keyword in ["国内", "天猫", "京东", "拼多多", "抖音", "小红书", "得物", "视频号", "商务"]):
        return "国内事业部"
    if any(keyword in channel_standard for keyword in ["跨境", "亚马逊"]):
        return "跨境事业部"
    if any(keyword in channel_raw for keyword in ["天猫", "京东", "拼多多", "抖音", "小红书", "得物", "视频号", "商务"]):
        return "国内事业部"
    if any(keyword in channel_raw for keyword in ["亚马逊", "Amazon"]):
        return "跨境事业部"
    if any(keyword in country_site for keyword in ["中国"]):
        return "国内事业部"
    if any(keyword in country_site for keyword in ["北美", "日本", "欧洲", "美国", "英国", "德国"]):
        return "跨境事业部"
    return "需确认"


def _has_formula(cell) -> bool:
    return isinstance(cell.value, str) and cell.value.startswith("=")


def _first_row_with_empty_keys(ws, data_start_row: int, key_columns: list[int]) -> int | None:
    """Find a reusable template row without treating calculation formulas as data."""
    for row_index in range(data_start_row, ws.max_row + 1):
        if all(ws.cell(row_index, column).value in (None, "") for column in key_columns):
            return row_index
    return None


def _bucket_from_days(days_value) -> str | None:
    if days_value not in (None, ""):
        days = float(days_value)
        if days > 540:
            return "18个月以上"
        if days > 360:
            return "12-18个月"
        if days > 180:
            return "6-12个月"
        return "6个月以下"
    return None


def _bucket_from_fba_columns(row: dict) -> str | None:
    # Prefer the oldest non-zero FBA bucket when bucket totals are available.
    bucket_fields = [
        ("FBA库存_18个月以上", "18个月以上"),
        ("FBA库存_12-18个月", "12-18个月"),
        ("FBA库存_6-12个月", "6-12个月"),
        ("FBA库存_181-270天", "6-12个月"),
        ("FBA库存_271-365天", "6-12个月"),
        ("FBA库存_365-540天", "12-18个月"),
        ("FBA库存_540天以上", "18个月以上"),
        ("FBA库存_0-90天", "6个月以下"),
        ("FBA库存_91-180天", "6个月以下"),
    ]
    for field, bucket in bucket_fields:
        value = row.get(field)
        if value not in (None, "", 0, 0.0):
            return bucket
    return None


def _age_bucket(days_value, inventory_type: str, notes: str, row: dict | None = None) -> str:
    bucket = _bucket_from_days(days_value)
    if bucket:
        return bucket
    bucket = _bucket_from_fba_columns(row or {})
    if bucket:
        return bucket
    if inventory_type == "FBA库存" and row and row.get("可售天数") not in (None, ""):
        return "6个月以下"
    if inventory_type == "国内库存":
        return "需确认"
    if inventory_type == "FBA库存":
        return "6个月以下"
    return "需确认"


def _upsert_sheet_rows(
    ws,
    required_fields: list[str],
    key_fields: list[str],
    rows: list[dict],
    workbook_role: str,
    source_lookup: dict[tuple[str, ...], tuple[str, str]] | None = None,
    force_write_fields: set[str] | None = None,
    preserve_formula_fields: set[str] | None = None,
) -> SheetWriteResult:
    context = build_sheet_context(ws, required_fields)
    # Templates may remove optional columns such as 仓库/区域. Only columns
    # that actually exist can identify a target row; repeated core keys retain
    # their source order so they are never silently merged.
    available_key_fields = [field for field in key_fields if field in context.columns]
    existing: dict[tuple[str, ...], list[int]] = defaultdict(list)
    for row_index in range(context.data_start_row, ws.max_row + 1):
        if is_logical_blank_row(ws, row_index):
            continue
        key = tuple(_key_value(ws.cell(row_index, context.columns[field]).value) for field in available_key_fields)
        if any(part for part in key):
            existing[key].append(row_index)

    result = SheetWriteResult(
        workbook_role=workbook_role,
        sheet_name=ws.title,
        written_rows=0,
        candidate_rows=len(rows),
        existing_key_count=0,
        original_cycle_rows=0,
    )
    template_limit = ws.max_row

    date_field = "日期"
    restrict_to_existing_period = False
    if date_field in context.columns and rows:
        date_value = _key_value(rows[0].get(date_field, ""))
        for row_index in range(context.data_start_row, ws.max_row + 1):
            if _key_value(ws.cell(row_index, context.columns[date_field]).value) == date_value:
                result.original_cycle_rows += 1
                restrict_to_existing_period = True

    seen_row_keys: dict[tuple[str, ...], int] = defaultdict(int)
    for row in rows:
        source_key = tuple(str(row.get(field, "") or "") for field in key_fields)
        key = tuple(_key_value(row.get(field, "")) for field in available_key_fields)
        occurrence = seen_row_keys[key]
        seen_row_keys[key] += 1
        source_file, source_sheet = ("", "")
        if source_lookup:
            source_file, source_sheet = source_lookup.get(source_key, ("", ""))

        matching_rows = existing.get(key, [])
        target_row = matching_rows[occurrence] if occurrence < len(matching_rows) else None
        if target_row is not None:
            result.existing_key_count += 1
            changed = False
            for field, value in row.items():
                column = context.columns.get(field)
                if column is None:
                    continue
                cell = ws.cell(target_row, column)
                if field in (preserve_formula_fields or set()) and _has_formula(cell):
                    continue
                should_refresh = field in PERIOD_METADATA_FIELDS
                if (should_refresh or cell.value in (None, "")) and value not in (None, ""):
                    cell.value = value
                    changed = True
            if changed:
                result.backfilled_rows += 1
                result.row_events.append(
                    RowWriteEvent(
                        sheet_name=ws.title,
                        row_number=target_row,
                        source_file=source_file,
                        source_sheet=source_sheet,
                        sku=str(row.get("SKU", row.get("SKU/分类", ""))),
                        reason="existing_key_backfill",
                    )
                )
            continue

        # A template that already contains this period declares the intended
        # reporting scope. Do not add unreviewed source candidates beside it.
        if restrict_to_existing_period:
            continue

        target_row = _first_row_with_empty_keys(
            ws,
            context.data_start_row,
            [context.columns[field] for field in available_key_fields],
        )
        if target_row is None:
            target_row = ws.max_row + 1
            template_row = first_non_blank_row_from_bottom(ws, context.data_start_row)
            copy_template_row(ws, template_row, target_row)
            result.appended_row_writes += 1
            reason = "append_after_template_exhausted"
        else:
            result.template_row_writes += 1
            reason = "reuse_formula_template_row"

        for field, value in row.items():
            column = context.columns.get(field)
            if column is None:
                continue
            cell = ws.cell(target_row, column)
            if field in (preserve_formula_fields or set()) and _has_formula(cell):
                continue
            if field in (force_write_fields or set()) or cell.value in (None, ""):
                cell.value = value
        existing[key].append(target_row)
        result.written_rows += 1
        result.row_events.append(
            RowWriteEvent(
                sheet_name=ws.title,
                row_number=target_row,
                source_file=source_file,
                source_sheet=source_sheet,
                sku=str(row.get("SKU", row.get("SKU/分类", ""))),
                reason=reason,
            )
        )

    result.ended_beyond_template = any(event.row_number > template_limit for event in result.row_events)
    return result


def fill_kepule_sales(workbook_path, date_value: str, week_label: str, rows: list[dict]) -> int:
    workbook = load_workbook(workbook_path)
    result = fill_kepule_sales_sheet(workbook["源_销售明细"], date_value, week_label, rows)
    workbook.save(workbook_path)
    workbook.close()
    return result.written_rows


def fill_kepule_sales_sheet(
    ws,
    date_value: str,
    week_label: str,
    rows: list[dict],
    source_lookup: dict[tuple[str, ...], tuple[str, str]] | None = None,
) -> SheetWriteResult:
    payload_rows = []
    ym = _year_month(date_value)
    for row in rows:
        channel = str(row.get("渠道_标准", "") or "")
        channel_raw = str(row.get("渠道_原始", "") or row.get("店铺/客户", "") or "")
        country_site = str(row.get("国家/站点", "") or "")
        sales_amount = row.get("销售额_元")
        sales_cost = row.get("销售成本_元")
        platform_profit = row.get("平台利润_元")
        gross_profit = platform_profit
        if gross_profit in (None, "") and sales_amount not in (None, "") and sales_cost not in (None, ""):
            gross_profit = sales_amount - sales_cost
        payload_rows.append(
            {
                "日期": _excel_date(date_value),
                "周次": week_label,
                "年月": ym,
                "事业部": _business_unit(channel, channel_raw, country_site),
                "毛利_元": gross_profit,
                "毛利率": (gross_profit / sales_amount) if gross_profit not in (None, "") and sales_amount not in (None, "", 0) else None,
                # Always include the operational flags in the payload. This
                # lets the new-row writer clear stale values carried by a
                # reusable template row when the source has no explicit flag.
                "是否新品": row.get("是否新品"),
                "是否清库": row.get("是否清库"),
                "是否B2B大单": row.get("是否B2B大单"),
                **row,
            }
        )
    result = _upsert_sheet_rows(
        ws,
        ["日期", "渠道_标准", "店铺/客户", "SKU"],
        ["日期", "渠道_标准", "店铺/客户", "SKU"],
        payload_rows,
        "kepule_target",
        source_lookup,
        KEPULE_SALES_FORCE_WRITE_FIELDS,
        {"毛利_元", "毛利率"},
    )
    _fill_missing_domestic_platform_profit_formulas(ws, date_value)
    return result


def fill_kepule_inventory(workbook_path, date_value: str, week_label: str, rows: list[dict]) -> int:
    workbook = load_workbook(workbook_path)
    result = fill_kepule_inventory_sheet(workbook["源_库存快照"], date_value, week_label, rows)
    workbook.save(workbook_path)
    workbook.close()
    return result.written_rows


def fill_kepule_inventory_sheet(
    ws,
    date_value: str,
    week_label: str,
    rows: list[dict],
    source_lookup: dict[tuple[str, ...], tuple[str, str]] | None = None,
) -> SheetWriteResult:
    ym = _year_month(date_value)
    payload_rows = []
    for row in rows:
        notes = str(row.get("备注", "") or "")
        payload_rows.append(
            {
                "日期": _excel_date(date_value),
                "周次": week_label,
                "年月": ym,
                "库龄段": row.get("库龄段") or _age_bucket(row.get("库龄天数"), str(row.get("类型", "") or ""), notes, row),
                **row,
            }
        )
    # A continuing workbook already defines this period's domestic inventory
    # scope.  Do not append every product in the archive beside that curated
    # scope; only backfill domestic rows whose business key is present. FBA
    # rows remain appendable because their source is a weekly feed.
    context = build_sheet_context(ws, ["日期", "SKU/分类", "渠道_标准", "库存数量"])
    key_fields = ["日期", "SKU/分类", "仓库/区域", "渠道_标准"]
    available_key_fields = [field for field in key_fields if field in context.columns]
    existing_period_keys = {
        tuple(_key_value(ws.cell(row_index, context.columns[field]).value) for field in available_key_fields)
        for row_index in range(context.data_start_row, ws.max_row + 1)
        if _key_value(ws.cell(row_index, context.columns["日期"]).value) == date_value
    }
    if existing_period_keys:
        payload_rows = [
            row for row in payload_rows
            if tuple(_key_value(row.get(field, "")) for field in available_key_fields) in existing_period_keys
        ]
    return _upsert_sheet_rows(
        ws,
        ["日期", "SKU/分类", "渠道_标准", "库存数量"],
        ["日期", "SKU/分类", "仓库/区域", "渠道_标准"],
        payload_rows,
        "kepule_target",
        source_lookup,
        KEPULE_INVENTORY_FORCE_WRITE_FIELDS,
        {"库存金额_元", "库龄段"},
    )
