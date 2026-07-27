from __future__ import annotations

from datetime import datetime

from openpyxl import load_workbook

from src.excel_utils import (
    build_sheet_context,
    copy_template_row,
    first_logical_blank_row,
    first_non_blank_row_from_bottom,
    is_logical_blank_row,
    row_key,
)
from src.models import RowWriteEvent, SheetWriteResult


SKU_DETAIL_KEY_FIELDS = ["周期", "平台", "店铺名称", "SKU编码"]
INVENTORY_KEY_FIELDS = ["周期", "平台", "店铺名称", "SKU编码"]
W27_FORCE_WRITE_FIELDS = {
    "周期类型",
    "周期",
    "开始日期",
    "结束日期",
    "平台",
    "店铺名称",
    "商品名称",
    "SKU编码",
    "商品分类",
    "销售额",
    "成交件数",
    "退款金额",
    "退款订单数",
    "单件成本",
    "在库库存",
    "在途库存",
    "近7天销量",
    "近30天销量",
}


def _original_cycle_rows(ws, context, cycle_code: str) -> int:
    cycle_col = context.columns["周期"]
    count = 0
    for row_index in range(context.data_start_row, ws.max_row + 1):
        if str(ws.cell(row_index, cycle_col).value or "") == cycle_code:
            count += 1
    return count


def _upsert_rows(
    ws,
    key_fields: list[str],
    cycle_type: str,
    cycle_code: str,
    start_date: str,
    end_date: str,
    rows: list[dict],
    source_lookup: dict[tuple[str, str, str, str], tuple[str, str]] | None = None,
) -> SheetWriteResult:
    context = build_sheet_context(ws, ["周期类型", "周期", "平台", "店铺名称", "SKU编码"])
    existing: dict[tuple[str, ...], int] = {}
    reserved_rows: list[int] = []
    for row_index in range(context.data_start_row, ws.max_row + 1):
        if str(ws.cell(row_index, context.columns["周期"]).value or "") == cycle_code:
            key = row_key(ws, row_index, context.columns, key_fields)
            non_cycle_key = key[1:] if key_fields and key_fields[0] == "周期" else key
            if not any(part for part in non_cycle_key):
                reserved_rows.append(row_index)
                continue
        if is_logical_blank_row(ws, row_index):
            continue
        key = row_key(ws, row_index, context.columns, key_fields)
        non_cycle_key = key[1:] if key_fields and key_fields[0] == "周期" else key
        if any(part for part in non_cycle_key):
            existing[key] = row_index

    result = SheetWriteResult(
        workbook_role="w27_target",
        sheet_name=ws.title,
        written_rows=0,
        existing_key_count=0,
        original_cycle_rows=_original_cycle_rows(ws, context, cycle_code),
        candidate_rows=len(rows),
    )
    template_limit = ws.max_row

    for row in rows:
        key = tuple(cycle_code if field == "周期" else str(row.get(field, "") or "") for field in key_fields)
        source_file, source_sheet = ("", "")
        if source_lookup:
            source_file, source_sheet = source_lookup.get(key, ("", ""))

        target_row = existing.get(key)
        if target_row is not None:
            result.existing_key_count += 1
            changed = False
            for field, value in {
                "周期类型": cycle_type,
                "周期": cycle_code,
                "开始日期": datetime.strptime(start_date, "%Y-%m-%d").date(),
                "结束日期": datetime.strptime(end_date, "%Y-%m-%d").date(),
                **row,
            }.items():
                column = context.columns.get(field)
                if column is None:
                    continue
                cell = ws.cell(target_row, column)
                if (field in W27_FORCE_WRITE_FIELDS or cell.value in (None, "")) and value not in (None, ""):
                    cell.value = value
                    changed = True
            if changed:
                result.backfilled_rows += 1
                result.row_events.append(
                    RowWriteEvent(
                        sheet_name=ws.title,
                        row_number=target_row,
                        source_file=source_file,
                        source_sheet=source_sheet,
                        sku=str(row.get("SKU编码", "")),
                        reason="existing_key_backfill",
                        key_values={"周期": cycle_code, "平台": str(row.get("平台", "")), "店铺名称": str(row.get("店铺名称", ""))},
                    )
                )
            continue

        target_row = reserved_rows.pop(0) if reserved_rows else first_logical_blank_row(ws, context.data_start_row)
        if target_row is None:
            target_row = ws.max_row + 1
            template_row = first_non_blank_row_from_bottom(ws, context.data_start_row)
            copy_template_row(ws, template_row, target_row)
            result.appended_row_writes += 1
            reason = "append_after_template_exhausted"
        else:
            result.template_row_writes += 1
            reason = "reuse_formula_template_row"

        payload = {
            "周期类型": cycle_type,
            "周期": cycle_code,
            "开始日期": datetime.strptime(start_date, "%Y-%m-%d").date(),
            "结束日期": datetime.strptime(end_date, "%Y-%m-%d").date(),
            **row,
        }
        for field, value in payload.items():
            column = context.columns.get(field)
            if column is None:
                continue
            cell = ws.cell(target_row, column)
            if field in W27_FORCE_WRITE_FIELDS or cell.value in (None, ""):
                cell.value = value
        existing[key] = target_row
        result.written_rows += 1
        result.row_events.append(
            RowWriteEvent(
                sheet_name=ws.title,
                row_number=target_row,
                source_file=source_file,
                source_sheet=source_sheet,
                sku=str(row.get("SKU编码", "")),
                reason=reason,
                key_values={"周期": cycle_code, "平台": str(row.get("平台", "")), "店铺名称": str(row.get("店铺名称", ""))},
            )
        )

    result.ended_beyond_template = any(event.row_number > template_limit for event in result.row_events)
    return result


def fill_w27_sku_detail(workbook_path, cycle_type: str, cycle_code: str, start_date: str, end_date: str, rows: list[dict]) -> int:
    workbook = load_workbook(workbook_path)
    result = fill_w27_sku_detail_sheet(workbook["3_SKU明细"], cycle_type, cycle_code, start_date, end_date, rows)
    workbook.save(workbook_path)
    workbook.close()
    return result.written_rows


def fill_w27_inventory(workbook_path, cycle_type: str, cycle_code: str, start_date: str, end_date: str, rows: list[dict]) -> int:
    workbook = load_workbook(workbook_path)
    result = fill_w27_inventory_sheet(workbook["7_库存补货"], cycle_type, cycle_code, start_date, end_date, rows)
    workbook.save(workbook_path)
    workbook.close()
    return result.written_rows


def fill_w27_sku_detail_sheet(
    ws,
    cycle_type: str,
    cycle_code: str,
    start_date: str,
    end_date: str,
    rows: list[dict],
    source_lookup: dict[tuple[str, str, str, str], tuple[str, str]] | None = None,
) -> SheetWriteResult:
    result = _upsert_rows(ws, SKU_DETAIL_KEY_FIELDS, cycle_type, cycle_code, start_date, end_date, rows, source_lookup)
    result.sheet_name = "3_SKU明细"
    return result


def fill_w27_inventory_sheet(
    ws,
    cycle_type: str,
    cycle_code: str,
    start_date: str,
    end_date: str,
    rows: list[dict],
    source_lookup: dict[tuple[str, str, str, str], tuple[str, str]] | None = None,
) -> SheetWriteResult:
    result = _upsert_rows(ws, INVENTORY_KEY_FIELDS, cycle_type, cycle_code, start_date, end_date, rows, source_lookup)
    result.sheet_name = "7_库存补货"
    return result
