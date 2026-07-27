from __future__ import annotations

import argparse
from datetime import date, datetime, timedelta
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

from src.excel_utils import LOGICAL_BUSINESS_COLUMNS, build_sheet_context
from src.app_paths import app_base_dir, config_dir
from src.amazon_inventory import fill_amazon_inventory
from src.file_scanner import scan_excel_files
from src.fillers.kepule_filler import fill_kepule_inventory_sheet, fill_kepule_sales_sheet
from src.fillers.w27_filler import fill_w27_inventory_sheet, fill_w27_sku_detail_sheet
from src.models import Issue, SheetWriteResult
from src.reports import (
    build_recognition_rows,
    write_audit_report,
    write_delivery_summary,
    write_issue_report,
    write_recognition_report,
    write_review_report,
    write_run_log,
    write_status_details,
)
from src.role_detector import detect_workbook_role
from src.runtime import build_runtime_config, load_fill_rules, resolve_fill_config
from src.shop_mapping import load_shop_config
from src.source_builders import (
    build_special_sku_audit,
    build_kepule_inventory_rows,
    build_kepule_sales_rows,
    parse_number,
    build_w27_inventory_rows,
    build_w27_sku_rows,
    week_label_from_cycle,
)
from src.validators import (
    PRIMARY_SOURCE_ROLES,
    determine_status,
    determine_status_details,
    issue_requires_manual_review,
    role_display_name,
    validate_required_roles,
)
from src.workbook_copier import copy_workbook


EXPECTED_ROLES = ["w27_target", "kepule_target", *PRIMARY_SOURCE_ROLES, "jd_amazon_inventory", "domestic_sales_ranking"]


def run_amazon_inventory_fill(input_dir: str, output_dir: str) -> int:
    """Run the standalone three-file Amazon inventory workflow."""
    source_dir = Path(input_dir)
    destination = Path(output_dir)
    destination.mkdir(parents=True, exist_ok=True)
    detections = _detect_files([entry.path for entry in scan_excel_files(source_dir, destination, recursive=True).input_files])
    role_files = {role: detection.path for role, detection in detections.items() if detection.path}
    required = ("amazon_inventory_target", "amazon_inventory_weekly", "fba_inventory")
    missing = [role_display_name(role) for role in required if role not in role_files]
    if missing:
        raise ValueError("亚马逊库存独立填写缺少：" + "、".join(missing))
    target = role_files["amazon_inventory_target"]
    output = destination / f"{target.stem}_已填报{target.suffix}"
    copy_workbook(target, output)
    issues: list[Issue] = []
    counts = fill_amazon_inventory(output, role_files["amazon_inventory_weekly"], role_files["fba_inventory"], issues)
    write_issue_report(destination / "建议人工复核.xlsx", issues)
    write_recognition_report(destination / "文件识别报告.xlsx", build_recognition_rows(detections, {}, list(required)))
    (destination / "status.txt").write_text(
        f"亚马逊库存已写入 {counts['written_rows']} 行。\n"
        f"缺每周更新站点 {counts['missing_weekly']}；缺 FBA {counts['missing_fba']}；新增 SKU 行 {counts['added_weekly_rows']}；清空旧 SKU 行 {counts['cleared_target_rows']}。\n",
        encoding="utf-8",
    )
    return 0


def _workflow_config(input_dir: str, output_dir: str, values: dict[str, str | None]):
    runtime = build_runtime_config(
        input_dir, output_dir, values.get("cycle_type"), values.get("cycle_code"),
        values.get("start_date"), values.get("end_date"),
    )
    rules = load_fill_rules(config_dir() / "fill_rules.yaml")
    return runtime, resolve_fill_config(cli_values=values, rule_values=rules, inferred_values={})


def run_ecommerce_fill(input_dir: str, output_dir: str, **values: str | None) -> int:
    """Fill only the e-commerce analysis workbook and its two business sheets."""
    runtime, fill_config = _workflow_config(input_dir, output_dir, values)
    runtime.output_dir.mkdir(parents=True, exist_ok=True)
    detections = _detect_files([entry.path for entry in scan_excel_files(runtime.input_dir, runtime.output_dir).input_files])
    if "w27_target" not in detections:
        raise ValueError("电商数据表填写缺少：W27周电商数据分析表目标模板")
    issues: list[Issue] = []
    role_files = {role: item.path for role, item in detections.items() if item.path}
    shop_config = load_shop_config(config_dir() / "shop_mapping.yaml")
    base = detections["w27_target"].path
    output = runtime.output_dir / _target_output_name("w27_target", "SUCCESS", fill_config.cycle_code)
    reference = _reference_target_path(base)
    existing_inventory = _collect_existing_keys(reference, "7_库存补货", ["周期", "平台", "店铺名称", "SKU编码"])
    copy_workbook(base, output)
    sku_build = build_w27_sku_rows(role_files, issues, fill_config.cycle_code, shop_config)
    sort_w27_rows_by_channel(sku_build.rows)
    allowed = {(row["平台"], row["店铺名称"], row["SKU编码"]) for row in sku_build.rows}
    inventory_build = build_w27_inventory_rows(
        role_files, issues, fill_config.cycle_code, shop_config,
        {key: value for key, value in existing_inventory.items() if key[0] == fill_config.cycle_code},
        allowed_sku_keys=allowed,
    )
    sort_w27_rows_by_channel(inventory_build.rows)
    wb = load_workbook(output)
    fill_w27_sku_detail_sheet(wb["3_SKU明细"], fill_config.cycle_type, fill_config.cycle_code, fill_config.start_date, fill_config.end_date, sku_build.rows, sku_build.source_lookup)
    fill_w27_inventory_sheet(wb["7_库存补货"], fill_config.cycle_type, fill_config.cycle_code, fill_config.start_date, fill_config.end_date, inventory_build.rows, inventory_build.source_lookup)
    wb.save(output)
    wb.close()
    write_issue_report(runtime.output_dir / "异常清单.xlsx", issues)
    write_review_report(runtime.output_dir / "建议人工复核.xlsx", _build_review_rows(issues, sku_build.audit_rows))
    write_recognition_report(runtime.output_dir / "文件识别报告.xlsx", build_recognition_rows(detections, {}, ["w27_target", *PRIMARY_SOURCE_ROLES]))
    return 0


def run_kepule_fill(input_dir: str, output_dir: str, **values: str | None) -> int:
    """Fill only the Kepule week/month workbook and its source sheets."""
    runtime, fill_config = _workflow_config(input_dir, output_dir, values)
    runtime.output_dir.mkdir(parents=True, exist_ok=True)
    detections = _detect_files([entry.path for entry in scan_excel_files(runtime.input_dir, runtime.output_dir).input_files])
    if "kepule_target" not in detections:
        raise ValueError("周月报填写缺少：开普乐周月报统一数据源目标模板")
    issues: list[Issue] = []
    role_files = {role: item.path for role, item in detections.items() if item.path}
    shop_config = load_shop_config(config_dir() / "shop_mapping.yaml")
    base = detections["kepule_target"].path
    output = runtime.output_dir / _target_output_name("kepule_target", "SUCCESS", fill_config.cycle_code, fill_config.kepule_sales_date)
    reference = _reference_target_path(base)
    previous_suppliers = _collect_prior_inventory_suppliers(reference, fill_config.kepule_inventory_date)
    copy_workbook(base, output)
    sales_build = build_kepule_sales_rows(role_files, issues, fill_config.kepule_sales_date, shop_config)
    inventory_build = build_kepule_inventory_rows(role_files, issues, fill_config.kepule_inventory_date, previous_supplier_by_display_name=previous_suppliers)
    sort_kepule_sales_rows_by_channel(sales_build.rows)
    wb = load_workbook(output)
    week_label = week_label_from_cycle(fill_config.cycle_code)
    fill_kepule_sales_sheet(wb["源_销售明细"], fill_config.kepule_sales_date, week_label, sales_build.rows, sales_build.source_lookup)
    fill_kepule_inventory_sheet(wb["源_库存快照"], fill_config.kepule_inventory_date, week_label, inventory_build.rows, inventory_build.source_lookup)
    if "控制台" in wb.sheetnames:
        write_kepule_console_sheet(wb["控制台"], fill_config.kepule_sales_date, fill_config.end_date)
    write_kepule_top10_sections(wb, fill_config.start_date, fill_config.end_date, fill_config.kepule_sales_date)
    wb.calculation.calcMode = "auto"
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(output)
    wb.close()
    write_issue_report(runtime.output_dir / "异常清单.xlsx", issues)
    write_review_report(runtime.output_dir / "建议人工复核.xlsx", _build_review_rows(issues, []))
    write_recognition_report(runtime.output_dir / "文件识别报告.xlsx", build_recognition_rows(detections, {}, ["kepule_target", *PRIMARY_SOURCE_ROLES]))
    return 0


W27_CHANNEL_ORDER = ("天猫", "拼多多", "京东POP", "京东自营")

KEPULE_SALES_CHANNEL_ORDER = (
    "亚马逊北美",
    "亚马逊日本",
    "亚马逊欧洲",
    "天猫",
    "拼多多",
    "京东POP",
    "小红书",
    "得物",
    "视频号",
    "抖音2店",
    "京东自营",
    "商务部",
    "达人带货合计",
)

def sort_w27_rows_by_channel(rows: list[dict]) -> None:
    """Group W27 rows by the confirmed business channel order, stably."""
    channel_rank = {channel: index for index, channel in enumerate(W27_CHANNEL_ORDER)}
    rows.sort(key=lambda row: channel_rank.get(str(row.get("平台", "")).strip(), len(channel_rank)))


def sort_kepule_sales_rows_by_channel(rows: list[dict]) -> None:
    """Group by the report's business channel order, stably in-group."""
    channel_rank = {channel: index for index, channel in enumerate(KEPULE_SALES_CHANNEL_ORDER)}
    rows.sort(key=lambda row: channel_rank.get(str(row.get("渠道_标准", "")).strip(), len(channel_rank)))


def build_cli_values(gui_values: dict[str, str]) -> dict[str, str]:
    return {
        "cycle_type": gui_values.get("周期类型", ""),
        "cycle_code": gui_values.get("周次编码", ""),
        "start_date": gui_values.get("开始日期", ""),
        "end_date": gui_values.get("结束日期", ""),
        "inventory_date": gui_values.get("库存快照日期", ""),
        "kepule_sales_date": gui_values.get("开普乐销售日期", ""),
        "kepule_inventory_date": gui_values.get("开普乐库存日期", ""),
    }


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--cycle-type", default=None)
    parser.add_argument("--cycle-code", default=None)
    parser.add_argument("--start-date", default=None)
    parser.add_argument("--end-date", default=None)
    parser.add_argument("--inventory-date", default=None)
    parser.add_argument("--kepule-sales-date", default=None)
    parser.add_argument("--kepule-inventory-date", default=None)
    return parser.parse_args()


def _detect_files(paths: list[Path]) -> dict[str, object]:
    detections = {}
    for path in paths:
        # A user can select a temporary, damaged, or non-Excel file renamed
        # with an .xlsx suffix.  Recognition must be isolated per file: one
        # unreadable file must not abort the GUI refresh after it has cleared
        # the visible file-slot rows.
        try:
            detection = detect_workbook_role(path)
        except Exception:
            continue
        if detection.role != "unknown" and detection.role not in detections:
            detections[detection.role] = detection
    return detections


def _find_reusable_targets(partial_paths: list[Path]) -> dict[str, Path]:
    reusable = {}
    for path in partial_paths:
        detection = detect_workbook_role(path)
        if detection.role in {"w27_target", "kepule_target"}:
            reusable[detection.role] = path
    return reusable


def _target_output_name(role: str, status: str, cycle_code: str, report_date: str | None = None) -> str:
    suffix = "已填报" if status in {"SUCCESS", "SUCCESS_WITH_REVIEW"} else "部分填报"
    if role == "w27_target":
        return f"{week_label_from_cycle(cycle_code)}周 电商数据分析表_{suffix}.xlsx"
    # The reporting workbook name follows the month the user selected for
    # sales, rather than the month embedded in the starter template name.
    parsed = _parse_date(report_date) if report_date else datetime.now()
    return f"开普乐{parsed.year % 100:02d}年{parsed.month}月周月报统一数据源_{suffix}.xlsx"


def _base_target_path(role: str, detections, reusable_targets) -> Path | None:
    if role in reusable_targets:
        return reusable_targets[role]
    detection = detections.get(role)
    return detection.path if detection else None


def month_start_from_date(date_value: str) -> str:
    parsed = _parse_date(date_value)
    return parsed.replace(day=1).strftime("%Y-%m-%d")


def month_end_from_date(date_value: str) -> str:
    parsed = _parse_date(date_value)
    next_month = (parsed.replace(day=28) + timedelta(days=4)).replace(day=1)
    return (next_month - timedelta(days=1)).strftime("%Y-%m-%d")


def _parse_date(value) -> datetime:
    if isinstance(value, datetime):
        return value
    text = str(value or "").strip()
    for fmt in ("%Y-%m-%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return datetime.fromisoformat(text)


def write_kepule_console_sheet(ws, kepule_start_date: str, end_date: str) -> None:
    # The console keeps the immediately preceding report period in B7/B8.
    # Shift the current period before replacing B5/B6 so weekly runs remain
    # continuous instead of leaving an older template period behind.
    previous_start = ws["B5"].value
    previous_end = ws["B6"].value
    if previous_start not in (None, ""):
        ws["B7"] = _parse_date(previous_start).date()
    if previous_end not in (None, ""):
        ws["B8"] = _parse_date(previous_end).date()
    ws["B5"] = _parse_date(kepule_start_date).date()
    ws["B6"] = _parse_date(end_date).date()
    ws["B11"] = _parse_date(month_start_from_date(kepule_start_date)).date()


def _load_kepule_sales_rows(ws) -> list[dict]:
    context = build_sheet_context(ws, ["日期", "渠道_标准", "SKU", "产品名称", "销量", "销售额_元", "销售成本_元"])
    rows: list[dict] = []
    for values in ws.iter_rows(min_row=context.data_start_row, values_only=True):
        row = {}
        has_value = False
        for field, column in context.columns.items():
            value = values[column - 1] if column - 1 < len(values) else None
            row[field] = value
            if value not in (None, ""):
                has_value = True
        if has_value:
            rows.append(row)
    return rows


def build_kepule_top10_rows(rows: list[dict], start_date: str, end_date: str) -> list[dict]:
    start = _parse_date(start_date).date()
    end = _parse_date(end_date).date()
    grouped: dict[tuple[str, str], dict] = {}
    for row in rows:
        date_text = str(row.get("日期") or "").strip()
        if not date_text:
            continue
        row_date = _parse_date(row.get("日期")).date()
        if not (start <= row_date <= end):
            continue
        sku = str(row.get("SKU") or "").strip()
        product_name = str(row.get("产品名称") or "").strip() or sku
        key = (sku, product_name)
        bucket = grouped.setdefault(
            key,
            {
                "SKU": sku,
                "产品名称": product_name,
                "销量": 0.0,
                "销售额_元": 0.0,
                "销售成本": 0.0,
                "毛利_元": 0.0,
                "channel_sales": {},
            },
        )
        qty = parse_number(row.get("销量")) or 0.0
        sales_amount = parse_number(row.get("销售额_元")) or 0.0
        sales_cost = parse_number(row.get("销售成本_元")) or 0.0
        gross_profit = parse_number(row.get("毛利_元"))
        if gross_profit is None:
            gross_profit = sales_amount - sales_cost
        bucket["销量"] += qty
        bucket["销售额_元"] += sales_amount
        bucket["销售成本"] += sales_cost
        bucket["毛利_元"] += gross_profit
        channel = str(row.get("渠道_标准") or "").strip()
        site = str(row.get("国家/站点") or "").strip()
        display_channel = site if channel.startswith("亚马逊") and site else (channel or "需确认")
        bucket["channel_sales"][display_channel] = bucket["channel_sales"].get(display_channel, 0.0) + sales_amount

    top_rows = []
    ordered = sorted(grouped.values(), key=lambda item: (-item["销售额_元"], item["SKU"], item["产品名称"]))
    for index, item in enumerate(ordered[:10], start=1):
        main_channel = sorted(item["channel_sales"].items(), key=lambda kv: (-kv[1], kv[0]))[0][0] if item["channel_sales"] else "需确认"
        gross_profit = item["毛利_元"]
        sales_amount = item["销售额_元"]
        top_rows.append(
            {
                "排名": index,
                # The target header means “SKU or product name”.  Product
                # names are the human-facing display identity in the manual
                # weekly/monthly reports; fall back to SKU only when absent.
                "SKU/产品名": item["产品名称"] or item["SKU"],
                "主销渠道": main_channel,
                "销量": int(item["销量"]) if float(item["销量"]).is_integer() else item["销量"],
                "销售额_元": int(sales_amount) if float(sales_amount).is_integer() else sales_amount,
                "销售成本": int(item["销售成本"]) if float(item["销售成本"]).is_integer() else item["销售成本"],
                "毛利_元": int(gross_profit) if float(gross_profit).is_integer() else gross_profit,
                "毛利率": (gross_profit / sales_amount) if sales_amount not in (None, 0) else None,
                "备注": "",
            }
        )
    return top_rows


def _write_top10_block(ws, header_row: int, data_start_row: int, rows: list[dict]) -> None:
    headers = [ws.cell(header_row, col).value for col in range(1, 9)]
    for row_index in range(10):
        target_row = data_start_row + row_index
        payload = rows[row_index] if row_index < len(rows) else {}
        rank = row_index + 1
        for col, header in enumerate(headers, start=1):
            if header == "排名":
                ws.cell(target_row, col).value = rank
            elif header == "SKU/产品名":
                ws.cell(target_row, col).value = payload.get("SKU/产品名")
            elif header == "主销渠道":
                ws.cell(target_row, col).value = payload.get("主销渠道")
            elif header == "销量":
                ws.cell(target_row, col).value = payload.get("销量")
            elif header == "销售额_元":
                ws.cell(target_row, col).value = payload.get("销售额_元")
            elif header == "销售成本":
                ws.cell(target_row, col).value = payload.get("销售成本")
            elif header == "毛利_元":
                ws.cell(target_row, col).value = payload.get("毛利_元")
            elif header == "平台利润_元":
                ws.cell(target_row, col).value = payload.get("毛利_元")
            elif header == "毛利率":
                ws.cell(target_row, col).value = payload.get("毛利率")
            elif header == "备注":
                ws.cell(target_row, col).value = payload.get("备注")


def _copy_block_values(source_ws, target_ws, source_start_row: int, source_end_row: int, target_start_row: int) -> None:
    for offset, source_row in enumerate(range(source_start_row, source_end_row + 1)):
        target_row = target_start_row + offset
        for col in range(1, 9):
            target_ws.cell(target_row, col).value = source_ws.cell(source_row, col).value


def write_kepule_top10_sections(workbook, start_date: str, end_date: str, sales_date: str) -> None:
    sales_rows = _load_kepule_sales_rows(workbook["源_销售明细"])
    week_rows = build_kepule_top10_rows(sales_rows, start_date, end_date)
    month_anchor = sales_date or start_date
    month_rows = build_kepule_top10_rows(sales_rows, month_start_from_date(month_anchor), month_end_from_date(month_anchor))

    week_agg = workbook["周报聚合"]
    month_agg = workbook["月报聚合"]
    _write_top10_block(week_agg, header_row=29, data_start_row=30, rows=week_rows)
    _write_top10_block(month_agg, header_row=101, data_start_row=102, rows=month_rows)

    if "周报输出" in workbook.sheetnames:
        _copy_block_values(week_agg, workbook["周报输出"], source_start_row=28, source_end_row=39, target_start_row=32)
    if "月报输出" in workbook.sheetnames:
        _copy_block_values(month_agg, workbook["月报输出"], source_start_row=100, source_end_row=111, target_start_row=111)


def _reference_target_path(base_path: Path) -> Path:
    canonical_input = Path("input") / base_path.name
    if canonical_input.exists() and canonical_input.resolve() != base_path.resolve():
        return canonical_input
    return base_path


def _collect_existing_keys(path: Path, sheet_name: str, key_fields: list[str]) -> dict[tuple[str, ...], dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=False)
    ws = workbook[sheet_name]
    context = build_sheet_context(ws, key_fields)
    business_indexes = [column_index_from_string(letter) for letter in LOGICAL_BUSINESS_COLUMNS[sheet_name]]
    rows: dict[tuple[str, ...], dict[str, str]] = {}
    for row_index, values in enumerate(ws.iter_rows(min_row=context.data_start_row, values_only=True), start=context.data_start_row):
        if all((values[index - 1] if index - 1 < len(values) else None) in (None, "") for index in business_indexes):
            continue
        key = tuple(str(values[context.columns[field] - 1] or "") for field in key_fields)
        if not any(key):
            continue
        row_data = {}
        for field, column in context.columns.items():
            value = values[column - 1] if column - 1 < len(values) else None
            row_data[field] = "" if value is None else str(value)
        rows[key] = row_data
    workbook.close()
    return rows


def _as_iso_date(value: object) -> str:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return str(value or "").strip()[:10]


def _collect_prior_inventory_suppliers(path: Path, before_date: str) -> dict[str, str]:
    """Return only unambiguous domestic supplier history for exact display names."""
    workbook = load_workbook(path, read_only=True, data_only=False)
    ws = workbook["源_库存快照"]
    # The business workbook does not expose “仓库/区域” in this source sheet;
    # its existing supplier channel is the historical domestic identity.
    context = build_sheet_context(ws, ["日期", "SKU/分类", "渠道_标准"])
    suppliers_by_name: dict[str, set[str]] = {}
    for values in ws.iter_rows(min_row=context.data_start_row, values_only=True):
        row_date = _as_iso_date(values[context.columns["日期"] - 1])
        if not row_date or row_date >= before_date:
            continue
        display_name = str(values[context.columns["SKU/分类"] - 1] or "").strip()
        supplier = str(values[context.columns["渠道_标准"] - 1] or "").strip()
        if not display_name or not supplier:
            continue
        suppliers_by_name.setdefault(display_name, set()).add(supplier)
    workbook.close()
    return {name: next(iter(suppliers)) for name, suppliers in suppliers_by_name.items() if len(suppliers) == 1}


def _apply_builder_stats(sheet_result: SheetWriteResult, stats: dict[str, int]) -> None:
    sheet_result.raw_candidate_count = stats.get("raw_candidate_count", 0)
    sheet_result.skip_zero_sales_count = stats.get("skip_zero_sales_count", 0)
    sheet_result.skip_no_effect_count = stats.get("skip_no_effect_count", 0)
    sheet_result.candidate_after_filter_count = stats.get("candidate_after_filter_count", 0)


def _collect_w27_formula_samples(path: Path) -> list[str]:
    workbook = load_workbook(path, read_only=False, data_only=False)
    ws = workbook["3_SKU明细"]
    sample_row = 241 if ws.max_row >= 241 else ws.max_row
    formulas = {
        "净销售额公式": ws[f"P{sample_row}"].value,
        "商品总成本公式": ws[f"AK{sample_row}"].value,
        "毛利额公式": ws[f"AL{sample_row}"].value,
        "毛利率公式": ws[f"AM{sample_row}"].value,
        "金额退款率公式": ws[f"AF{sample_row}"].value,
    }
    workbook.close()
    return [f"{label}={value}" for label, value in formulas.items()]


def _build_w27_diff_rows(
    sheet_name: str,
    cycle_code: str,
    key_fields: list[str],
    row_payloads: list[dict],
    audit_rows: list[dict],
    source_lookup: dict[tuple[str, ...], tuple[str, str]],
    sheet_result: SheetWriteResult,
    existing_rows: dict[tuple[str, ...], dict[str, str]],
) -> list[dict]:
    rows: list[dict] = []
    written_keys = {
        (event.key_values.get("周期", ""), event.key_values.get("平台", ""), event.key_values.get("店铺名称", ""), event.sku): event
        for event in sheet_result.row_events
    }
    candidate_keys = set()
    for payload in row_payloads:
        if key_fields[0] == "周期":
            key = (cycle_code, str(payload.get("平台", "")), str(payload.get("店铺名称", "")), str(payload.get("SKU编码", "")))
        else:
            key = tuple(str(payload.get(field, "") or "") for field in key_fields)
        candidate_keys.add(key)
        source_file, source_sheet = source_lookup.get(key, ("", ""))
        event = written_keys.get(key)
        original_exists = "是" if key in existing_rows else "否"
        this_run_written = "是" if event else "否"
        diff_type = "原有保留" if key in existing_rows and not event else "新增写入" if key not in existing_rows and event else "原有保留"
        source_field = (
            "近30天销量=来源字段优先；若缺失则按同业务键使用补充来源"
            if sheet_name == "7_库存补货"
            else "SKU编码=来源字段原样映射；店铺名称=通过店铺映射规则归一后写入；商品名称=当前来源字段优先，缺失时按同SKU补充；商品分类=销售源优先，缺失时按同SKU使用商品档案补充；销售额/成交件数/单件成本按已确认字段映射写入"
        )
        key_value = payload.get("近30天销量") if sheet_name == "7_库存补货" else payload.get("销售额")
        rows.append(
            {
                "目标Sheet": sheet_name,
                "平台": payload.get("平台", ""),
                "店铺名称": payload.get("店铺名称", ""),
                "SKU编码": payload.get("SKU编码", ""),
                "商品名称": payload.get("商品名称", ""),
                "原目标表是否已有": original_exists,
                "本次是否写入": this_run_written,
                "差异类型": diff_type,
                "来源文件": source_file,
                "来源Sheet": source_sheet,
                "来源字段": source_field,
                "关键值": "" if key_value is None else str(key_value),
                "说明": event.reason if event else "",
            }
        )
    rows.extend(audit_rows)
    for key, original in existing_rows.items():
        if key not in candidate_keys:
            rows.append(
                {
                    "目标Sheet": sheet_name,
                    "平台": original.get("平台", ""),
                    "店铺名称": original.get("店铺名称", ""),
                    "SKU编码": original.get("SKU编码", ""),
                    "商品名称": original.get("商品名称", ""),
                    "原目标表是否已有": "是",
                    "本次是否写入": "否",
                    "差异类型": "原有但本次未写入",
                    "来源文件": "",
                    "来源Sheet": "",
                    "来源字段": "",
                    "关键值": "",
                    "说明": "原目标表存在，但本次候选中没有对应记录。",
                }
            )
    return rows


def _collect_missing_source_names(issues: list[Issue]) -> list[str]:
    names = []
    for issue in issues:
        if issue.module == "validator" and "缺少“" in issue.message:
            names.append(issue.message.replace("缺少“", "").replace("”来源文件", "").replace("”", ""))
    return sorted(set(names))


def _build_review_rows(issues: list[Issue], special_audit_rows: list[dict]) -> list[dict]:
    product_name_by_sku = {str(row.get("SKU编码", "")): str(row.get("商品名称", "")) for row in special_audit_rows}
    rows: list[dict] = []
    for issue in issues:
        if not issue_requires_manual_review(issue):
            continue
        sku = issue.sku
        product_name = product_name_by_sku.get(sku, "")
        note = issue.message
        action = issue.suggestion or "请结合来源表人工确认。"
        result_table = "W27周电商数据分析表" if issue.target_table == "W27" else "开普乐周月报统一数据源" if issue.target_table == "开普乐" else "结果文件"
        result_sheet = issue.target_field or "待确认"
        result_location = f"SKU {sku}" if sku else "请按异常说明查找"
        source_sheet = issue.sheet or "Sheet待确认"
        source_row = issue.row_number or ""
        source_field = issue.field or "待确认字段"
        source_location = f"{issue.file_name or '来源表待确认'} / {source_sheet} / 行{source_row or '待确认'} / 字段：{source_field}"
        if sku == "6959291008957":
            note = "来源表真实异常；已正常写入结果表；建议人工确认销售金额 / 销售数量 / 销售成本是否真实"
            action = "打开来源表原始记录，核对销售金额、销售数量、销售成本是否为真实业务数据。"
        rows.append(
            {
                "结果表": result_table,
                "结果Sheet": result_sheet,
                "结果定位": result_location,
                "结果字段": issue.field,
                "SKU": sku,
                "商品名称": product_name,
                "来源文件": issue.file_name,
                "来源Sheet": source_sheet,
                "来源行号": source_row,
                "来源字段": source_field,
                "来源定位": source_location,
                "异常说明": note,
                "建议核对动作": action,
            }
        )
    return rows


def _print_run_intro(fill_config, detections: dict[str, object], reusable_targets: dict[str, Path], issues: list[Issue], output_dir: Path) -> None:
    target_count = sum(1 for role in ["w27_target", "kepule_target"] if role in detections or role in reusable_targets)
    source_count = sum(1 for role in detections.keys() if role not in {"w27_target", "kepule_target"})
    missing = _collect_missing_source_names(issues)
    print(f"本次周期：{fill_config.cycle_code}")
    print(f"已识别到目标模板：{target_count} 份")
    print(f"已识别到来源表：{source_count} 份")
    if missing:
        print("缺失的来源表：" + "、".join(missing))
    else:
        print("缺失的来源表：无")
    print(f"将输出到：{output_dir}")


def _print_run_summary(status_details: dict[str, object], primary_files: list[Path], review_rows: list[dict]) -> None:
    print("")
    print(f"核心填表状态：{status_details['core_fill_status']}")
    print("已生成结果文件：")
    for path in primary_files:
        print(f"- {path.name}")
    if review_rows:
        print(f"存在建议人工复核项：{len(review_rows)} 条，请查看 建议人工复核.xlsx")
    else:
        print("本次没有建议人工复核项。")


def run_fill(
    input_dir: str,
    output_dir: str,
    cycle_type: str | None = None,
    cycle_code: str | None = None,
    start_date: str | None = None,
    end_date: str | None = None,
    inventory_date: str | None = None,
    kepule_sales_date: str | None = None,
    kepule_inventory_date: str | None = None,
) -> int:
    """Run the fill pipeline for either the CLI or the desktop GUI."""
    args = argparse.Namespace(
        input_dir=input_dir,
        output_dir=output_dir,
        cycle_type=cycle_type,
        cycle_code=cycle_code,
        start_date=start_date,
        end_date=end_date,
        inventory_date=inventory_date,
        kepule_sales_date=kepule_sales_date,
        kepule_inventory_date=kepule_inventory_date,
    )
    runtime = build_runtime_config(args.input_dir, args.output_dir, args.cycle_type, args.cycle_code, args.start_date, args.end_date)
    rules = load_fill_rules(config_dir() / "fill_rules.yaml")
    shop_config = load_shop_config(config_dir() / "shop_mapping.yaml")
    fill_config = resolve_fill_config(
        cli_values={
            "cycle_type": args.cycle_type,
            "cycle_code": args.cycle_code,
            "start_date": args.start_date,
            "end_date": args.end_date,
            "inventory_date": args.inventory_date,
            "kepule_sales_date": args.kepule_sales_date,
            "kepule_inventory_date": args.kepule_inventory_date,
        },
        rule_values=rules,
        inferred_values={},
    )

    runtime.output_dir.mkdir(parents=True, exist_ok=True)
    log_dir = app_base_dir() / "logs"
    log_dir.mkdir(parents=True, exist_ok=True)

    scan_result = scan_excel_files(runtime.input_dir, runtime.output_dir)
    detections = _detect_files([entry.path for entry in scan_result.input_files])
    reusable_targets = _find_reusable_targets([entry.path for entry in scan_result.partial_output_files])
    issues: list[Issue] = validate_required_roles(detections, reusable_targets)
    _print_run_intro(fill_config, detections, reusable_targets, issues, runtime.output_dir)
    status_details = determine_status_details(issues)
    status = str(status_details["status"])
    sheet_results: list[SheetWriteResult] = []
    audit_report_rows: dict[str, list[dict]] = {}
    special_audit_rows: list[dict] = []

    if status != "FAILED":
        w27_base = _base_target_path("w27_target", detections, reusable_targets)
        kepule_base = _base_target_path("kepule_target", detections, reusable_targets)
        w27_reference = _reference_target_path(w27_base)
        kepule_reference = _reference_target_path(kepule_base)
        w27_output = runtime.output_dir / _target_output_name("w27_target", status, fill_config.cycle_code)
        kepule_output = runtime.output_dir / _target_output_name("kepule_target", status, fill_config.cycle_code, fill_config.kepule_sales_date)

        existing_w27_sku = _collect_existing_keys(w27_reference, "3_SKU明细", ["周期", "平台", "店铺名称", "SKU编码"])
        existing_w27_inventory = _collect_existing_keys(w27_reference, "7_库存补货", ["周期", "平台", "店铺名称", "SKU编码"])
        existing_kepule_sales = _collect_existing_keys(kepule_reference, "源_销售明细", ["日期", "渠道_标准", "店铺/客户", "SKU"])
        previous_inventory_suppliers = _collect_prior_inventory_suppliers(kepule_reference, fill_config.kepule_inventory_date)
        copy_workbook(w27_base, w27_output)
        copy_workbook(kepule_base, kepule_output)

        role_files = {role: detection.path for role, detection in detections.items() if detection.path}
        w27_sku_build = build_w27_sku_rows(role_files, issues, fill_config.cycle_code, shop_config)
        sort_w27_rows_by_channel(w27_sku_build.rows)
        allowed_inventory_keys = {(row["平台"], row["店铺名称"], row["SKU编码"]) for row in w27_sku_build.rows}
        w27_inventory_build = build_w27_inventory_rows(
            role_files,
            issues,
            fill_config.cycle_code,
            shop_config,
            {key: value for key, value in existing_w27_inventory.items() if key[0] == fill_config.cycle_code},
            allowed_sku_keys=allowed_inventory_keys,
        )
        kepule_sales_build = build_kepule_sales_rows(role_files, issues, fill_config.kepule_sales_date, shop_config)
        kepule_inventory_build = build_kepule_inventory_rows(
            role_files,
            issues,
            fill_config.kepule_inventory_date,
            previous_supplier_by_display_name=previous_inventory_suppliers,
        )
        sort_w27_rows_by_channel(w27_inventory_build.rows)
        sort_kepule_sales_rows_by_channel(kepule_sales_build.rows)
        # 源_库存快照 follows the product archive from top to bottom.  Do not
        # reorder by supplier/SKU: that destroys the established source order
        # and can shift row-level manual classifications onto another product.

        week_label = week_label_from_cycle(fill_config.cycle_code)

        w27_wb = load_workbook(w27_output)
        w27_sku_result = fill_w27_sku_detail_sheet(
            w27_wb["3_SKU明细"],
            fill_config.cycle_type,
            fill_config.cycle_code,
            fill_config.start_date,
            fill_config.end_date,
            w27_sku_build.rows,
            w27_sku_build.source_lookup,
        )
        w27_sku_result.workbook_role = "w27_target"
        _apply_builder_stats(w27_sku_result, w27_sku_build.stats)
        w27_sku_result.filled_fields = ["周期类型", "周期", "开始日期", "结束日期", "平台", "店铺名称", "商品名称", "SKU编码", "商品分类", "销售额", "成交件数", "退款金额", "单件成本"]
        w27_sku_result.blank_fields = ["SPU编码", "分层(运营判断)", "售价", "订单数", "支付买家数", "访客数UV", "浏览量PV", "点击率", "加购人数", "收藏人数", "新客数", "老客数", "退款订单数", "平台活动标记"]
        sheet_results.append(w27_sku_result)

        w27_inventory_result = fill_w27_inventory_sheet(
            w27_wb["7_库存补货"],
            fill_config.cycle_type,
            fill_config.cycle_code,
            fill_config.start_date,
            fill_config.end_date,
            w27_inventory_build.rows,
            w27_inventory_build.source_lookup,
        )
        w27_inventory_result.workbook_role = "w27_target"
        _apply_builder_stats(w27_inventory_result, w27_inventory_build.stats)
        w27_inventory_result.filled_fields = ["周期类型", "周期", "开始日期", "结束日期", "平台", "店铺名称", "商品名称", "SKU编码", "在库库存", "在途库存", "近7天销量", "近30天销量"]
        w27_inventory_result.blank_fields = ["锁定库存", "安全库存天数", "建议补货量", "建议补货日期", "备注"]
        sheet_results.append(w27_inventory_result)
        if w27_sku_result.written_rows or w27_inventory_result.written_rows or w27_sku_result.backfilled_rows or w27_inventory_result.backfilled_rows:
            w27_wb.save(w27_output)
        w27_wb.close()

        kepule_wb = load_workbook(kepule_output)
        kepule_sales_result = fill_kepule_sales_sheet(
            kepule_wb["源_销售明细"],
            fill_config.kepule_sales_date,
            week_label,
            kepule_sales_build.rows,
            kepule_sales_build.source_lookup,
        )
        kepule_sales_result.workbook_role = "kepule_target"
        _apply_builder_stats(kepule_sales_result, kepule_sales_build.stats)
        kepule_sales_result.filled_fields = ["日期", "渠道_标准", "渠道_原始", "国家/站点", "店铺/客户", "SKU", "产品名称", "产品分类", "销量", "销售额_元", "销售成本_元", "平台利润_元", "退款金额_元", "广告费_元", "备注"]
        kepule_sales_result.blank_fields = ["是否新品", "是否清库", "是否B2B大单"]
        sheet_results.append(kepule_sales_result)

        kepule_inventory_result = fill_kepule_inventory_sheet(
            kepule_wb["源_库存快照"],
            fill_config.kepule_inventory_date,
            week_label,
            kepule_inventory_build.rows,
            kepule_inventory_build.source_lookup,
        )
        kepule_inventory_result.workbook_role = "kepule_target"
        _apply_builder_stats(kepule_inventory_result, kepule_inventory_build.stats)
        kepule_inventory_result.filled_fields = ["日期", "SKU/分类", "仓库/区域", "渠道_标准", "库存数量", "库存金额_元", "可售天数", "类型", "备注"]
        kepule_inventory_result.blank_fields = ["库龄天数", "状态", "处置进度"]
        sheet_results.append(kepule_inventory_result)
        console_updated = False
        if "控制台" in kepule_wb.sheetnames:
            write_kepule_console_sheet(kepule_wb["控制台"], fill_config.kepule_sales_date, fill_config.end_date)
            console_updated = True
        # The blank template does not contain a complete TOP10 calculation
        # block. Build it from the filled sales source so weekly/monthly
        # outputs are complete for a newly appended period.
        write_kepule_top10_sections(
            kepule_wb,
            fill_config.start_date,
            fill_config.end_date,
            fill_config.kepule_sales_date,
        )
        kepule_wb.calculation.calcMode = "auto"
        kepule_wb.calculation.fullCalcOnLoad = True
        kepule_wb.calculation.forceFullCalc = True
        if console_updated or kepule_sales_result.written_rows or kepule_inventory_result.written_rows or kepule_sales_result.backfilled_rows or kepule_inventory_result.backfilled_rows:
            kepule_wb.save(kepule_output)
        kepule_wb.close()

        audit_report_rows = {
            "W27_3_SKU明细_原有与重填差异": _build_w27_diff_rows(
                "3_SKU明细",
                fill_config.cycle_code,
                ["周期", "平台", "店铺名称", "SKU编码"],
                w27_sku_build.rows,
                w27_sku_build.audit_rows,
                w27_sku_build.source_lookup,
                w27_sku_result,
                {key: value for key, value in existing_w27_sku.items() if key[0] == fill_config.cycle_code},
            ),
            "W27_7_库存补货_原有与重填差异": _build_w27_diff_rows(
                "7_库存补货",
                fill_config.cycle_code,
                ["周期", "平台", "店铺名称", "SKU编码"],
                w27_inventory_build.rows,
                [row for row in w27_inventory_build.audit_rows if row.get("目标Sheet") == "7_库存补货"],
                w27_inventory_build.source_lookup,
                w27_inventory_result,
                {key: value for key, value in existing_w27_inventory.items() if key[0] == fill_config.cycle_code},
            ),
            "W27_近30天销量来源核对": [
                {
                    "平台": row.get("平台", ""),
                    "店铺名称": row.get("店铺名称", ""),
                    "SKU编码": row.get("SKU编码", ""),
                    "商品名称": row.get("商品名称", ""),
                    "近30天销量值": row.get("近30天销量值", ""),
                    "来源文件": row.get("近30天销量来源文件", ""),
                    "来源Sheet": row.get("近30天销量来源Sheet", ""),
                    "来源字段": row.get("近30天销量来源字段", ""),
                    "来源原始值": row.get("近30天销量来源原始值", ""),
                    "是否补充来源": row.get("近30天销量是否补充来源", ""),
                }
                for row in w27_inventory_build.audit_rows
                if "近30天销量来源字段" in row
            ],
            "W27_异常SKU源表对账": [],
            "开普乐_源销售明细_来源核对": [
                {
                    "日期": fill_config.kepule_sales_date,
                    "渠道_标准": row.get("渠道_标准", ""),
                    "店铺/客户": row.get("店铺/客户", ""),
                    "SKU": row.get("SKU", ""),
                    "产品名称": row.get("产品名称", ""),
                    "销售额_元": row.get("销售额_元", ""),
                    "销售成本_元": row.get("销售成本_元", ""),
                    "平台利润_元": row.get("平台利润_元", ""),
                    "退款金额_元": row.get("退款金额_元", ""),
                    "广告费_元": row.get("广告费_元", ""),
                    "来源文件": kepule_sales_build.source_lookup.get((fill_config.kepule_sales_date, row.get("渠道_标准", ""), row.get("店铺/客户", ""), row.get("SKU", "")), ("", ""))[0],
                    "来源Sheet": kepule_sales_build.source_lookup.get((fill_config.kepule_sales_date, row.get("渠道_标准", ""), row.get("店铺/客户", ""), row.get("SKU", "")), ("", ""))[1],
                    "来源字段": "销售额/销售成本/平台利润",
                    "说明": row.get("备注", ""),
                }
                for row in kepule_sales_build.rows
            ],
            "开普乐_源库存快照_来源核对": [
                {
                    "日期": fill_config.kepule_inventory_date,
                    "渠道_标准": row.get("渠道_标准", ""),
                    "仓库/区域": row.get("仓库/区域", ""),
                    "SKU/分类": row.get("SKU/分类", ""),
                    "库存数量": row.get("库存数量", ""),
                    "库存金额_元": row.get("库存金额_元", ""),
                    "可售天数": row.get("可售天数", ""),
                    "类型": row.get("类型", ""),
                    "来源文件": kepule_inventory_build.source_lookup.get((fill_config.kepule_inventory_date, row.get("SKU/分类", ""), row.get("仓库/区域", ""), row.get("渠道_标准", "")), ("", ""))[0],
                    "来源Sheet": kepule_inventory_build.source_lookup.get((fill_config.kepule_inventory_date, row.get("SKU/分类", ""), row.get("仓库/区域", ""), row.get("渠道_标准", "")), ("", ""))[1],
                    "来源字段": "库存数量/库存金额_元/可售天数",
                    "说明": row.get("备注", ""),
                }
                for row in kepule_inventory_build.rows
            ],
        }

        special_audit_rows.append(build_special_sku_audit(role_files, "6959291008957"))
        audit_report_rows["W27_异常SKU源表对账"].extend(special_audit_rows)

    status_details = determine_status_details(issues)
    status = str(status_details["status"])
    recognition_rows = build_recognition_rows(detections, reusable_targets, EXPECTED_ROLES)
    extra_log_lines = []
    if status != "FAILED":
        extra_log_lines.extend(
            [
                "W27_3_SKU明细_销售额=销售金额（未扣退款）",
                "W27_3_SKU明细_成交件数=销售数量（未扣退款）",
                "W27_3_SKU明细_退款金额=单独写入",
                "W27_3_SKU明细_净销售额=模板公式 销售额-退款金额",
                "W27_3_SKU明细_SKU编码=来源字段原样映射",
                "W27_3_SKU明细_店铺名称=通过店铺映射规则归一后写入",
                "W27_3_SKU明细_商品名称=当前来源字段优先；若缺失，则按同SKU从补充来源补齐",
                "W27_3_SKU明细_商品分类=销售源优先；若销售源缺失且商品档案存在同SKU记录，则用商品档案分类补充",
                "W27_7_库存补货=严格跟随本周期3_SKU明细业务键范围",
                "开普乐_TOP10=周报/月报均按源_销售明细全渠道统一按销售额排序",
                "开普乐_国内_平台利润来源字段=净销售毛利",
                "开普乐_跨境_平台利润来源字段=毛利润",
                "开普乐_跨境_销售成本来源字段=合计成本",
            ]
        )
        extra_log_lines.extend(_collect_w27_formula_samples(w27_reference))
    review_rows = _build_review_rows(issues, special_audit_rows)
    primary_delivery_files = [
        runtime.output_dir / _target_output_name("w27_target", status, fill_config.cycle_code),
        runtime.output_dir / _target_output_name("kepule_target", status, fill_config.cycle_code),
        runtime.output_dir / "建议人工复核.xlsx",
    ]
    technical_files = [
        "本周期写入差异核对.xlsx",
        "异常清单.xlsx",
        "status.txt",
        "run_log.txt",
    ]
    write_recognition_report(runtime.output_dir / "文件识别报告.xlsx", recognition_rows, fill_config)
    write_issue_report(runtime.output_dir / "异常清单.xlsx", issues)
    write_review_report(runtime.output_dir / "建议人工复核.xlsx", review_rows)
    write_audit_report(runtime.output_dir / "本周期写入差异核对.xlsx", audit_report_rows)
    write_status_details(
        runtime.output_dir / "status.txt",
        status,
        str(status_details["core_fill_status"]),
        bool(status_details["has_manual_review"]),
        int(status_details["manual_review_count"]),
    )
    write_delivery_summary(
        runtime.output_dir / "交付说明.txt",
        status,
        str(status_details["core_fill_status"]),
        bool(status_details["has_manual_review"]),
        [path.name for path in primary_delivery_files],
        technical_files,
    )
    write_run_log(log_dir / "run_log.txt", status, sheet_results, issues, fill_config, extra_log_lines)
    _print_run_summary(status_details, primary_delivery_files, review_rows)
    return 0


def main() -> int:
    return run_fill(**vars(parse_args()))


if __name__ == "__main__":
    raise SystemExit(main())
