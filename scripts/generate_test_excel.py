import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
import os

output_dir = r"e:\trea\AItool\test_files"
os.makedirs(output_dir, exist_ok=True)

header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(color="FFFFFF", bold=True)
example_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")

columns = [
    '名称', 'SKU', '条形码', '商品中文名称', '商品英文名称',
    '上架渠道', '售卖地区', '品牌', '系列', '系统分类',
    '商品分级', '上市时间', '生命周期', '负责人',
    '尺寸信息', '容量信息', '毛重(g)', '主体材质', '主色系',
    '表面处理', '适用热源', '功率（炉具类）', '技术优势',
    '认证信息', '使用说明',
    '核心卖点 TOP5', '目标人群', '差异化定位', '价格定位带',
    '情感价值', '使用场景', '竞品对标',
    '标题（英文）', '标题（中文）', '产品长描述（英文）', '产品长描述（中文）',
    '搜索关键词库',
]

# ============================================================
# 1. L1-L4 test file
# ============================================================
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "产品数据"

ws.append(["产品部"] + [""] * (len(columns) - 1))
ws.append(["请从第6行开始填写数据，SKU列不能为空"] + [""] * (len(columns) - 1))
ws.append(["", "L1", "", "", "", "L1", "L1", "L1", "L1", "L1",
           "L1", "L1", "L1", "L1",
           "L2", "L2", "L2", "L2", "L2", "L2", "L2", "L2", "L2",
           "L2", "L2",
           "L3", "L3", "L3", "L3", "L3", "L3", "L3",
           "L4", "L4", "L4", "L4", "L4"])

ws.append(columns)
for cell in ws[4]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center')

example_data = [
    '示例产品', 'CW-C01-37', '6920000000001', '野营锅7件套', 'Camping Cookware Set 7pcs',
    '淘宝,京东,Amazon', '中国,日本,美国', 'NatureHike', '挪客户外', '炊具',
    'A级', '2026/1/1', '成长期', '张三',
    '展开:14*28.5 cm\n收纳:14*16.5 cm\n锅:12.5*10 cm\n碗:12*5 cm',
    '锅900ml\n碗450ml', '1200', '铝合金', '黑色',
    '硬质氧化', '明火\n电磁炉\n电陶炉', '2000W',
    '1.无涂层更安全\n2.越用越好用\n3.导热均匀',
    'FDA\nLFGB\nSGS', '首次使用前请用中性洗涤剂清洗',
    '1.轻量化设计2.一锅多用3.快速导热4.易清洗5.耐用',
    '户外爱好者,家庭露营', '轻量化+多功能一体化', '200-400元',
    '自由探索', '家庭露营\n徒步旅行\n野餐', '火枫\n凯斯',
    'Ultralight Camping Cookware Set', '超轻户外野营锅具套装',
    'High quality camping cookware...', '高品质户外野营锅具套装...',
    'camping cookware\nultralight pot\noutdoor cooking',
]
ws.append(example_data)
for cell in ws[5]:
    cell.fill = example_fill

# Row 6: CW-C01-01 (will be existing draft, different data -> conflict)
ws.append([
    '测试产品1', 'CW-C01-01', '6920000000100', '测试产品1-新版', 'Test Product 1 New',
    '淘宝,京东', '中国,美国', 'TestBrand', 'TestSeries', '炊具',
    'B级', '2026/3/15', '成长期', '李四',
    '展开:20*30 cm\n收纳:15*20 cm', '锅1200ml', '800', '不锈钢', '银色',
    '抛光', '明火\n电磁炉', '最小功率：900W\n最大功率：3200W',
    '1.不锈钢材质\n2.耐用防腐',
    'FDA\nSGS', '使用前清洗',
    '1.不锈钢耐用\n2.导热性好\n3.易清洁\n4.大容量\n5.多功能',
    '家庭用户,户外爱好者', '不锈钢+大容量', '150-300元',
    '健康烹饪', '家庭烹饪\n户外露营', '品牌A\n品牌B',
    'Stainless Steel Cookware', '不锈钢炊具',
    'High quality stainless steel cookware...', '高品质不锈钢炊具...',
    'stainless steel\ncookware\nkitchen',
])

# Row 7: CW-C01-02 (will be existing draft, same data -> skip)
ws.append([
    '测试产品2', 'CW-C01-02', '', '测试产品2', 'Test2',
    '天猫', '中国', '', '', '茶具',
    '', '', '', '',
    '', '', '', '', '',
    '', '', '',
    '', '', '',
    '', '', '', '', '', '', '',
    '测试产品2', 'Test2', '', '',
    '',
])

# Row 8: CW-C01-03 (new SKU -> create)
ws.append([
    '全新产品', 'CW-C01-03', '6920000000200', '全新产品中文', 'New Product EN',
    'Amazon', '美国,欧洲', 'NewBrand', 'NewSeries', '户外家具',
    'A级', '2026/6/1', '导入期', '王五',
    '展开:50*50 cm', '容量5L', '2000', '钛合金', '白色',
    '磨砂', '明火', '最大功率：2250W',
    '1.钛合金轻量\n2.强度高\n3.耐腐蚀',
    'FDA\nCE', '轻拿轻放',
    '1.极致轻量\n2.超高强度\n3.耐腐蚀\n4.易携带\n5.多功能',
    '专业户外,极限运动', '钛合金+超轻量', '500-800元',
    '极致体验', '高山徒步\n极限探险', '竞品X\n竞品Y',
    'Ultralight Titanium Gear', '超轻钛合金装备',
    'Premium ultralight titanium outdoor gear...', '高端超轻钛合金户外装备...',
    'titanium\nultralight\noutdoor gear',
])

# Row 9: Empty SKU -> should stop parsing
ws.append(['', '', '', '应被忽略', 'Should be ignored'])

for col_idx in range(1, len(columns) + 1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 18

l1l4_path = os.path.join(output_dir, "L1L4_test_products.xlsx")
wb.save(l1l4_path)
print(f"L1-L4 test file saved: {l1l4_path}")

# ============================================================
# 2. L5 test file (CW-C01-01 knowledge base)
# ============================================================
wb2 = openpyxl.Workbook()
ws_qa = wb2.active
ws_qa.title = "QA库"
ws_qa.column_dimensions['B'].width = 50
ws_qa.column_dimensions['C'].width = 60

ws_qa.append(["序号", "问题", "回答"])
ws_qa.append([1, "Q：这个锅可以在电磁炉上使用吗？", "A：可以，适用于明火、电磁炉、电陶炉等多种热源。"])
ws_qa.append([2, "Q：产品是否含有害涂层？", "A：不含，采用物理不粘技术，无化学涂层，安全健康。"])
ws_qa.append(["", "", ""])
ws_qa.append([3, "Q：如何清洗和保养？", "A：使用中性洗涤剂手洗，避免使用钢丝球，晾干后存放。"])
ws_qa.append([4, "Q：质保期多久？", "A：提供2年质保，非人为损坏免费换新。"])
ws_qa.append(["", "", ""])
ws_qa.append([5, "Q：可以用洗碗机清洗吗？", "A：建议手洗，洗碗机可能影响表面处理效果。"])

for cell in ws_qa[1]:
    cell.fill = header_fill
    cell.font = header_font

ws_review = wb2.create_sheet("差评应对")
ws_review.column_dimensions['B'].width = 40
ws_review.column_dimensions['C'].width = 60

ws_review.append(["序号", "差评高频词", "应对话术"])
ws_review.append([1, "差评词：生锈", "话术：本品采用食品级不锈钢材质，正常使用不会生锈。建议使用后及时擦干。"])
ws_review.append([2, "差评词：太重", "话术：本品采用加厚材质确保耐用性，我们也提供轻量版供选择。"])
ws_review.append(["", "", ""])
ws_review.append([3, "差评词：粘锅", "话术：使用前请先预热并加入适量油，掌握正确使用方法后不易粘锅。"])
ws_review.append([4, "差评词：掉漆", "话术：外部涂层经过高温固化处理，正常使用不会脱落。"])
ws_review.append(["", "", ""])
ws_review.append([5, "差评词：盖子不密封", "话术：硅胶密封圈可拆卸清洗，如密封不严请检查密封圈是否安装到位。"])

for cell in ws_review[1]:
    cell.fill = header_fill
    cell.font = header_font

l5_path = os.path.join(output_dir, "CW-C01-01_product_knowledge.xlsx")
wb2.save(l5_path)
print(f"L5 test file saved: {l5_path}")

# ============================================================
# 3. L5 with bad filename
# ============================================================
wb3 = openpyxl.Workbook()
ws3 = wb3.active
ws3.title = "QA库"
ws3.append(["序号", "问题", "回答"])
ws3.append([1, "Q：测试问题", "A：测试回答"])
bad_path = os.path.join(output_dir, "bad_filename_no_sku.xlsx")
wb3.save(bad_path)
print(f"Bad filename test file saved: {bad_path}")

print("\nDone! All test files generated.")
